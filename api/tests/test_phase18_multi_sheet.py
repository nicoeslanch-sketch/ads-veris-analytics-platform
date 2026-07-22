"""Regression coverage for intelligent selection, chained analysis and export."""

import io
import json
import os
from pathlib import Path

import openpyxl
import pandas as pd
import pytest

from app.engine.audit import build_audit_dataframe
from app.engine.export import safe_export_dataframe
from app.engine.loader import load_dataframe_with_report, load_dataframes_with_reports
from app.engine.metrics import compute_metrics
from app.engine.multi_sheet import (
    build_analysis_frame,
    detect_relationships,
    join_related_frames,
    validate_analysis_scope,
)
from app.engine.standardize import column_date_profile, parse_date, standardize_dataframe
from app.routes.pipeline import _clean_download_book_sync
from app.routes.pipeline import _analysis_export_type_columns
from app.routes.pipeline import _cache_key
from app.routes.pipeline import _relationships_sync
from app.routes.pipeline import _write_clean_sheet


def _classified_workbook() -> bytes:
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        pd.DataFrame(
            {
                "ID": list(range(1, 9)),
                "Monto": [100, 200, 300, 400, 500, 600, 700, 800],
            }
        ).to_excel(writer, sheet_name="Ventas", index=False)
        pd.DataFrame({"LEEME": ["No modificar"], "Paso": ["Conservar"]}).to_excel(
            writer, sheet_name="LEEME_NO_PROCESAR", index=False
        )
        pd.DataFrame({"Titulo": ["Portada"]}).to_excel(
            writer, sheet_name="Presentacion", index=False
        )
    return output.getvalue()


def test_sheet_classification_is_explainable_and_never_removes_sheets():
    content = _classified_workbook()
    _, report = load_dataframe_with_report("libro.xlsx", content, sheet="Ventas")
    profiles = {item["nombre"]: item for item in report["clasificacion_hojas"]}

    assert list(profiles) == ["Ventas", "LEEME_NO_PROCESAR", "Presentacion"]
    assert profiles["Ventas"]["clasificacion"] == "datos"
    assert profiles["Ventas"]["recomendacion"] == "procesar"
    assert profiles["LEEME_NO_PROCESAR"]["clasificacion"] == "auxiliar"
    assert profiles["LEEME_NO_PROCESAR"]["recomendacion"] == "conservar_sin_procesar"
    assert profiles["Presentacion"]["clasificacion"] == "ambigua"
    assert all(profile["motivos"] for profile in profiles.values())


def test_bulk_sheet_loader_matches_individual_results_and_opens_excel_once(monkeypatch):
    content = _classified_workbook()
    real_excel_file = pd.ExcelFile
    openings = 0

    def counted_excel_file(*args, **kwargs):
        nonlocal openings
        openings += 1
        return real_excel_file(*args, **kwargs)

    monkeypatch.setattr(pd, "ExcelFile", counted_excel_file)
    loaded, available = load_dataframes_with_reports(
        "libro.xlsx", content, ["Ventas", "LEEME_NO_PROCESAR"]
    )

    assert openings == 1
    assert available == ["Ventas", "LEEME_NO_PROCESAR", "Presentacion"]
    for sheet in ("Ventas", "LEEME_NO_PROCESAR"):
        expected, expected_report = load_dataframe_with_report(
            "libro.xlsx", content, sheet=sheet
        )
        actual, actual_report = loaded[sheet]
        pd.testing.assert_frame_equal(actual, expected)
        assert actual.attrs == expected.attrs
        assert actual_report == expected_report


def test_bulk_sheet_loader_audits_formulas_in_every_requested_sheet():
    workbook = openpyxl.Workbook()
    first = workbook.active
    first.title = "Enero"
    first.append(["ID", "Monto"])
    first.append(["A", "=100+50"])
    first.append(["B", 200])
    second = workbook.create_sheet("Febrero")
    second.append(["ID", "Monto"])
    second.append(["C", "=300+50"])
    second.append(["D", 400])
    output = io.BytesIO()
    workbook.save(output)

    loaded, _ = load_dataframes_with_reports(
        "formulas.xlsx", output.getvalue(), ["Enero", "Febrero"]
    )

    for sheet in ("Enero", "Febrero"):
        formula_report = loaded[sheet][1]["formulas"]
        assert formula_report["disponible"] is True
        assert formula_report["total"] == 1
        assert formula_report["por_columna"]["Monto"]["valores_fijos"] == 1


@pytest.mark.parametrize(
    ("value", "dayfirst", "expected"),
    [
        ("01/05/2026", True, "2026-05-01"),
        ("05/22/2026", True, "2026-05-22"),
        ("2026-05-01 00:00:00", True, "2026-05-01"),
        ("1/5/26", True, "2026-05-01"),
        ("02/03/2026", False, "2026-02-03"),
    ],
)
def test_fast_numeric_date_parser_preserves_existing_semantics(value, dayfirst, expected):
    parsed = parse_date(value, dayfirst=dayfirst)
    assert parsed is not None
    assert parsed.strftime("%Y-%m-%d") == expected


def test_append_then_join_derives_cost_without_changing_rows_or_income():
    january = pd.DataFrame(
        {"ID_Producto": ["A", "B"], "Cantidad": [2, 1], "Monto": [500, 300]}
    )
    february = pd.DataFrame(
        {"ID_Producto": ["A", "B"], "Cantidad": [1, 3], "Monto": [250, 900]}
    )
    products = pd.DataFrame(
        {"ID_Producto": ["A", "B"], "Producto": ["Uno", "Dos"], "Categoria": ["X", "Y"], "Costo_Unitario": [100, 200]}
    )
    frames = {"Enero": january, "Febrero": february, "Productos": products}
    mappings = {
        "Enero": {"producto": "ID_Producto", "cantidad": "Cantidad", "monto": "Monto"},
        "Febrero": {"producto": "ID_Producto", "cantidad": "Cantidad", "monto": "Monto"},
        "Productos": {"producto": "Producto", "categoria": "Categoria", "costo": "Costo_Unitario"},
    }
    scope = validate_analysis_scope(
        {
            "mode": "append_join",
            "sheets": ["Enero", "Febrero", "Productos"],
            "append_sheets": ["Enero", "Febrero"],
            "active_sheet": "Enero",
            "join": {
                "left_sheet": "Enero",
                "right_sheet": "Productos",
                "left_keys": ["ID_Producto"],
                "right_keys": ["ID_Producto"],
                "type": "left",
            },
        },
        list(frames),
    )

    joined, mapping, provenance = build_analysis_frame(frames, mappings, scope)

    assert len(joined) == 4
    assert joined["Monto"].sum() == 1950
    assert joined["Costo_Venta"].sum() == 1100
    assert joined["Utilidad_Bruta"].sum() == 850
    assert mapping["costo"] == "Costo_Venta"
    assert provenance["join"]["filas_sin_correspondencia"] == 0
    assert provenance["join"]["costo_derivado"]["cobertura_costos_pct"] == 100.0


def test_append_join_accepts_one_sales_sheet_and_one_catalog():
    sales = pd.DataFrame(
        {"ID_Producto": ["A", "B"], "Cantidad": [2, 1], "Monto": [500, 300]}
    )
    products = pd.DataFrame(
        {"ID_Producto": ["A", "B"], "Costo_Unitario": [100, 200]}
    )
    frames = {"Enero": sales, "Productos": products}
    mappings = {
        "Enero": {"producto": "ID_Producto", "cantidad": "Cantidad", "monto": "Monto"},
        "Productos": {"costo": "Costo_Unitario"},
    }
    scope = validate_analysis_scope(
        {
            "mode": "append_join",
            "sheets": ["Enero", "Productos"],
            "append_sheets": ["Enero"],
            "active_sheet": "Enero",
            "join": {
                "left_sheet": "Enero",
                "right_sheet": "Productos",
                "left_keys": ["ID_Producto"],
                "right_keys": ["ID_Producto"],
                "type": "left",
            },
        },
        list(frames),
    )

    joined, mapping, provenance = build_analysis_frame(frames, mappings, scope)

    assert scope["append_sheets"] == ["Enero"]
    assert len(joined) == 2
    assert joined["Costo_Venta"].sum() == 400
    assert joined["Utilidad_Bruta"].sum() == 400
    assert mapping["costo"] == "Costo_Venta"
    assert provenance["join"]["filas_sin_correspondencia"] == 0


def test_product_catalog_has_unit_statistics_not_fake_cost_sum():
    frame = pd.DataFrame(
        {
            "Producto": ["A", "B", "C"],
            "Categoria": ["X", "X", "Y"],
            "Marca": ["M1", "M2", "M1"],
            "Costo_Unitario": [50, 80, 120],
            "Precio_Lista": [100, 160, 200],
            "Estado": ["Activo", "Activo", "Inactivo"],
        }
    )
    metrics = compute_metrics(
        frame,
        {"producto": "Producto", "categoria": "Categoria", "costo": "Costo_Unitario", "monto": "Precio_Lista"},
    )

    assert metrics["tipo_analisis"] == "catalogo_productos"
    assert metrics["analisis_productos"]["productos"] == 3
    assert metrics["analisis_productos"]["costos"] == {
        "promedio": 83.33,
        "mediana": 80.0,
        "minimo": 50.0,
        "maximo": 120.0,
    }
    assert metrics["kpis"]["gastos_totales"] is None
    assert metrics["kpis"]["ingresos_totales"] is None


def test_product_catalog_with_stock_quantity_stays_a_catalog():
    """Stock is not a sold quantity and must not turn a product master into sales."""
    frame = pd.DataFrame(
        {
            "ID_Producto": ["A", "B"],
            "Producto": ["Uno", "Dos"],
            "Stock": [10, 20],
            "Costo_Unitario": [50, 80],
            "Precio_Lista": [100, 160],
        }
    )
    metrics = compute_metrics(
        frame,
        {
            "producto": "Producto",
            "cantidad": "Stock",
            "costo": "Costo_Unitario",
            "monto": "Precio_Lista",
        },
    )

    assert metrics["tipo_analisis"] == "catalogo_productos"
    assert metrics["analisis_productos"]["totales_catalogo_unitario"] == {
        "costo": 130.0,
        "precio_lista": 260.0,
        "utilidad_potencial": 130.0,
        "productos_con_comparacion": 2,
    }


@pytest.mark.parametrize(
    "metadata_date",
    [
        "Fecha_Actualizacion",
        "Fecha_Creacion",
        "Fecha_Modificacion",
        "Fecha_Vigencia",
    ],
)
def test_product_catalog_metadata_date_is_not_a_sales_date(metadata_date):
    frame = pd.DataFrame(
        {
            "ID_Producto": ["A", "B"],
            "Producto": ["Uno", "Dos"],
            "Stock": [10, 20],
            "Costo_Unitario": [50, 80],
            "Precio_Lista": [100, 160],
            metadata_date: ["01/01/2025", "02/01/2025"],
        }
    )
    metrics = compute_metrics(
        frame,
        {
            "fecha": metadata_date,
            "producto": "Producto",
            "cantidad": "Stock",
            "costo": "Costo_Unitario",
            "monto": "Precio_Lista",
        },
    )

    assert metrics["tipo_analisis"] == "catalogo_productos"
    assert metrics["kpis"]["ingresos_totales"] is None
    assert metrics["analisis_productos"]["productos"] == 2


def test_product_table_with_sales_date_is_transactional():
    frame = pd.DataFrame(
        {
            "Producto": ["A", "B"],
            "Cantidad": [2, 3],
            "Costo_Unitario": [50, 80],
            "Precio_Lista": [100, 160],
            "Fecha_Venta": ["01/01/2025", "02/01/2025"],
        }
    )
    metrics = compute_metrics(
        frame,
        {
            "fecha": "Fecha_Venta",
            "producto": "Producto",
            "cantidad": "Cantidad",
            "costo": "Costo_Unitario",
            "monto": "Precio_Lista",
        },
    )

    assert metrics.get("tipo_analisis") != "catalogo_productos"
    assert metrics["kpis"]["ingresos_totales"]["valor"] == 260.0


def test_stock_product_catalog_is_a_cost_reference_for_sales_relationships():
    sales = pd.DataFrame(
        {
            "ID_Producto": ["A", "A", "B"],
            "Cantidad": [1, 2, 1],
            "Monto": [200, 400, 300],
        }
    )
    products = pd.DataFrame(
        {
            "ID_Producto": ["A", "B"],
            "Producto": ["Uno", "Dos"],
            "Stock": [10, 20],
            "Costo_Unitario": [50, 80],
            "Precio_Lista": [100, 160],
            "Fecha_Actualizacion": ["01/01/2025", "02/01/2025"],
        }
    )
    candidates = detect_relationships(
        {"Ventas": sales, "Productos": products},
        {
            "Ventas": {
                "producto": "ID_Producto",
                "cantidad": "Cantidad",
                "monto": "Monto",
            },
            "Productos": {
                "producto": "Producto",
                "cantidad": "Stock",
                "costo": "Costo_Unitario",
                "monto": "Precio_Lista",
                "fecha": "Fecha_Actualizacion",
            },
        },
    )

    relation = next(
        item
        for item in candidates
        if item["left_sheet"] == "Ventas" and item["right_sheet"] == "Productos"
    )
    assert relation["safe"] is True
    assert relation["purpose"] == "enriquecer_costos"
    assert relation["recommended"] is True


def test_sales_with_transaction_id_and_amount_can_relates_without_date_or_quantity():
    sales = pd.DataFrame(
        {
            "ID_Venta": ["V1", "V2", "V3"],
            "ID_Producto": ["A", "A", "B"],
            "Monto": [200, 400, 300],
        }
    )
    products = pd.DataFrame(
        {
            "ID_Producto": ["A", "B"],
            "Producto": ["Uno", "Dos"],
            "Costo_Unitario": [50, 80],
            "Precio_Lista": [100, 160],
        }
    )
    mappings = {
        "Ventas": {"producto": "ID_Producto", "monto": "Monto"},
        "Productos": {
            "producto": "Producto",
            "costo": "Costo_Unitario",
            "monto": "Precio_Lista",
        },
    }

    candidates = detect_relationships(
        {"Ventas": sales, "Productos": products}, mappings
    )
    relation = next(
        item
        for item in candidates
        if item["left_sheet"] == "Ventas" and item["right_sheet"] == "Productos"
    )
    metrics = compute_metrics(sales, mappings["Ventas"])

    assert relation["safe"] is True
    assert relation["purpose"] == "enriquecer_costos"
    assert relation["recommended"] is True
    assert metrics.get("tipo_analisis") != "generico"
    assert metrics["kpis"]["ingresos_totales"]["valor"] == 900.0


def test_relationships_sync_blocks_clp_sales_with_usd_stock_catalog():
    source = io.BytesIO()
    with pd.ExcelWriter(source, engine="openpyxl") as writer:
        pd.DataFrame(
            {
                "Fecha_Venta": ["01/01/2025", "02/01/2025", "03/01/2025"],
                "ID_Producto": ["A", "A", "B"],
                "Cantidad": [1, 2, 1],
                "Monto": ["CLP 200", "CLP 400", "CLP 300"],
            }
        ).to_excel(writer, sheet_name="Ventas", index=False)
        pd.DataFrame(
            {
                "ID_Producto": ["A", "B"],
                "Producto": ["Uno", "Dos"],
                "Stock": [10, 20],
                "Costo_Unitario": ["USD 2", "USD 3"],
                "Precio_Lista": ["USD 4", "USD 6"],
                "Fecha_Actualizacion": ["01/01/2025", "02/01/2025"],
            }
        ).to_excel(writer, sheet_name="Productos", index=False)
    manifest = {
        "hojas": [
            {
                "nombre": "Ventas",
                "procesar": True,
                "rules": {},
                "mapping": {
                    "fecha": "Fecha_Venta",
                    "producto": "ID_Producto",
                    "cantidad": "Cantidad",
                    "monto": "Monto",
                },
                "scope": {},
                "eliminar_duplicados": False,
            },
            {
                "nombre": "Productos",
                "procesar": True,
                "rules": {},
                "mapping": {
                    "fecha": "Fecha_Actualizacion",
                    "producto": "Producto",
                    "cantidad": "Stock",
                    "costo": "Costo_Unitario",
                    "monto": "Precio_Lista",
                },
                "scope": {},
                "eliminar_duplicados": False,
            },
        ]
    }

    response = _relationships_sync("monedas.xlsx", source.getvalue(), manifest)
    relation = next(
        item
        for item in response["candidates"]
        if item["left_sheet"] == "Ventas" and item["right_sheet"] == "Productos"
    )

    assert relation["purpose"] == "enriquecer_costos"
    assert relation["safe"] is False
    assert relation["recommended"] is False
    assert relation["currency_compatible"] is False
    assert "CLP" in relation["reason"] and "USD" in relation["reason"]


def test_relationship_focus_discovers_neutral_cost_catalog_from_headers():
    source = io.BytesIO()
    with pd.ExcelWriter(source, engine="openpyxl") as writer:
        for sheet, prefix in (("Enero", "E"), ("Febrero", "F")):
            pd.DataFrame(
                {
                    "ID_Venta": [f"{prefix}1", f"{prefix}2"],
                    "ID_Producto": ["A", "B"],
                    "Monto": [200, 300],
                }
            ).to_excel(writer, sheet_name=sheet, index=False)
        pd.DataFrame(
            {
                "ID_Producto": ["A", "B"],
                "Descripcion": ["Uno", "Dos"],
                "Costo_Unitario": [50, 80],
                "Precio_Lista": [100, 160],
            }
        ).to_excel(writer, sheet_name="Catalogo_SKU", index=False)
    manifest = {
        "hojas": [
            {
                "nombre": name,
                "procesar": True,
                "rules": {},
                "mapping": (
                    {"producto": "ID_Producto", "monto": "Monto"}
                    if name != "Catalogo_SKU"
                    else {}
                ),
                "scope": {},
                "eliminar_duplicados": False,
            }
            for name in ("Enero", "Febrero", "Catalogo_SKU")
        ]
    }

    response = _relationships_sync(
        "catalogo-neutro.xlsx",
        source.getvalue(),
        manifest,
        focus={"sheets": ["Enero", "Febrero"]},
    )

    catalog_relations = [
        item
        for item in response["candidates"]
        if item["right_sheet"] == "Catalogo_SKU"
    ]
    assert catalog_relations
    assert all(item["purpose"] == "enriquecer_costos" for item in catalog_relations)
    assert any(item["safe"] and item["recommended"] for item in catalog_relations)


def test_total_cost_reference_is_not_multiplied_or_recommended_as_unit_cost():
    sales = pd.DataFrame(
        {
            "ID_Producto": ["A", "B"],
            "Cantidad": [2, 3],
            "Monto": [500, 900],
        }
    )
    products = pd.DataFrame(
        {
            "ID_Producto": ["A", "B"],
            "Producto": ["Uno", "Dos"],
            "Costo_Total": [1000, 2000],
            "Precio_Lista": [300, 350],
        }
    )
    mappings = {
        "Ventas": {
            "producto": "ID_Producto",
            "cantidad": "Cantidad",
            "monto": "Monto",
        },
        "Productos": {
            "producto": "Producto",
            "costo": "Costo_Total",
            "monto": "Precio_Lista",
        },
    }
    join = {
        "left_sheet": "Ventas",
        "right_sheet": "Productos",
        "left_keys": ["ID_Producto"],
        "right_keys": ["ID_Producto"],
        "type": "left",
    }

    merged, mapping, provenance = join_related_frames(
        {"Ventas": sales, "Productos": products}, mappings, join
    )
    relation = detect_relationships(
        {"Ventas": sales, "Productos": products}, mappings
    )[0]

    assert merged["Costo_Total"].tolist() == [1000, 2000]
    assert "Costo_Venta" not in merged.columns
    assert mapping.get("costo") is None
    assert provenance["costo_derivado"] is None
    assert relation["purpose"] == "enriquecer_referencia"
    assert relation["recommended"] is False


def test_transaction_with_amount_and_price_list_is_not_hidden_as_catalog():
    frame = pd.DataFrame(
        {
            "Producto": ["A", "B"],
            "Cantidad": [2, 3],
            "Monto": [500, 900],
            "Costo_Unitario": [100, 200],
            "Precio_Lista": [300, 350],
        }
    )
    metrics = compute_metrics(
        frame,
        {
            "producto": "Producto",
            "cantidad": "Cantidad",
            "monto": "Monto",
            "costo": "Costo_Unitario",
        },
    )

    assert metrics.get("tipo_analisis") != "catalogo_productos"
    assert metrics["kpis"]["ingresos_totales"]["valor"] == 1400.0
    assert metrics["kpis"]["gastos_totales"]["valor"] == 800.0
    assert metrics["kpis"]["ganancia_neta"]["valor"] == 600.0


def test_single_sales_sheet_multiplies_quantity_by_unit_cost():
    frame = pd.DataFrame(
        {
            "Fecha": ["01/01/2025", "02/01/2025"],
            "ID_Producto": ["A", "B"],
            "Cantidad": [2, 3],
            "Costo_Unitario": [50, 80],
            "Monto": [200, 400],
            "Categoria": ["X", "Y"],
        }
    )
    metrics = compute_metrics(
        frame,
        {
            "fecha": "Fecha",
            "producto": "ID_Producto",
            "cantidad": "Cantidad",
            "costo": "Costo_Unitario",
            "monto": "Monto",
            "categoria": "Categoria",
        },
    )

    assert metrics["calculo_costos"]["origen"] == "cantidad_por_costo_unitario"
    assert metrics["kpis"]["gastos_totales"]["valor"] == 340.0
    assert metrics["kpis"]["ganancia_neta"]["valor"] == 260.0
    assert {row["nombre"]: row["costo"] for row in metrics["por_categoria"]} == {
        "X": 100.0,
        "Y": 240.0,
    }


def test_declared_total_cost_is_not_multiplied_again():
    metrics = compute_metrics(
        pd.DataFrame(
            {
                "Fecha": ["01/01/2025"],
                "Producto": ["A"],
                "Cantidad": [4],
                "Costo_Total": [300],
                "Monto": [500],
            }
        ),
        {
            "fecha": "Fecha",
            "producto": "Producto",
            "cantidad": "Cantidad",
            "costo": "Costo_Total",
            "monto": "Monto",
        },
    )

    assert metrics["calculo_costos"]["origen"] == "columna_costo"
    assert metrics["kpis"]["gastos_totales"]["valor"] == 300.0


def test_english_month_names_and_explicit_percentages_are_normalized():
    assert parse_date("12-Jan-2025").strftime("%Y-%m-%d") == "2025-01-12"
    assert parse_date("03-Aug-25").strftime("%Y-%m-%d") == "2025-08-03"

    standardized, _ = standardize_dataframe(
        pd.DataFrame({"Descuento_Pct": ["20%", "20", "0.2", "110%", "150"]})
    )
    assert standardized["Descuento_Pct"].tolist() == ["0.2", "0.2", "0.2", "1.1", "1.5"]


def test_cost_master_is_recommended_for_sales_relationship():
    sales = pd.DataFrame(
        {
            "Fecha": ["01/01/2026", "02/01/2026", "03/01/2026"],
            "SKU_Producto": ["SKU-1", "SKU-2", "SKU-3"],
            "Cantidad": [2, 1, 3],
            "Venta_CLP": [2000, 1500, 3600],
        }
    )
    costs = pd.DataFrame(
        {
            "SKU_Producto": ["SKU-1", "SKU-2", "SKU-3"],
            "Producto": ["A", "B", "C"],
            "Costo Unitario": [600, 900, 700],
            "Costo Total Unitario": [650, 950, 750],
            "Fecha Vigencia": ["01/01/2026"] * 3,
        }
    )
    mappings = {
        "Ventas": {
            "fecha": "Fecha",
            "producto": "SKU_Producto",
            "cantidad": "Cantidad",
            "monto": "Venta_CLP",
        },
        "Costos_Productos": {
            "producto": "Producto",
            "costo": "Costo Unitario",
            "monto": "Costo Total Unitario",
            "fecha": "Fecha Vigencia",
        },
    }

    candidates = detect_relationships(
        {"Ventas": sales, "Costos_Productos": costs}, mappings
    )

    assert candidates
    assert candidates[0]["left_sheet"] == "Ventas"
    assert candidates[0]["right_sheet"] == "Costos_Productos"
    assert candidates[0]["left_keys"] == ["SKU_Producto"]
    assert candidates[0]["right_keys"] == ["SKU_Producto"]
    assert candidates[0]["safe"] is True
    assert candidates[0]["recommended"] is True
    assert candidates[0]["purpose"] == "enriquecer_costos"

    metrics = compute_metrics(costs, mappings["Costos_Productos"])
    assert metrics["tipo_analisis"] == "catalogo_productos"
    assert metrics["analisis_productos"]["referencia_tipo"] == "costo_total_unitario"
    assert metrics["kpis"]["ingresos_totales"] is None


def test_metrics_honors_canonical_decimal_chosen_by_standardization():
    standardized, _ = standardize_dataframe(
        pd.DataFrame({"Fecha": ["01/01/2025", "02/01/2025"], "Monto": ["1,234", "2,5"]}),
        {"fecha": "Fecha", "monto": "Monto"},
    )
    assert standardized["Monto"].tolist() == ["1.234", "2.5"]
    metrics = compute_metrics(standardized, {"fecha": "Fecha", "monto": "Monto"})
    assert metrics["kpis"]["ingresos_totales"]["valor"] == 3.73


def test_canonical_export_keeps_decimal_dates_and_negative_text_roles():
    exported = safe_export_dataframe(
        pd.DataFrame(
            {
                "Monto": ["1.234"],
                "Fecha": ["12-Jan-2025"],
                "SKU": ["-12.990"],
            }
        ),
        numeric_columns={"Monto"},
        date_columns={"Fecha"},
        canonical_numeric=True,
    )

    assert exported.loc[0, "Monto"] == 1.234
    assert exported.loc[0, "Fecha"].strftime("%Y-%m-%d") == "2025-01-12"
    assert exported.loc[0, "SKU"] == "'-12.990"


def test_xlsx_writer_preserves_numeric_percent_and_date_types():
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)
    _write_clean_sheet(
        workbook,
        "Ventas",
        pd.DataFrame(
            {
                "Fecha": ["12-Jan-2025"],
                "Monto": ["1.234"],
                "Descuento_Pct": ["0.2"],
            }
        ),
        {},
        {},
        numeric_columns={"Monto", "Descuento_Pct"},
        date_columns={"Fecha"},
    )
    sheet = workbook["Ventas"]

    assert sheet["A2"].is_date
    assert sheet["A2"].number_format == "dd/mm/yyyy"
    assert sheet["B2"].value == 1.234
    assert sheet["B2"].data_type == "n"
    assert sheet["C2"].value == 0.2
    assert sheet["C2"].number_format == "0.00%"


def test_numeric_export_types_only_values_that_parse():
    exported = safe_export_dataframe(
        pd.DataFrame({"Cantidad": ["2", "N/D"], "Monto": ["1.500", "1,234"], "Texto": ["001", "=2+2"]}),
        numeric_columns={"Cantidad", "Monto"},
    )

    assert exported.loc[0, "Cantidad"] == 2
    assert exported.loc[0, "Monto"] == 1500
    assert exported.loc[1, "Cantidad"] == "N/D"
    assert exported.loc[1, "Monto"] == "1,234"
    assert exported.loc[0, "Texto"] == "001"
    assert exported.loc[1, "Texto"] == "'=2+2"


def test_related_export_keeps_all_inferred_numeric_and_date_columns_typed():
    frame = pd.DataFrame(
        {
            "Fecha": ["12-Jan-2025"],
            "Monto": ["1.234"],
            "Precio_Unitario": ["2500"],
            "Costo_Unitario_Productos": ["800"],
            "Costo_Venta": [1600.0],
        }
    )
    numeric, dates = _analysis_export_type_columns(
        frame,
        {
            "Ventas": {
                "column_types": {
                    "Fecha": "fecha",
                    "Monto": "numero",
                    "Precio_Unitario": "numero",
                }
            },
            "Productos": {"column_types": {"Costo_Unitario": "numero"}},
        },
        {"fecha": "Fecha", "monto": "Monto", "costo": "Costo_Venta"},
    )
    exported = safe_export_dataframe(
        frame,
        numeric_columns=numeric,
        date_columns=dates,
        canonical_numeric=True,
    )

    assert numeric == {
        "Monto",
        "Precio_Unitario",
        "Costo_Unitario_Productos",
        "Costo_Venta",
    }
    assert dates == {"Fecha"}
    assert exported.loc[0, "Monto"] == 1.234
    assert exported.loc[0, "Precio_Unitario"] == 2500
    assert exported.loc[0, "Costo_Unitario_Productos"] == 800
    assert exported.loc[0, "Fecha"].strftime("%Y-%m-%d") == "2025-01-12"


def test_related_export_does_not_infer_identifier_type_from_numeric_prefix():
    frame = pd.DataFrame(
        {"ID_Producto": ["001"], "ID_Referencia": ["002"]}
    )
    numeric, dates = _analysis_export_type_columns(
        frame,
        {
            "Ventas": {"column_types": {"ID_Producto": "texto"}},
            "Referencia": {"column_types": {"ID": "numero"}},
        },
    )
    exported = safe_export_dataframe(
        frame,
        numeric_columns=numeric,
        date_columns=dates,
        canonical_numeric=True,
    )

    assert numeric == set()
    assert dates == set()
    assert exported.loc[0, "ID_Producto"] == "001"
    assert exported.loc[0, "ID_Referencia"] == "002"


def test_joined_xlsx_keeps_sales_costs_percentages_and_dates_typed_end_to_end():
    source = io.BytesIO()
    with pd.ExcelWriter(source, engine="openpyxl") as writer:
        pd.DataFrame(
            {
                "Fecha": ["12-Jan-2025", "13-Jan-2025"],
                "ID_Producto": ["A", "B"],
                "Cantidad": ["2", "3"],
                "Precio_Unitario": ["2500", "3000"],
                "Descuento_Pct": ["20%", "0.1"],
                "Monto": ["4000", "8100"],
            }
        ).to_excel(writer, sheet_name="Ventas", index=False)
        pd.DataFrame(
            {
                "ID_Producto": ["A", "B"],
                "Producto": ["Uno", "Dos"],
                "Costo_Unitario": ["800", "1000"],
            }
        ).to_excel(writer, sheet_name="Productos", index=False)

    manifest = {
        "hojas": [
            {
                "nombre": "Ventas",
                "procesar": True,
                "rules": {},
                "mapping": {
                    "fecha": "Fecha",
                    "producto": "ID_Producto",
                    "cantidad": "Cantidad",
                    "monto": "Monto",
                },
                "scope": {},
                "eliminar_duplicados": False,
                "status": "limpia",
                "error": "",
            },
            {
                "nombre": "Productos",
                "procesar": True,
                "rules": {},
                "mapping": {
                    "producto": "Producto",
                    "costo": "Costo_Unitario",
                },
                "scope": {},
                "eliminar_duplicados": False,
                "status": "limpia",
                "error": "",
            },
        ]
    }
    scope = {
        "mode": "join",
        "sheets": ["Ventas", "Productos"],
        "active_sheet": "Ventas",
        "join": {
            "left_sheet": "Ventas",
            "right_sheet": "Productos",
            "left_keys": ["ID_Producto"],
            "right_keys": ["ID_Producto"],
            "type": "left",
        },
    }
    payload, _, _ = _clean_download_book_sync(
        "ventas_productos.xlsx", source.getvalue(), manifest, "xlsx", scope
    )
    workbook = openpyxl.load_workbook(io.BytesIO(payload), data_only=False)
    related = workbook["Datos_relacionados"]
    positions = {cell.value: cell.column for cell in related[1]}

    assert related.cell(2, positions["Fecha"]).is_date
    for column in (
        "Cantidad",
        "Precio_Unitario",
        "Descuento_Pct",
        "Monto",
        "Costo_Unitario",
        "Costo_Venta",
        "Utilidad_Bruta",
        "Margen_Bruto",
    ):
        assert related.cell(2, positions[column]).data_type == "n"
    assert related.cell(2, positions["Descuento_Pct"]).value == 0.2
    assert related.cell(2, positions["Descuento_Pct"]).number_format == "0.00%"
    assert related.cell(2, positions["Costo_Venta"]).value == 1600
    assert related.cell(2, positions["Utilidad_Bruta"]).value == 2400
    assert related.cell(2, positions["Margen_Bruto"]).value == 0.6
    assert related.cell(2, positions["Margen_Bruto"]).number_format == "0.00%"


def test_cache_key_includes_dataset_revision_rules_mapping_sheet_and_engine():
    base = _cache_key(
        b"data",
        {"textos": True},
        True,
        {"monto": "Venta"},
        {},
        "Enero",
        False,
        "dataset-1",
        7,
    )
    assert base != _cache_key(
        b"data",
        {"textos": True},
        True,
        {"monto": "Venta"},
        {},
        "Enero",
        False,
        "dataset-1",
        8,
    )
    assert base != _cache_key(
        b"data",
        {"textos": True},
        True,
        {"monto": "Venta"},
        {},
        "Febrero",
        False,
        "dataset-1",
        7,
    )


def test_audit_records_duplicate_even_when_user_keeps_it():
    original = pd.DataFrame({"ID": ["A", "A"], "Monto": ["100", "100"]})
    cleaned = original.copy()
    audit = build_audit_dataframe(
        filename="ventas.csv",
        original=original,
        cleaned=cleaned,
        original_source_rows=[2, 3],
        cleaned_source_rows=[2, 3],
        source_sheet=None,
        column_types={"ID": "texto", "Monto": "numero"},
        column_confidence={"ID": 1.0, "Monto": 1.0},
        mapping={"monto": "Monto"},
        rules={},
        scope=None,
        removed_rows=[],
        detected_duplicate_rows=[{"fila_origen": 3, "motivo": "exacta"}],
        source_sha256="abc",
    )

    kept = audit[audit["accion"] == "duplicado_detectado_y_conservado"]
    assert kept["fila_origen"].tolist() == [3]


def test_single_scope_export_does_not_create_related_sheet():
    content = _classified_workbook()
    names = ["Ventas", "LEEME_NO_PROCESAR", "Presentacion"]
    manifest = {
        "hojas": [
            {
                "nombre": name,
                "procesar": name == "Ventas",
                "rules": {},
                "mapping": {"monto": "Monto"} if name == "Ventas" else {},
                "scope": {},
                "eliminar_duplicados": False,
                "status": "pendiente",
                "error": "",
            }
            for name in names
        ]
    }
    payload, _, _ = _clean_download_book_sync(
        "libro.xlsx",
        content,
        manifest,
        "xlsx",
        {"mode": "single", "sheets": ["Ventas"], "active_sheet": "Ventas"},
    )
    workbook = openpyxl.load_workbook(io.BytesIO(payload), data_only=False)
    assert "Datos_relacionados" not in workbook.sheetnames
    assert workbook["LEEME_NO_PROCESAR"]["A2"].value == "No modificar"


FIXTURE_DIR = Path(__file__).resolve().parent / "fixtures"
_stress_path = os.getenv("ADSVERIS_STRESS_XLSX")
_small_path = os.getenv("ADSVERIS_SMALL_XLSX")
STRESS_PATH = (
    Path(_stress_path)
    if _stress_path
    else FIXTURE_DIR / "Prueba_Estres_Multihoja_ADS_VerIs_2025.xlsx"
)
SMALL_PATH = (
    Path(_small_path)
    if _small_path
    else FIXTURE_DIR / "Prueba_Fase17_Multihoja_ADS_VerIs.xlsx"
)


@pytest.mark.skipif(
    SMALL_PATH is None or not SMALL_PATH.is_file(),
    reason="Define ADSVERIS_SMALL_XLSX para ejecutar la regresion del libro pequeno",
)
def test_small_workbook_exact_regression():
    from app.engine.standardize import parse_number
    from app.routes.pipeline import _analyze_cached

    content = SMALL_PATH.read_bytes()
    names = ["Ventas_Enero", "Ventas_Febrero"]
    before = [
        _analyze_cached(SMALL_PATH.name, content, None, True, sheet=name, eliminar_duplicados=False)
        for name in names
    ]
    after = [
        _analyze_cached(SMALL_PATH.name, content, None, True, sheet=name, eliminar_duplicados=True)
        for name in names
    ]

    def amount(results):
        return sum(
            float(result["_df_limpio"][result["mapeo"]["monto"]].map(parse_number).dropna().sum())
            for result in results
        )

    assert sum(len(result["_df_limpio"]) for result in before) == 24
    assert amount(before) == 5_789_480
    assert sum(len(result["_df_limpio"]) for result in after) == 23
    assert amount(after) == 5_699_510


@pytest.mark.skipif(
    STRESS_PATH is None or not STRESS_PATH.is_file(),
    reason="Define ADSVERIS_STRESS_XLSX para ejecutar la regresion del libro de estres",
)
def test_stress_workbook_verifiable_regressions():
    from app.routes.pipeline import _analyze_cached

    content = STRESS_PATH.read_bytes()
    names = ["Ventas_Ene_Abr_2025", "Ventas_May_Ago_2025", "Ventas_Sep_Dic_2025"]
    before = [
        _analyze_cached(STRESS_PATH.name, content, None, True, sheet=name, eliminar_duplicados=False)
        for name in names
    ]
    results = [
        _analyze_cached(STRESS_PATH.name, content, None, True, sheet=name, eliminar_duplicados=True)
        for name in names
    ]
    assert sum(result["resumen"]["filas_antes"] for result in before) == 5505
    assert sum(result["problemas"]["duplicados"] for result in results) == 75
    assert sum(result["resumen"]["filas_despues"] for result in results) == 5430
    from app.engine.standardize import parse_number
    assert sum(result["_df_limpio"][result["mapeo"]["monto"]].map(parse_number).notna().sum() for result in before) == 5302
    assert sum(result["_df_limpio"][result["mapeo"]["monto"]].map(parse_number).notna().sum() for result in results) == 5233


@pytest.mark.skipif(
    STRESS_PATH is None or not STRESS_PATH.is_file(),
    reason="Define ADSVERIS_STRESS_XLSX para ejecutar la regresion del libro de estres",
)
def test_stress_date_ambiguities_use_full_columns_not_samples():
    content = STRESS_PATH.read_bytes()
    expected = {
        "Ventas_Ene_Abr_2025": 142,
        "Ventas_May_Ago_2025": 133,
        "Ventas_Sep_Dic_2025": 141,
    }

    observed = {}
    for name, expected_count in expected.items():
        frame, _ = load_dataframe_with_report(STRESS_PATH.name, content, sheet=name)
        profile = column_date_profile(frame["Fecha"])
        observed[name] = profile["ambiguas"]
        assert profile["ambiguas"] == expected_count
        assert len(profile["muestras_ambiguas"]) == 5

    assert sum(observed.values()) == 416


@pytest.mark.skipif(
    STRESS_PATH is None or not STRESS_PATH.is_file(),
    reason="Define ADSVERIS_STRESS_XLSX para ejecutar la regresion del libro de estres",
)
def test_stress_append_join_exact_cost_regression():
    from app.routes.pipeline import _analyze_cached

    content = STRESS_PATH.read_bytes()
    sales = ["Ventas_Ene_Abr_2025", "Ventas_May_Ago_2025", "Ventas_Sep_Dic_2025"]
    names = [*sales, "Productos"]
    results = {
        name: _analyze_cached(
            STRESS_PATH.name,
            content,
            None,
            True,
            sheet=name,
            eliminar_duplicados=name in sales,
        )
        for name in names
    }
    joined, mapping, provenance = build_analysis_frame(
        {name: result["_df_limpio"] for name, result in results.items()},
        {name: result["mapeo"] for name, result in results.items()},
        {
            "mode": "append_join",
            "sheets": names,
            "append_sheets": sales,
            "active_sheet": sales[0],
            "join": {
                "left_sheet": sales[0],
                "right_sheet": "Productos",
                "left_keys": ["ID_Producto"],
                "right_keys": ["ID_Producto"],
                "type": "left",
            },
        },
    )

    assert len(joined) == 5430
    # El frame ya usa representación numérica canónica. Reutilizar aquí
    # parse_number con su convención de miles volvería a convertir 1.234 en
    # 1234 y reproduciría precisamente el bug de doble interpretación.
    metrics = compute_metrics(joined, mapping)
    assert metrics["kpis"]["ingresos_totales"]["valor"] == 3_165_894_176
    assert metrics["kpis"]["gastos_totales"]["valor"] == 1_853_487_400
    cost = provenance["join"]["costo_derivado"]
    assert cost["filas_con_costo"] == 5043
    assert cost["cobertura_costos_pct"] == 92.87
    assert provenance["join"]["filas_sin_correspondencia"] == 54


@pytest.mark.skipif(
    STRESS_PATH is None or not STRESS_PATH.is_file(),
    reason="Define ADSVERIS_STRESS_XLSX para ejecutar la regresion del libro de estres",
)
def test_stress_append_join_preserves_reproducible_visible_data_totals():
    from app.routes.pipeline import _analyze_cached

    content = STRESS_PATH.read_bytes()
    sales = ["Ventas_Ene_Abr_2025", "Ventas_May_Ago_2025", "Ventas_Sep_Dic_2025"]
    names = [*sales, "Productos"]
    results = {
        name: _analyze_cached(
            STRESS_PATH.name,
            content,
            None,
            True,
            sheet=name,
            eliminar_duplicados=False,
        )
        for name in names
    }
    joined, mapping, provenance = build_analysis_frame(
        {name: result["_df_limpio"] for name, result in results.items()},
        {name: result["mapeo"] for name, result in results.items()},
        {
            "mode": "append_join",
            "sheets": names,
            "append_sheets": sales,
            "active_sheet": sales[0],
            "join": {
                "left_sheet": sales[0],
                "right_sheet": "Productos",
                "left_keys": ["ID_Producto"],
                "right_keys": ["ID_Producto"],
                "type": "left",
            },
        },
    )

    metrics = compute_metrics(joined, mapping)
    kpis = metrics["kpis"]
    base_costs = kpis["base_costos"]
    assert len(joined) == 5_505
    assert kpis["ingresos_totales"]["valor"] == 3_200_889_420
    assert kpis["gastos_totales"]["valor"] == 1_879_040_000
    assert base_costs == {
        "filas_con_costo": 5_115,
        "costo_total_conocido": 1_879_040_000.0,
        "filas_pareadas": 4_931,
        "costo_pareado": 1_806_140_400.0,
        "ingresos_pareados": 2_967_594_838.0,
    }
    assert kpis["ganancia_neta"]["valor"] == 1_161_454_438
    assert kpis["margen_utilidad_pct"]["valor"] == 39.1
    assert kpis["cobertura_costos"] == {
        "filas_con_ingreso": 5_302,
        "filas_con_ingreso_y_costo": 4_931,
        "pct": 93.0,
    }
    discount_grouping = next(
        item for item in metrics["agrupaciones_flexibles"]
        if item["columna"] == "Descuento_Pct"
    )
    # Desde 0.21.2, 20, 20% y 0.2 representan el mismo descuento de 20%.
    # La expectativa anterior (360) contaba como inválidas 183 filas con
    # enteros porcentuales válidos. Las 177 restantes sí están fuera de
    # 0–100%; incluyen cinco copias exactas que se conservan en este alcance.
    assert discount_grouping["fuera_de_rango"] == {
        "filas": 177,
        "monto_asociado": 127_950_949.0,
    }
    assert "Fuera de rango" in {
        group["nombre"] for group in discount_grouping["grupos"]
    }
    assert provenance["join"]["filas_sin_correspondencia"] == 54


@pytest.mark.skipif(
    STRESS_PATH is None or not STRESS_PATH.is_file(),
    reason="Define ADSVERIS_STRESS_XLSX para ejecutar la regresion del libro de estres",
)
def test_stress_append_join_export_reconciles_scope_totals_and_ambiguities():
    content = STRESS_PATH.read_bytes()
    sales = ["Ventas_Ene_Abr_2025", "Ventas_May_Ago_2025", "Ventas_Sep_Dic_2025"]
    selected = [*sales, "Productos"]
    source_book = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=False)
    all_sheets = list(source_book.sheetnames)
    source_book.close()
    manifest = {
        "hojas": [
            {
                "nombre": name,
                "procesar": name in selected,
                "rules": {},
                "mapping": {},
                "scope": {},
                "eliminar_duplicados": False,
                "revision": 0,
            }
            for name in all_sheets
        ]
    }
    scope = {
        "mode": "append_join",
        "sheets": selected,
        "append_sheets": sales,
        "active_sheet": sales[0],
        "join": {
            "left_sheet": sales[0],
            "right_sheet": "Productos",
            "left_keys": ["ID_Producto"],
            "right_keys": ["ID_Producto"],
            "type": "left",
        },
    }

    payload, _, _ = _clean_download_book_sync(
        STRESS_PATH.name, content, manifest, "xlsx", scope
    )
    workbook = openpyxl.load_workbook(io.BytesIO(payload), data_only=False)
    related = workbook["Datos_relacionados"]
    positions = {cell.value: cell.column for cell in related[1]}
    assert related.max_row - 1 == 5_505
    amount_values = [
        related.cell(row, positions["Monto"]).value
        for row in range(2, related.max_row + 1)
    ]
    numeric_amounts = [value for value in amount_values if isinstance(value, (int, float))]
    assert len(numeric_amounts) == 5_302
    assert sum(numeric_amounts) == 3_200_889_420
    cost_values = [
        related.cell(row, positions["Costo_Venta"]).value
        for row in range(2, related.max_row + 1)
    ]
    utility_values = [
        related.cell(row, positions["Utilidad_Bruta"]).value
        for row in range(2, related.max_row + 1)
    ]
    assert sum(value for value in cost_values if isinstance(value, (int, float))) == 1_879_040_000
    assert sum(value for value in utility_values if isinstance(value, (int, float))) == 1_161_454_438
    assert related.cell(2, positions["Fecha"]).is_date
    for column in ("Costo_Unitario", "Costo_Venta", "Utilidad_Bruta", "Margen_Bruto"):
        assert related.cell(2, positions[column]).data_type == "n"

    manifest_sheet = workbook["Manifest"]
    manifest_columns = {cell.value: cell.column for cell in manifest_sheet[1]}
    exported_scope = json.loads(
        manifest_sheet.cell(2, manifest_columns["alcance_analisis"]).value
    )
    assert exported_scope["mode"] == "append_join"
    assert exported_scope["append_sheets"] == sales
    assert exported_scope["join"]["right_sheet"] == "Productos"

    observations = pd.read_excel(
        io.BytesIO(payload), sheet_name="Observaciones", dtype=str, keep_default_na=False
    )
    summaries = observations[observations["Tipo"] == "ambiguedad_numerica_resumen"]
    counts_by_column = summaries.groupby("Columna")["Detalle"].apply(
        lambda details: sum(int(detail.split(" ", 1)[0]) for detail in details)
    )
    assert counts_by_column.to_dict() == {"Monto": 86, "Precio_Unitario": 98}
    total = observations[observations["Tipo"] == "ambiguedad_numerica_total"]
    assert len(total) == 1
    assert total.iloc[0]["Detalle"].startswith("184 valores ambiguos")

    audit = pd.read_excel(
        io.BytesIO(payload), sheet_name="Auditoria", dtype=str, keep_default_na=False
    )
    audited_ambiguous = audit.apply(
        lambda column: column.astype(str).str.contains("1,234", regex=False)
    ).any(axis=1)
    assert int(audited_ambiguous.sum()) == 184


@pytest.mark.skipif(
    STRESS_PATH is None or not STRESS_PATH.is_file(),
    reason="Define ADSVERIS_STRESS_XLSX para ejecutar la regresion del libro de estres",
)
def test_stress_relationship_allow_and_block_matrix():
    from app.engine.multi_sheet import detect_relationships
    from app.routes.pipeline import _analyze_cached

    content = STRESS_PATH.read_bytes()
    names = [
        "Ventas_Ene_Abr_2025",
        "Productos",
        "Clientes",
        "Sucursales",
        "Inventario_Prod_Sucursal",
        "Promociones",
    ]
    results = {
        name: _analyze_cached(STRESS_PATH.name, content, None, True, sheet=name)
        for name in names
    }
    candidates = detect_relationships(
        {name: result["_df_limpio"] for name, result in results.items()},
        {name: result["mapeo"] for name, result in results.items()},
    )
    safe = {
        (item["left_sheet"], item["right_sheet"], tuple(item["left_keys"]))
        for item in candidates
        if item["safe"]
    }
    assert ("Ventas_Ene_Abr_2025", "Productos", ("ID_Producto",)) in safe
    assert ("Ventas_Ene_Abr_2025", "Clientes", ("ID_Cliente",)) in safe
    assert ("Ventas_Ene_Abr_2025", "Sucursales", ("ID_Sucursal",)) in safe
    assert (
        "Ventas_Ene_Abr_2025",
        "Inventario_Prod_Sucursal",
        ("ID_Producto", "ID_Sucursal"),
    ) in safe
    promotions = [
        item
        for item in candidates
        if {item["left_sheet"], item["right_sheet"]}
        == {"Ventas_Ene_Abr_2025", "Promociones"}
    ]
    assert promotions and not any(item["safe"] for item in promotions)


@pytest.mark.skipif(
    STRESS_PATH is None or not STRESS_PATH.is_file(),
    reason="Define ADSVERIS_STRESS_XLSX para ejecutar la regresion del libro de estres",
)
def test_stress_control_amounts_reconcile_with_duplicate_decision():
    from app.routes.pipeline import _analyze_cached

    content = STRESS_PATH.read_bytes()
    names = ["Ventas_Ene_Abr_2025", "Ventas_May_Ago_2025", "Ventas_Sep_Dic_2025"]
    before_total = 0.0
    after_total = 0.0
    for name in names:
        before = _analyze_cached(STRESS_PATH.name, content, None, True, sheet=name, eliminar_duplicados=False)
        result = _analyze_cached(STRESS_PATH.name, content, None, True, sheet=name, eliminar_duplicados=True)
        before_total += compute_metrics(before["_df_limpio"], before["mapeo"])["kpis"]["ingresos_totales"]["valor"]
        after_total += compute_metrics(result["_df_limpio"], result["mapeo"])["kpis"]["ingresos_totales"]["valor"]
    assert before_total == 3_200_889_420
    assert after_total == 3_165_894_176


@pytest.mark.skipif(
    STRESS_PATH is None or not STRESS_PATH.is_file(),
    reason="Define ADSVERIS_STRESS_XLSX para ejecutar la regresion del libro de estres",
)
def test_stress_declared_control_totals_are_not_present_in_visible_sales_cells():
    """Impide volver a afirmar que CONTROL es reconciliable sin imputar datos.

    CONTROL_ESPERADO conserva totales declarados al generar la base, pero las
    hojas entregadas ya reemplazaron montos por vacíos/N/D. El pipeline solo
    puede sumar los 5.302 valores visibles y legibles: inventar los originales
    ocultos violaría la política no destructiva.
    """
    import openpyxl

    workbook = openpyxl.load_workbook(STRESS_PATH, read_only=True, data_only=True)
    control = workbook["CONTROL_ESPERADO"]
    declared_before = control["B14"].value
    declared_after = control["C14"].value
    workbook.close()

    assert declared_before == 3_278_490_880
    assert declared_after == 3_243_120_100
    assert declared_before - 3_200_889_420 == 77_601_460
    assert declared_after - 3_165_894_176 == 77_225_924
