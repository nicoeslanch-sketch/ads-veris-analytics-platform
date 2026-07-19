"""Fase 18 — verificación con el archivo de estrés multihoja (sintético).

Cubre los hallazgos de la auditoría independiente sobre la Prueba de Estrés:
convención de '1,234' por magnitud, canónicos estables entre hojas, grupos sin
etiqueta 'nan', filtro mensual consistente, perfiles adaptativos enriquecidos,
agrupaciones flexibles y cobertura de negocio en Observaciones.
"""

import io

import pandas as pd
import pytest

from app.engine.metrics import _group_sum, compute_metrics
from app.engine.standardize import column_comma3_convention
from app.routes.pipeline import (
    _clean_download_book_uncached_sync,
    _metrics_sync,
)
from tests.fixture_phase18 import build_stress_book


@pytest.fixture(scope="module")
def stress_book() -> bytes:
    return build_stress_book()


@pytest.fixture(scope="module")
def export_book(stress_book) -> bytes:
    sheets = [
        "Ventas_A", "Ventas_B", "Productos", "Sucursales",
        "Inventario", "Meta_Campanas", "Trabajadores",
    ]
    manifest = {
        "hojas": [
            {
                "nombre": name,
                "procesar": True,
                "rules": {},
                "mapping": {},
                "scope": {},
                "eliminar_duplicados": False,
                "revision": 0,
            }
            for name in sheets
        ]
    }
    data, _, _ = _clean_download_book_uncached_sync(
        "stress18.xlsx", stress_book, manifest, "xlsx", None
    )
    return data


# ── Convención "1,234" por magnitud ──────────────────────────────────────────

def test_comma3_magnitud_decide_miles_en_columna_de_montos():
    serie = pd.Series(["38000", "1,234", "154660", "0", "245600"])
    assert column_comma3_convention(serie) == ("miles", 1)


def test_comma3_decimales_reales_mantienen_decimal():
    serie = pd.Series(["1,5", "2,25", "1,234"])
    assert column_comma3_convention(serie) == ("decimal", 1)


def test_comma3_enteros_chicos_mantienen_decimal():
    serie = pd.Series(["1,234", "2", "5"])
    assert column_comma3_convention(serie) == ("decimal", 1)


def test_metrics_incluye_1234_como_miles_y_avisa(stress_book):
    result = _metrics_sync("x.xlsx", stress_book, None, None, None, sheet="Ventas_A")
    # El fixture suma montos de miles + una celda "1,234": si se leyera como
    # decimal, el total terminaría en ~.234 en vez de entero.
    total = result["kpis"]["ingresos_totales"]["valor"]
    assert total == int(total)
    avisos = [a for a in result.get("advertencias", []) if "1,234" in a]
    assert avisos and "MILES" in avisos[0]


# ── Etiquetas de grupo sin 'nan' ─────────────────────────────────────────────

def test_group_sum_ausencias_caen_en_sin_clasificar_y_literales_se_conservan():
    # NaN (relación sin correspondencia) y vacío → "Sin clasificar"; un literal
    # textual "null" es un DATO conservado (Fase 16) y mantiene su grupo.
    groups = pd.Series([None, float("nan"), "", "null", "Hogar"])
    amounts = pd.Series([100.0, 200.0, 400.0, 300.0, 500.0])
    rows = _group_sum(groups, amounts, None)
    labels = {row["nombre"] for row in rows}
    assert labels == {"Sin clasificar", "null", "Hogar"}
    assert all(row["nombre"].casefold() != "nan" for row in rows)
    sin_clasificar = next(row for row in rows if row["nombre"] == "Sin clasificar")
    assert sin_clasificar["ingresos"] == 700.0


# ── Filtro mensual consistente ───────────────────────────────────────────────

def test_date_to_mes_cubre_el_mes_completo():
    df = pd.DataFrame({
        "Fecha": ["05/01/2025", "20/01/2025", "28/02/2025"],
        "Monto": ["1000", "2000", "4000"],
    })
    mapping = {"fecha": "Fecha", "monto": "Monto"}
    con_mes = compute_metrics(df, mapping, date_from="2025-01", date_to="2025-02")
    con_dia = compute_metrics(df, mapping, date_from="2025-01", date_to="2025-02-28")
    assert con_mes["kpis"]["ingresos_totales"]["valor"] == 7000
    assert (
        con_mes["kpis"]["ingresos_totales"]["valor"]
        == con_dia["kpis"]["ingresos_totales"]["valor"]
    )


# ── Canónico estable entre hojas ─────────────────────────────────────────────

def test_tipocliente_canonico_estable_entre_hojas(stress_book):
    from app.routes.pipeline import _analyze_cached

    canonicos: dict[str, set[str]] = {}
    for sheet in ("Ventas_A", "Ventas_B"):
        result = _analyze_cached("x.xlsx", stress_book, {}, apply=True, sheet=sheet)
        values = set(result["_df_limpio"]["TipoCliente"].unique())
        canonicos[sheet] = values
    # Las variantes de mayúsculas deben converger a la MISMA forma Título en
    # ambas hojas (Persona/Empresa), sin importar qué variante era más común.
    assert "Persona" in canonicos["Ventas_A"] and "Empresa" in canonicos["Ventas_A"]
    assert "Persona" in canonicos["Ventas_B"] and "Empresa" in canonicos["Ventas_B"]
    for values in canonicos.values():
        assert not any(v in {"PERSONA", "persona", "empresa", "EMPRESA"} for v in values)


# ── Perfiles adaptativos enriquecidos ────────────────────────────────────────

def test_inventario_reporta_por_sucursal_y_negativos(stress_book):
    result = _metrics_sync("x.xlsx", stress_book, None, None, None, sheet="Inventario")
    analisis = result["analisis_inventario"]
    assert analisis["stocks_negativos"] >= 0
    assert len(analisis["por_sucursal"]) == 3
    assert {"nombre", "stock", "bajo_minimo", "stocks_negativos"} <= set(
        analisis["por_sucursal"][0]
    )


def test_campanas_reporta_por_plataforma_y_ctr_imposible(stress_book):
    result = _metrics_sync("x.xlsx", stress_book, None, None, None, sheet="Meta_Campanas")
    analisis = result["analisis_campanas"]
    assert analisis["clics_sobre_impresiones"] == 1
    plataformas = {item["nombre"]: item for item in analisis["por_plataforma"]}
    assert plataformas["TikTok Ads"]["ctr_pct"] > 100
    assert any("clics que impresiones" in a for a in result["advertencias"])


def test_generico_clasifica_trabajadores_con_distribuciones(stress_book):
    result = _metrics_sync("x.xlsx", stress_book, None, None, None, sheet="Trabajadores")
    generico = result["analisis_generico"]
    assert generico["subtipo"] == "trabajadores"
    columnas_distro = {d["columna"] for d in generico["distribuciones"]}
    assert "Cargo" in columnas_distro
    assert any(n["columna"] == "Sueldo" for n in generico["numericas"])


def test_generico_clasifica_sucursales(stress_book):
    result = _metrics_sync("x.xlsx", stress_book, None, None, None, sheet="Sucursales")
    assert result["analisis_generico"]["subtipo"] == "sucursales"


# ── Agrupaciones flexibles ───────────────────────────────────────────────────

def test_ventas_exponen_agrupaciones_flexibles(stress_book):
    result = _metrics_sync("x.xlsx", stress_book, None, None, None, sheet="Ventas_A")
    flexibles = {item["columna"]: item for item in result["agrupaciones_flexibles"]}
    assert "Region" in flexibles
    grupos = {g["nombre"] for g in flexibles["Region"]["grupos"]}
    assert {"Norte", "Centro", "Sur"} <= grupos
    # Ningún grupo debe llamarse "nan".
    for item in flexibles.values():
        assert all(g["nombre"].casefold() != "nan" for g in item["grupos"])


# ── Observaciones con cobertura de negocio ───────────────────────────────────

def test_observaciones_incluyen_duplicados_conservados_y_ambiguos(export_book):
    obs = pd.read_excel(
        io.BytesIO(export_book), sheet_name="Observaciones", dtype=str,
        keep_default_na=False,
    )
    tipos = set(obs["Tipo"])
    assert "duplicado_conservado" in tipos
    assert obs["Detalle"].str.contains("1,234", regex=False).any()
    # ID_Sucursal es clave foránea en ventas: sus repeticiones NO son conflicto.
    conflictos = obs[obs["Tipo"] == "conflicto_id"]
    assert not (conflictos["Columna"] == "ID_Sucursal").any()


def test_export_hojas_limpias_son_legibles(export_book):
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(export_book))
    ws = wb["Ventas_A"]
    assert ws.freeze_panes == "A2"
    assert ws["A1"].font.bold and ws["A1"].fill.start_color.rgb.endswith("1A3A52")
    assert (ws.column_dimensions["B"].width or 0) >= 10


def test_export_pide_recalculo_de_formulas(export_book):
    import zipfile

    workbook_xml = zipfile.ZipFile(io.BytesIO(export_book)).read("xl/workbook.xml").decode()
    assert "fullCalcOnLoad" in workbook_xml
