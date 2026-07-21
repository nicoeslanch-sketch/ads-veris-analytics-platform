"""Fase 17: contratos multihoja y relaciones sin doble conteo."""

import io
import json
import zipfile

import openpyxl
import pandas as pd
import pytest

from app.engine.mapping import detect_column_roles, detect_columns_extended
from app.engine.multi_sheet import (
    RELATION_MIN_OVERLAP,
    append_compatible_frames,
    detect_relationships,
    join_related_frames,
    relation_stats,
    validate_analysis_scope,
)
from app.routes.pipeline import (
    _clean_download_book_sync,
    _metrics_multi_sync,
    _parse_analysis_scope,
    _validate_restore_state,
    _workbook_for_clean_export,
)


def _workbook_bytes() -> bytes:
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        pd.DataFrame({"Fecha": ["01/01/2026"], "Venta": ["$ 1.000"]}).to_excel(
            writer, sheet_name="Enero", index=False
        )
        pd.DataFrame({"Fecha": ["01/02/2026"], "Venta": ["$ 2.000"]}).to_excel(
            writer, sheet_name="Febrero", index=False
        )
        pd.DataFrame({"Nota": ["NO TOCAR"], "Formula": ["=1+1"]}).to_excel(
            writer, sheet_name="Notas", index=False
        )
    return output.getvalue()


def _manifest(*processed: str) -> dict:
    return {
        "hojas": [
            {
                "nombre": name,
                "procesar": name in processed,
                "rules": {},
                "mapping": {"fecha": "Fecha", "monto": "Venta"} if name != "Notas" else {},
                "scope": {},
                "eliminar_duplicados": False,
                "status": "pendiente",
                "error": "",
            }
            for name in ("Enero", "Febrero", "Notas")
        ]
    }


def test_all_processed_export_starts_from_empty_workbook(monkeypatch):
    def fail_if_source_is_reopened(*_args, **_kwargs):
        raise AssertionError(
            "el libro original no debe reabrirse si todas las hojas se reemplazan"
        )

    monkeypatch.setattr(openpyxl, "load_workbook", fail_if_source_is_reopened)

    workbook = _workbook_for_clean_export(b"not-read", all_sheets_processed=True)

    assert workbook.write_only is True
    assert workbook.sheetnames == []


def test_tipo_cliente_is_not_product_category_or_customer_name():
    columns = ["TipoCliente", "Total Venta"]
    extended = detect_columns_extended(columns)

    assert extended["TipoCliente"].rol == "tipo_cliente"
    assert not extended["TipoCliente"].rol_motor
    mapping = detect_column_roles(columns)
    assert mapping.get("categoria") != "TipoCliente"
    assert mapping.get("cliente") != "TipoCliente"


def test_extended_role_without_safe_equivalence_is_not_forced_into_metrics():
    mapping = detect_column_roles(["Precio Unitario", "SKU"])
    assert "monto" not in mapping


def test_analysis_scope_is_strict_and_normalized():
    assert validate_analysis_scope(
        {"mode": "single", "sheets": ["Ventas"], "active_sheet": "Ventas"},
        ["Ventas", "Productos"],
    ) == {"mode": "single", "sheets": ["Ventas"], "active_sheet": "Ventas"}
    with pytest.raises(ValueError, match="Hojas desconocidas"):
        validate_analysis_scope(
            {"mode": "append", "sheets": ["Ventas", "Otra"]},
            ["Ventas", "Productos"],
        )


def test_private_selection_mode_never_leaves_restore_storage():
    parsed = _parse_analysis_scope(
        json.dumps({
            "mode": "single",
            "sheets": ["Ventas"],
            "active_sheet": "Ventas",
            "_selection_mode": "all",
        }),
        ["Ventas"],
    )

    assert parsed == {
        "mode": "single",
        "sheets": ["Ventas"],
        "active_sheet": "Ventas",
    }


def test_append_compatible_sheets_adds_origin_without_changing_rows():
    frames = {
        "Enero": pd.DataFrame({"Fecha": ["01/01/2026"], "Venta": [1000]}),
        "Febrero": pd.DataFrame({"Fecha": ["01/02/2026"], "Venta": [2000]}),
    }
    mappings = {
        "Enero": {"fecha": "Fecha", "monto": "Venta"},
        "Febrero": {"fecha": "Fecha", "monto": "Venta"},
    }

    combined, mapping, provenance = append_compatible_frames(frames, mappings)

    assert combined["hoja_origen"].tolist() == ["Enero", "Febrero"]
    assert combined["Venta"].sum() == 3000
    assert mapping["monto"] == "Venta"
    assert provenance["rows"] == 2


def test_append_rejects_incompatible_schemas():
    with pytest.raises(ValueError, match="columnas compatibles"):
        append_compatible_frames(
            {
                "Ventas": pd.DataFrame({"Venta": [1]}),
                "Costos": pd.DataFrame({"Costo": [1]}),
            },
            {"Ventas": {"monto": "Venta"}, "Costos": {"costo": "Costo"}},
        )


def test_many_to_one_join_preserves_rows_and_transaction_totals():
    frames = {
        "Ventas": pd.DataFrame(
            {"ID Producto": ["A", "A", "B"], "Venta": [100, 200, 300], "Cantidad": [1, 2, 3]}
        ),
        "Productos": pd.DataFrame(
            {"ID Producto": ["A", "B"], "Producto": ["Uno", "Dos"], "Categoria": ["X", "Y"]}
        ),
    }
    mappings = {
        "Ventas": {"monto": "Venta", "cantidad": "Cantidad"},
        "Productos": {"producto": "Producto", "categoria": "Categoria"},
    }
    join = {
        "left_sheet": "Ventas",
        "right_sheet": "Productos",
        "left_keys": ["ID Producto"],
        "right_keys": ["ID Producto"],
        "type": "left",
    }

    merged, mapping, provenance = join_related_frames(frames, mappings, join)

    assert len(merged) == len(frames["Ventas"])
    assert merged["Venta"].sum() == frames["Ventas"]["Venta"].sum()
    assert merged["Cantidad"].sum() == frames["Ventas"]["Cantidad"].sum()
    assert merged["Categoria"].tolist() == ["X", "X", "Y"]
    assert mapping["monto"] == "Venta"
    assert provenance["cardinality"] == "muchos_a_uno"


def test_many_to_many_relation_is_blocked():
    left = pd.DataFrame({"ID": ["A", "A", "B"]})
    right = pd.DataFrame({"ID": ["A", "A", "B"]})
    stats = relation_stats(left, ["ID"], right, ["ID"])

    assert stats.cardinality == "muchos_a_muchos"
    assert stats.safe is False
    with pytest.raises(ValueError, match="duplicadas"):
        join_related_frames(
            {"Ventas": left, "Productos": right},
            {"Ventas": {}, "Productos": {}},
            {
                "left_sheet": "Ventas",
                "right_sheet": "Productos",
                "left_keys": ["ID"],
                "right_keys": ["ID"],
                "type": "left",
            },
        )


def test_low_overlap_relation_is_not_safe():
    left = pd.DataFrame({"ID": ["A", "B", "C", "D"]})
    right = pd.DataFrame({"ID": ["A", "X", "Y", "Z"]})
    stats = relation_stats(left, ["ID"], right, ["ID"])

    assert stats.overlap < RELATION_MIN_OVERLAP
    assert stats.safe is False
    assert stats.reason == "El solapamiento entre hojas es insuficiente."


def test_composite_key_can_be_safe_when_single_columns_are_not_unique():
    left = pd.DataFrame(
        {"Sucursal": ["N", "N", "S", "S"], "Codigo": ["A", "B", "A", "B"]}
    )
    right = pd.DataFrame(
        {"Sucursal": ["N", "N", "S", "S"], "Codigo": ["A", "B", "A", "B"]}
    )

    stats = relation_stats(left, ["Sucursal", "Codigo"], right, ["Sucursal", "Codigo"])

    assert stats.cardinality == "uno_a_uno"
    assert stats.safe is True


def test_relationship_response_never_contains_cell_values():
    frames = {
        "Ventas": pd.DataFrame(
            {
                "ID Producto": ["SECRETO-1", "SECRETO-2"],
                "Fecha": ["01/01/2025", "02/01/2025"],
                "Cantidad": [1, 2],
                "Monto": [100, 200],
            }
        ),
        "Productos": pd.DataFrame(
            {
                "ID Producto": ["SECRETO-1", "SECRETO-2"],
                "Producto": ["Uno", "Dos"],
                "Costo Unitario": [50, 80],
                "Precio Lista": [120, 140],
            }
        ),
    }
    candidates = detect_relationships(
        frames,
        {
            "Ventas": {
                "fecha": "Fecha",
                "producto": "ID Producto",
                "cantidad": "Cantidad",
                "monto": "Monto",
            },
            "Productos": {
                "producto": "Producto",
                "costo": "Costo Unitario",
                "monto": "Precio Lista",
            },
        },
    )

    assert candidates
    assert "SECRETO" not in str(candidates)


def test_multi_xlsx_preserves_unselected_sheet_and_original_order():
    content = _workbook_bytes()
    payload, name, media = _clean_download_book_sync(
        "ventas.xlsx", content, _manifest("Enero", "Febrero"), "xlsx", None
    )

    assert name.endswith(".xlsx") and "spreadsheetml" in media
    workbook = openpyxl.load_workbook(io.BytesIO(payload), data_only=False)
    assert workbook.sheetnames[:3] == ["Enero", "Febrero", "Notas"]
    assert workbook["Notas"]["A2"].value == "NO TOCAR"
    assert workbook["Notas"]["B2"].value == "=1+1"
    assert {"Observaciones", "Auditoria", "Manifest"} <= set(workbook.sheetnames)


def test_multi_csv_is_zip_with_one_csv_per_processed_sheet():
    payload, name, media = _clean_download_book_sync(
        "ventas.xlsx", _workbook_bytes(), _manifest("Enero", "Febrero"), "csv", None
    )

    assert name.endswith(".zip") and media == "application/zip"
    with zipfile.ZipFile(io.BytesIO(payload)) as archive:
        assert {"Enero_limpio.csv", "Febrero_limpio.csv", "Auditoria.csv", "manifest.json"} <= set(
            archive.namelist()
        )
        manifest = json.loads(archive.read("manifest.json"))
        assert [entry["estado"] for entry in manifest["hojas"]] == [
            "procesada", "procesada", "no_procesada"
        ]


def test_multi_metrics_cache_contract_is_separated_by_analysis_scope():
    content = _workbook_bytes()
    manifest = _manifest("Enero", "Febrero")
    single = _metrics_multi_sync(
        "ventas.xlsx",
        content,
        manifest,
        {"mode": "single", "sheets": ["Enero"], "active_sheet": "Enero"},
        None,
        None,
    )
    appended = _metrics_multi_sync(
        "ventas.xlsx",
        content,
        manifest,
        {"mode": "append", "sheets": ["Enero", "Febrero"], "active_sheet": "Enero"},
        None,
        None,
    )

    assert single["kpis"]["ingresos_totales"]["valor"] == 1000
    assert appended["kpis"]["ingresos_totales"]["valor"] == 3000
    assert appended["analysis_provenance"]["origin_column"] == "hoja_origen"


def test_mixed_currencies_block_combined_analysis():
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        pd.DataFrame({"Fecha": ["01/01/2026"], "Venta": ["CLP 1000"]}).to_excel(
            writer, sheet_name="Chile", index=False
        )
        pd.DataFrame({"Fecha": ["01/01/2026"], "Venta": ["USD 10"]}).to_excel(
            writer, sheet_name="USA", index=False
        )
    manifest = {
        "hojas": [
            {
                "nombre": name,
                "procesar": True,
                "rules": {},
                "mapping": {"fecha": "Fecha", "monto": "Venta"},
                "scope": {},
                "eliminar_duplicados": False,
            }
            for name in ("Chile", "USA")
        ]
    }
    with pytest.raises(Exception, match="monedas incompatibles"):
        _metrics_multi_sync(
            "monedas.xlsx",
            output.getvalue(),
            manifest,
            {"mode": "append", "sheets": ["Chile", "USA"], "active_sheet": "Chile"},
            None,
            None,
        )


def test_multi_export_preserves_textual_null_literals():
    output = io.BytesIO()
    literals = ["None", "none", "nan", "NaT", "NA", "null"]
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        pd.DataFrame({"Categoria": literals, "Venta": range(1, 7)}).to_excel(
            writer, sheet_name="Datos", index=False
        )
        pd.DataFrame({"Nota": ["intacta"]}).to_excel(writer, sheet_name="Notas", index=False)
    manifest = {
        "hojas": [
            {"nombre": "Datos", "procesar": True, "rules": {}, "mapping": {"monto": "Venta"}, "scope": {}, "eliminar_duplicados": False},
            {"nombre": "Notas", "procesar": False, "rules": {}, "mapping": {}, "scope": {}, "eliminar_duplicados": False},
        ]
    }
    payload, _, _ = _clean_download_book_sync(
        "literales.xlsx", output.getvalue(), manifest, "xlsx", None
    )
    workbook = openpyxl.load_workbook(io.BytesIO(payload), data_only=False)
    values = [workbook["Datos"].cell(row=index, column=1).value for index in range(2, 8)]
    assert values == literals


def test_restore_state_keeps_selection_errors_and_confirmed_scope():
    state = _validate_restore_state(json.dumps({
        "active_sheet": "Ventas",
        "available_sheets": ["Ventas", "Productos", "Notas"],
        "excluded_sheets": ["Notas"],
        "selected_sheets": ["Ventas", "Productos"],
        "sheet_errors": {"Productos": "fallo temporal", "Otra": "ignorar"},
        "analysis_scope": {
            "mode": "join",
            "sheets": ["Ventas", "Productos"],
            "active_sheet": "Ventas",
            "join": {
                "left_sheet": "Ventas",
                "right_sheet": "Productos",
                "left_keys": ["ID"],
                "right_keys": ["ID"],
                "type": "left",
            },
        },
    }))

    assert state["selected_sheets"] == ["Ventas", "Productos"]
    assert state["sheet_errors"] == {"Productos": "fallo temporal"}
    assert state["analysis_scope"]["mode"] == "join"


def test_restore_state_preserves_null_analysis_scope_and_custom_selection():
    state = _validate_restore_state(json.dumps({
        "active_sheet": "Ventas",
        "available_sheets": ["Ventas", "Productos"],
        "excluded_sheets": ["Productos"],
        "selected_sheets": ["Ventas"],
        "sheet_errors": {},
        "analysis_scope": None,
        "selection_mode": "custom",
    }))

    assert state["analysis_scope"] is None
    assert state["selection_mode"] == "custom"
