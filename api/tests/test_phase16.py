"""Regresiones de la auditoría crítica Fase 16."""

from __future__ import annotations

import csv
import io
import json
import zipfile
from pathlib import Path

import openpyxl
import pandas as pd
from fastapi.testclient import TestClient

from app.engine.clean import analyze_and_clean
from app.engine.loader import load_dataframe_with_report
from app.engine.metrics import detect_currency
from app.engine.standardize import standardize_dataframe
from app.routes.pipeline import _clean_download_sync, _metrics_sync
from app.main import app
from app.version import ENGINE_VERSION, LATEST_MIGRATION

REPO = Path(__file__).resolve().parents[2]
LITERALS = ["None", "none", "nan", "NaT", "NA", "null"]


def _literal_csv() -> bytes:
    rows = [
        "Categoria;Ventas",
        "None;100",
        "none;NA",
        "nan;300",
        "NaT;400",
        "NA;500",
        "null;600",
        ";700",
        "VacioMonto;",
    ]
    return ("\n".join(rows) + "\n").encode("utf-8")


def _literal_xlsx() -> bytes:
    frame = pd.DataFrame(
        {
            "Categoria": [*LITERALS, None, "VacioMonto"],
            "Ventas": ["100", "NA", "300", "400", "500", "600", "700", None],
        }
    )
    output = io.BytesIO()
    frame.to_excel(output, index=False)
    return output.getvalue()


def _assert_pipeline_preserves(filename: str, content: bytes) -> None:
    loaded, _ = load_dataframe_with_report(filename, content)
    assert loaded["Categoria"].tolist()[:6] == LITERALS

    standardized, report = standardize_dataframe(loaded)
    assert standardized["Categoria"].tolist()[:6] == LITERALS

    result = analyze_and_clean(
        loaded,
        None,
        apply=True,
        standardized=(standardized, report),
    )
    assert result["_df_limpio"]["Categoria"].tolist()[:6] == LITERALS
    assert result["problemas"]["nulos_fisicos"] == 2
    assert result["problemas"]["nulos_semanticos"] == 1
    assert result["reporte_calidad"]["Categoria"]["nulos_semanticos"] == 0


def test_csv_load_standardize_clean_export_conserva_literales():
    content = _literal_csv()
    _assert_pipeline_preserves("literales.csv", content)

    payload, name, media = _clean_download_sync(
        "literales.csv", content, {}, "csv"
    )
    assert name.endswith(".zip") and media == "application/zip"
    with zipfile.ZipFile(io.BytesIO(payload)) as archive:
        assert sorted(archive.namelist()) == [
            "literales_auditoria.csv",
            "literales_limpio.csv",
            "manifest.json",
        ]
        rows = list(
            csv.DictReader(
                io.StringIO(
                    archive.read("literales_limpio.csv").decode("utf-8-sig")
                ),
                delimiter=";",
            )
        )
        assert [row["Categoria"] for row in rows[:6]] == LITERALS
        audit = list(
            csv.DictReader(
                io.StringIO(
                    archive.read("literales_auditoria.csv").decode("utf-8-sig")
                ),
                delimiter=";",
            )
        )
        assert audit
        assert {
            "valor_original",
            "valor_final",
            "regla",
            "accion",
            "confianza",
            "confirmacion",
            "version_motor",
            "metadatos",
        } <= set(audit[0])
        assert {row["version_motor"] for row in audit} == {ENGINE_VERSION}


def test_excel_load_standardize_clean_export_conserva_literales():
    content = _literal_xlsx()
    _assert_pipeline_preserves("literales.xlsx", content)

    payload, name, media = _clean_download_sync(
        "literales.xlsx", content, {}, "xlsx"
    )
    assert name.endswith(".xlsx") and "spreadsheetml" in media
    workbook = openpyxl.load_workbook(io.BytesIO(payload), data_only=False)
    assert "Auditoria" in workbook.sheetnames
    rows = list(workbook["Datos_limpios"].iter_rows(min_row=2, values_only=True))
    assert [row[0] for row in rows[:6]] == LITERALS
    assert all(
        workbook["Datos_limpios"][f"A{row}"].fill.fill_type is None
        for row in range(2, 8)
    )
    audit_headers = [cell.value for cell in workbook["Auditoria"][1]]
    assert "valor_original" in audit_headers and "version_motor" in audit_headers


def test_literales_textuales_cuentan_como_clientes_validos_en_metricas():
    content = (
        "Cliente;Ventas\n"
        "None;100\n"
        "none;200\n"
        "nan;300\n"
        "NaT;400\n"
        "NA;500\n"
        "null;600\n"
        ";700\n"
    ).encode()
    metrics = _metrics_sync("clientes.csv", content, None, None, None)
    assert metrics["clientes"]["unicos"] == 6
    assert len(metrics["clientes"]["top"]) == 5
    assert {row["nombre"] for row in metrics["clientes"]["top"]} <= set(LITERALS)


def test_moneda_detecta_toda_la_columna_y_bloquea_backend_real():
    values = ["$ 100"] * 1_001 + ["USD 200"]
    detection = detect_currency(pd.Series(values))
    assert detection.mixta is True
    assert detection.detectadas == ("CLP", "USD")
    assert detection.conteos["CLP"] == 1_001
    assert detection.conteos["USD"] == 1

    csv_text = "Fecha;Ventas\n" + "\n".join(
        f"01/01/2026;{value}" for value in values
    )
    metrics = _metrics_sync("mixta.csv", csv_text.encode(), None, None, None)
    assert metrics["moneda_mixta"] is True
    assert metrics["moneda_detalle"]["mixta"] is True
    assert metrics["kpis"]["ingresos_totales"] is None
    assert metrics["kpis"]["ticket_promedio"] is None
    assert metrics["evolucion_mensual"] == []
    assert metrics["proyeccion"] is None
    assert metrics["datos_monetarios_disponibles"] is False


def test_moneda_mixta_tambien_se_detecta_entre_ventas_y_costos():
    content = (
        "Fecha;Ventas;Costo\n"
        "01/01/2026;CLP 1000;USD 10\n"
        "02/01/2026;CLP 2000;USD 20\n"
    ).encode()
    metrics = _metrics_sync("costos.csv", content, None, None, None)
    detail = metrics["moneda_detalle"]
    assert detail["mixta"] is True
    assert detail["conteos_por_columna"]["monto"]["CLP"] == 2
    assert detail["conteos_por_columna"]["costo"]["USD"] == 2
    assert metrics["kpis"]["ganancia_neta"] is None
    assert metrics.get("top_productos", []) == []


def test_calidad_dimensiones_antes_y_despues_misma_formula_y_valores_validos():
    frame = pd.DataFrame(
        {
            "Fecha": ["01/01/2026", "01/01/2026", "fecha rota"],
            "Ventas": ["100", "100", "NA"],
            "Producto": ["A", "A", ""],
        }
    )
    result = analyze_and_clean(
        frame,
        None,
        apply=True,
        eliminar_duplicados=True,
        mapping={"fecha": "Fecha", "monto": "Ventas", "producto": "Producto"},
    )
    before = result["resumen"]["calidad_dimensiones_antes"]
    after = result["resumen"]["calidad_dimensiones_despues"]
    assert result["resumen"]["calidad_dimensiones"] == before
    assert before.keys() == after.keys()
    assert after["unicidad"] > before["unicidad"]
    # Tres roles ausentes aportan 0 y los presentes no están 100% válidos.
    assert 0 < before["cobertura_analitica"] < 50


def test_migracion_snapshot_v3_es_atomica_y_sin_fallback_de_escritura():
    sql = (REPO / "supabase" / "migrations" / "0020_restore_state_v3.sql").read_text(
        encoding="utf-8"
    ).lower()
    source = (REPO / "api" / "app" / "restore_cache.py").read_text(encoding="utf-8")
    assert "reserve_restore_snapshot_revision" in sql
    assert "store_restore_snapshot_guarded" in sql
    assert "where excluded.revision > dataset_sheet_snapshots.revision" in sql
    assert "and s.revision = p_revision" in sql
    assert "dataset_sheet_snapshots" in sql
    assert "datasets.restore_snapshot" not in sql
    assert "drop column" not in sql
    assert "delete from public.datasets" not in sql
    assert "httpx.patch" not in source
    assert "restore snapshot omitted because it is too large" not in source.lower()


def test_migracion_0019_no_reclasifica_historial_empresarial():
    sql = (REPO / "supabase" / "migrations" / "0019_contratacion_basico.sql").read_text(
        encoding="utf-8"
    ).lower()
    assert "upgrade_basico" in sql
    assert "update public.addon_requests" not in sql


def test_smoke_rls_staging_exige_fixtures_criticos():
    source = (REPO / "api" / "scripts" / "smoke_rls.py").read_text(encoding="utf-8")
    assert '"STORAGE_PATH_B": STORAGE_PATH_B' in source
    assert '"BILLING_ID_B": BILLING_ID_B' in source
    for table in (
        "datasets",
        "dataset_columns",
        "cleaning_jobs",
        "analyses",
        "activity_log",
        "billing_identities",
        "addon_requests",
    ):
        assert f'"{table}"' in source


def test_origen_cors_rechazado_no_derrumba_la_api(caplog):
    with TestClient(app) as client:
        response = client.get(
            "/health",
            headers={"Origin": "https://origen-no-permitido.example"},
        )
    assert response.status_code == 200
    assert "Origen NO permitido" in caplog.text


def test_identidad_de_despliegue_declara_fase_16():
    with TestClient(app) as client:
        response = client.get("/version")
    assert response.status_code == 200
    assert response.json()["engine_version"] == ENGINE_VERSION == "0.20.0"
    assert response.json()["database_migration"] == LATEST_MIGRATION == "0021"


def test_standardize_reserva_revision_antes_de_leer_archivo(
    client, auth_headers, monkeypatch
):
    from app.routes import pipeline as pipeline_module

    events: list[str] = []

    def reserve(*_args):
        events.append("reserve")
        return 41

    async def read_input(*_args):
        events.append("read")
        return "orden.csv", b"Fecha;Ventas\n01/01/2026;100\n"

    monkeypatch.setattr(pipeline_module, "reserve_restore_snapshot_revision", reserve)
    monkeypatch.setattr(pipeline_module, "_read_input", read_input)
    monkeypatch.setattr(
        pipeline_module,
        "store_restore_snapshot",
        lambda *_args, **_kwargs: True,
    )

    response = client.post(
        "/standardize",
        data={"dataset_id": "00000000-0000-0000-0000-000000000041"},
        headers=auth_headers,
    )
    assert response.status_code == 200
    assert events[:2] == ["reserve", "read"]
