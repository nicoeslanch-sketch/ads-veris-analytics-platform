"""Fase 12, Bloque 1: duplicados no destructivos e identidad de filas."""

import io
import json

import openpyxl
import pandas as pd
import pytest

from app.engine.clean import (
    analyze_and_clean,
    classify_identifier_kind,
    exclude_from_statistical_outliers,
)
from app.engine.loader import load_dataframe_with_report
from app.engine.standardize import _normalize_text_column, _repair_mojibake


def _clean(client, auth_headers, csv: str, **data):
    return client.post(
        "/clean",
        files={"file": ("fase12.csv", csv.encode("utf-8"), "text/csv")},
        data={key: str(value).lower() if isinstance(value, bool) else value for key, value in data.items()},
        headers=auth_headers,
    )


def test_exactos_se_detectan_siempre_y_solo_se_eliminan_con_flag(client, auth_headers):
    csv = "Producto;Ventas\nA;100\nA;100\n"
    safe = _clean(client, auth_headers, csv, apply=True).json()

    assert safe["problemas"]["duplicados"] == 1
    assert safe["correcciones"]["filas_duplicadas_a_eliminar"] == 0
    assert safe["resumen"]["filas_despues"] == 2
    assert safe["duplicados_detalle"] == {
        **safe["duplicados_detalle"],
        "exactos": 1,
        "grupos": 1,
        "filas_involucradas": 2,
        "tamano_maximo_grupo": 2,
        "filas_seleccionadas_para_eliminar": 0,
        "filas_eliminadas": 0,
        "eliminacion_habilitada": False,
    }

    confirmed = _clean(
        client, auth_headers, csv, apply=True, eliminar_duplicados=True
    ).json()
    assert confirmed["resumen"]["filas_despues"] == 1
    assert confirmed["correcciones"]["filas_duplicadas_a_eliminar"] == 1
    assert confirmed["correcciones"]["filas_duplicadas_eliminadas"] == 1
    assert confirmed["opciones_aplicacion"]["eliminar_duplicados"] is True


def test_normalizados_no_son_eliminables_ni_con_flag(client, auth_headers):
    csv = "Cliente;Ventas\nJuan  Perez;100\nJuan Perez;100\n"
    body = _clean(
        client, auth_headers, csv, apply=True, eliminar_duplicados=True
    ).json()

    assert body["problemas"]["duplicados"] == 0
    assert body["problemas"]["duplicados_probables"] == 1
    assert body["resumen"]["filas_despues"] == 2
    assert body["resumen"]["calidad_despues"] == body["resumen"]["calidad_antes"]
    assert body["correcciones"]["filas_duplicadas_eliminadas"] == 0


@pytest.mark.parametrize("header", ["Año", "RUT Cliente", "Teléfono", "Código Producto"])
def test_nombre_de_atributo_no_cambia_el_criterio_de_duplicados(
    client, auth_headers, header
):
    csv = f"{header};Detalle\n100;Mismo\n100;Mismo\n"
    body = _clean(client, auth_headers, csv, apply=True).json()

    assert body["duplicados_criterio"] == "fila_exacta_original_con_confirmacion"
    assert body["problemas"]["duplicados"] == 1
    assert body["resumen"]["filas_despues"] == 2


@pytest.mark.parametrize("header", ["Año", "RUT", "Teléfono", "Código", "SKU"])
def test_exclusion_iqr_es_independiente_de_la_taxonomia_de_duplicados(header):
    values = [str(value) for value in range(10, 19)] + ["99999999"]
    df = pd.DataFrame({header: values})
    result = analyze_and_clean(df, None, False, mapping={"monto": header})

    assert exclude_from_statistical_outliers(header, "monto") is True
    assert result["reporte_calidad"][header].get("outliers", 0) == 0


def test_folio_es_documento_y_no_habilita_borrado(client, auth_headers):
    csv = "Folio;Producto;Ventas\nF-1;A;100\nF-1;A;100\n"
    body = _clean(client, auth_headers, csv, apply=True).json()

    assert classify_identifier_kind("Folio") == "documento"
    assert body["duplicados_detalle"]["identificadores_fila"] == []
    assert body["resumen"]["filas_despues"] == 2


def test_conflicto_de_id_se_reporta_como_conflicto_no_duplicado(client, auth_headers):
    csv = "ID;Producto;Ventas\nMOV-1;A;100\nMOV-1;B;200\n"
    body = _clean(client, auth_headers, csv, apply=True).json()

    assert classify_identifier_kind("ID") == "fila"
    assert body["problemas"]["duplicados"] == 0
    assert body["duplicados_detalle"]["conflictos_id"] == 1
    assert body["duplicados_detalle"]["ejemplos_conflictos_id"][0]["columna"] == "ID"
    assert body["resumen"]["filas_despues"] == 2


def test_grupo_grande_advierte_granularidad_omitida(client, auth_headers):
    csv = "Producto;Ventas\nA;100\nA;100\nA;100\n"
    body = _clean(client, auth_headers, csv).json()

    detail = body["duplicados_detalle"]
    assert detail["exactos"] == 2
    assert detail["tamano_maximo_grupo"] == 3
    assert detail["posible_granularidad_omitida"] is True
    assert any("variable diferenciadora" in aviso for aviso in body["avisos"])


def test_regla_heredada_duplicados_false_no_desactiva_deteccion(client, auth_headers):
    csv = "Producto;Ventas\nA;100\nA;100\n"
    body = _clean(
        client,
        auth_headers,
        csv,
        apply=True,
        rules='{"duplicados": false}',
    ).json()

    assert body["reglas_activas"]["duplicados"] is False
    assert body["problemas"]["duplicados"] == 1
    assert body["correcciones"]["filas_duplicadas_a_eliminar"] == 0
    assert body["resumen"]["filas_despues"] == 2


def test_cache_distingue_decision_de_eliminar(client, auth_headers, monkeypatch):
    from app.routes import pipeline as pipeline_module

    with pipeline_module._CACHE_LOCK:
        pipeline_module._CLEAN_CACHE.clear()
    calls = {"count": 0}
    real = pipeline_module.analyze_and_clean

    def counted(*args, **kwargs):
        calls["count"] += 1
        return real(*args, **kwargs)

    monkeypatch.setattr(pipeline_module, "analyze_and_clean", counted)
    csv = "Producto;Ventas\nCACHE-F12;123\nCACHE-F12;123\n"
    safe = _clean(client, auth_headers, csv, apply=True).json()
    confirmed = _clean(
        client, auth_headers, csv, apply=True, eliminar_duplicados=True
    ).json()
    again = _clean(client, auth_headers, csv, apply=True).json()

    assert calls["count"] == 2
    assert safe["resumen"]["filas_despues"] == again["resumen"]["filas_despues"] == 2
    assert confirmed["resumen"]["filas_despues"] == 1


def test_metricas_respetan_la_misma_decision_de_sesion(client, auth_headers):
    csv = "Fecha;Producto;Ventas\n01/05/2026;A;100\n01/05/2026;A;100\n"
    common = {
        "files": {"file": ("metricas_f12.csv", csv.encode("utf-8"), "text/csv")},
        "headers": auth_headers,
    }
    safe = client.post("/metrics", **common).json()
    confirmed = client.post(
        "/metrics", data={"eliminar_duplicados": "true"}, **common
    ).json()

    assert safe["kpis"]["transacciones"] == 2
    assert safe["kpis"]["ingresos_totales"]["valor"] == 200
    assert confirmed["kpis"]["transacciones"] == 1
    assert confirmed["kpis"]["ingresos_totales"]["valor"] == 100


def test_limpieza_asistida_no_borra_por_instruccion_libre(client, auth_headers):
    csv = "Producto;Ventas\nA;100\nA;100\n"
    response = client.post(
        "/clean/assisted",
        files={"file": ("asistida_f12.csv", csv.encode("utf-8"), "text/csv")},
        data={"instructions": "elimina duplicados"},
        headers=auth_headers,
    )
    assert response.status_code == 200
    body = response.json()
    assert body["resumen"]["filas_despues"] == 2
    assert body["opciones_aplicacion"]["eliminar_duplicados"] is False
    assert any("confirmar" in aviso for aviso in body["dirigida"]["avisos"])


def test_fila_origen_real_llega_a_preview_y_observaciones(client, auth_headers, monkeypatch):
    from app.routes import pipeline as pipeline_module

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Datos"
    ws.append(["REPORTE DE PRUEBA"])
    ws.append(["Producto", "Ventas"])
    ws.append(["A", 100])
    ws.append(["A", 100])
    buf = io.BytesIO()
    wb.save(buf)
    content = buf.getvalue()
    files = {
        "file": (
            "origen_f12.xlsx",
            content,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    }

    cleaned = client.post(
        "/clean",
        files=files,
        data={"apply": "true", "eliminar_duplicados": "true"},
        headers=auth_headers,
    ).json()
    duplicate_issue = next(issue for issue in cleaned["preview"]["issues"] if issue["tipo"] == "duplicado")
    assert duplicate_issue["fila_origen"] == 4

    monkeypatch.setattr(
        pipeline_module,
        "require_capability_for_user",
        lambda user_id, capability, settings: "analista",
    )
    download = client.post(
        "/clean/download",
        files=files,
        data={"fmt": "xlsx", "eliminar_duplicados": "true"},
        headers=auth_headers,
    )
    assert download.status_code == 200
    exported = openpyxl.load_workbook(io.BytesIO(download.content), data_only=False)
    assert exported["Datos_limpios"].max_row == 2
    observations = list(exported["Observaciones"].iter_rows(min_row=2, values_only=True))
    removed = next(row for row in observations if row[3] == "duplicado_eliminado")
    assert removed[0] == 4
    assert removed[1] == "Datos"


def test_respuesta_conserva_campos_anteriores_y_agrega_detalle(client, auth_headers):
    body = _clean(client, auth_headers, "Producto;Ventas\nA;100\nA;100\n").json()
    for field in (
        "resumen", "problemas", "correcciones", "reglas_activas", "preview",
        "estandarizacion", "column_types", "mapeo", "reporte_calidad", "avisos",
        "duplicados_criterio", "fusiones_texto",
    ):
        assert field in body
    assert "duplicados_detalle" in body
    assert "opciones_aplicacion" in body


def test_contador_textual_no_cuenta_dos_veces_la_misma_celda():
    values = pd.Series(["JUAN  PEREZ"] + ["Juan Perez"] * 10)

    unified, detail, _, _, _ = _normalize_text_column(values)

    assert unified.eq("Juan Perez").all()
    assert detail["celdas_con_espacios_normalizados"] == 1
    assert detail["celdas_con_variantes_unificadas"] == 1
    assert detail["celdas_textuales_unicas_modificadas"] == 1
    assert detail["placeholders_detectados"] == 0
    assert detail["mojibake_detectado"] == 0
    assert detail["mojibake_reparado"] == 0


def test_controles_monetarios_e_iqr_son_solo_senalizacion(client, auth_headers):
    amounts = [0, -1, 1, 2, 3, 4, 5, 6, 7, 100]
    csv = "Ventas\n" + "\n".join(map(str, amounts)) + "\n"

    body = _clean(client, auth_headers, csv, apply=True).json()

    assert body["problemas"]["montos_cero"] == 1
    assert body["problemas"]["montos_negativos"] == 1
    assert body["problemas"]["outliers_iqr"] == 1
    assert body["problemas"]["valores_fuera_de_rango"] == 1
    detail = body["reporte_calidad"]["Ventas"]["outliers_iqr"]
    assert detail["total"] == detail["bajo_limite"] + detail["sobre_limite"] == 1
    assert detail["q1"] < detail["q3"]
    assert body["resumen"]["filas_despues"] == len(amounts)
    assert body["correcciones"]["valores_fuera_de_rango_a_revisar"] == 1


def test_respuesta_no_inventa_total_unico_de_categorias(client, auth_headers):
    body = _clean(client, auth_headers, "Ventas;Descripcion\n0;\n0;Informada\n").json()

    assert body["problemas"]["montos_cero"] == 2
    assert body["problemas"]["valores_nulos"] == 1
    assert not {
        "total_problemas", "problemas_totales", "incidencias_totales"
    } & set(body)


def test_placeholder_cliente_se_conserva_y_no_se_unifica_como_nombre():
    values = pd.Series(["Sin Nombre", "SIN NOMBRE", "Juan Pérez"])

    clients, client_detail, _, _, _ = _normalize_text_column(values, role="cliente")
    categories, category_detail, _, _, _ = _normalize_text_column(values, role="categoria")

    assert clients.tolist()[:2] == ["Sin Nombre", "SIN NOMBRE"]
    assert client_detail["placeholders_detectados"] == 2
    assert all(value != "" for value in clients)
    assert category_detail["placeholders_detectados"] == 0
    assert all(value != "" for value in categories)


def test_nulos_fisicos_y_semanticos_se_reportan_separados(client, auth_headers):
    csv = "Razon Social;Ventas\n;10\nSin Nombre;20\nCliente Real;30\n"

    body = _clean(client, auth_headers, csv).json()

    assert body["problemas"]["nulos_fisicos"] == 1
    assert body["problemas"]["nulos_semanticos"] == 1
    assert body["nulos_detalle"]["fisicos"] == 1
    assert body["nulos_detalle"]["semanticos"] == 1
    assert body["preview"]["filas"][1][0] == "Sin Nombre"
    assert any(issue["tipo"] == "nulo_semantico" for issue in body["preview"]["issues"])


def test_cliente_placeholder_no_cuenta_como_dimension_valida(client, auth_headers):
    csv = "Razon Social;Ventas\nSin Nombre;10\nCliente desconocido;20\n"
    response = client.post(
        "/metrics",
        files={"file": ("clientes.csv", csv.encode("utf-8"), "text/csv")},
        headers=auth_headers,
    )

    assert response.status_code == 200
    assert response.json()["dimensiones"]["cliente"] is False


def test_nulo_estructural_se_detecta_sin_modificar():
    df = pd.DataFrame(
        {
            "Categoria": ["A"] * 20 + ["B"] * 20,
            "Detalle": [""] * 20 + [f"Dato {index}" for index in range(20)],
            "Ventas": [str(index + 1) for index in range(40)],
        }
    )

    result = analyze_and_clean(df, None, apply=True)

    patterns = result["nulos_detalle"]["posibles_estructurales"]
    pattern = next(item for item in patterns if item["columna"] == "Detalle")
    assert pattern["agrupado_por"] == "Categoria"
    assert pattern["vacio_en_grupo_pct"] == 100.0
    assert pattern["informado_fuera_pct"] == 100.0
    assert result["_df_limpio"]["Detalle"].iloc[:20].eq("").all()


def test_mojibake_strict_latin1_cp1252_y_ambiguo():
    latin, latin_audit = _repair_mojibake("JosÃ©")
    cp1252, cp_audit = _repair_mojibake("Lâ€™Oréal")
    ambiguous, ambiguous_audit = _repair_mojibake("Texto � incompleto")

    assert latin == "José"
    assert latin_audit and latin_audit["aplicado"] is True
    assert cp1252 == "L’Oréal"
    assert cp_audit and cp_audit["aplicado"] is True
    assert "cp1252" in cp_audit["metodo"]
    assert ambiguous == "Texto � incompleto"
    assert ambiguous_audit and ambiguous_audit["aplicado"] is False


def test_mojibake_expone_auditoria_sin_perder_original(client, auth_headers):
    csv = "Cliente;Ventas\nJosÃ©;10\nTexto � incompleto;20\n"

    body = _clean(client, auth_headers, csv).json()

    assert body["estandarizacion"]["mojibake_detectado"] == 2
    assert body["estandarizacion"]["mojibake_reparado"] == 1
    audits = body["mojibake_auditoria"]
    assert any(item["valor_original"] == "JosÃ©" and item["aplicado"] for item in audits)
    assert any(item["valor_original"] == "Texto � incompleto" and not item["aplicado"] for item in audits)


def test_identidad_nombre_con_varios_ids_usa_normalizacion_del_motor():
    df = pd.DataFrame(
        {
            "Producto": ["Coca-Cola", "COCA COLA", "Otro"],
            "SKU": ["SKU-1", "SKU-2", "SKU-3"],
            "Ventas": ["10", "20", "30"],
        }
    )

    result = analyze_and_clean(df, None, apply=True)
    identity = result["inconsistencias_identidad"]

    assert identity["nombre_con_varios_ids"]["conteo"] == 1
    example = identity["nombre_con_varios_ids"]["ejemplos"][0]
    assert example["cantidad_ids"] == 2
    assert set(example["ids_ejemplo"]) == {"SKU-1", "SKU-2"}
    assert result["_df_limpio"]["SKU"].tolist() == ["SKU-1", "SKU-2", "SKU-3"]


def test_identidad_id_con_varios_nombres_no_fusiona_entidades():
    df = pd.DataFrame(
        {
            "Producto": ["Producto Uno", "Producto Dos"],
            "Codigo Producto": ["P-1", "P-1"],
            "Ventas": ["10", "20"],
        }
    )

    result = analyze_and_clean(df, None, apply=True)
    identity = result["inconsistencias_identidad"]

    assert identity["id_con_varios_nombres"]["conteo"] == 1
    assert identity["id_con_varios_nombres"]["ejemplos"][0]["cantidad_nombres"] == 2
    assert result["_df_limpio"]["Producto"].nunique() == 2


def test_formulas_solo_se_escanean_en_area_real_y_alertan_id_volatil():
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Datos"
    sheet.append(["=NOW()"])
    sheet.append(["ID Producto", "Producto", "Ventas"])
    sheet.append(["P-1", "A", 10])
    sheet.append(["=RANDBETWEEN(1,999)", "B", 20])
    sheet.append(["P-3", "C", 30])
    sheet.append(["Total", "", "=SUM(C3:C5)"])
    buffer = io.BytesIO()
    workbook.save(buffer)

    frame, report = load_dataframe_with_report("formulas.xlsx", buffer.getvalue())

    assert len(frame) == 3
    formulas = report["formulas"]
    assert formulas["total"] == 1
    assert formulas["volatiles"] == 1
    assert formulas["identificadores_volatiles"] == ["ID Producto"]
    detail = formulas["por_columna"]["ID Producto"]
    assert detail["total"] == 1
    assert detail["valores_fijos"] == 2
    assert detail["ejemplos"][0]["fila_origen"] == 4
    assert any("Advertencia fuerte" in warning for warning in report["avisos"])


def _multi_sheet_workbook() -> bytes:
    workbook = openpyxl.Workbook()
    january = workbook.active
    january.title = "Enero"
    january.append(["Fecha", "Producto", "Ventas"])
    january.append(["01/01/2026", "A", 100])
    january.append(["02/01/2026", "B", 200])
    february = workbook.create_sheet("Febrero")
    february.append(["Producto", "Ventas", "Fecha"])
    february.append(["C", 300, "01/02/2026"])
    february.append(["D", 400, "02/02/2026"])
    notes = workbook.create_sheet("Notas")
    notes.append(["Comentario"])
    notes.append(["Hoja informativa que el usuario no procesó"])
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def test_descarga_multihoja_usa_manifiesto_y_combina_solo_con_confirmacion(
    client, auth_headers, monkeypatch
):
    from app.routes import pipeline as pipeline_module

    monkeypatch.setattr(
        pipeline_module,
        "require_capability_for_user",
        lambda user_id, capability, settings: "analista",
    )
    manifest = {
        "hojas": [
            {
                "nombre": "Enero",
                "procesar": True,
                "rules": {},
                "mapping": {},
                "scope": {},
                "eliminar_duplicados": False,
            },
            {
                "nombre": "Febrero",
                "procesar": True,
                "rules": {},
                "mapping": {},
                "scope": {},
                "eliminar_duplicados": False,
            },
            {
                "nombre": "Notas",
                "procesar": False,
                "rules": {},
                "mapping": {},
                "scope": {},
                "eliminar_duplicados": False,
            },
        ]
    }
    response = client.post(
        "/clean/download",
        files={
            "file": (
                "meses.xlsx",
                _multi_sheet_workbook(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        data={
            "fmt": "xlsx",
            "manifest": json.dumps(manifest),
            "combinar_hojas": "true",
        },
        headers=auth_headers,
    )

    assert response.status_code == 200
    exported = openpyxl.load_workbook(io.BytesIO(response.content), data_only=False)
    assert exported.sheetnames == [
        "Enero", "Febrero", "Datos_combinados", "Observaciones", "Auditoria"
    ]
    assert exported["Enero"].max_row == 3
    assert exported["Febrero"].max_row == 3
    combined = list(exported["Datos_combinados"].iter_rows(values_only=True))
    assert combined[0][0] == "hoja_origen"
    assert [row[0] for row in combined[1:]] == ["Enero", "Enero", "Febrero", "Febrero"]
    observations = list(exported["Observaciones"].iter_rows(min_row=2, values_only=True))
    assert any(row[1] == "Notas" and row[3] == "hoja_no_procesada" for row in observations)


def test_descarga_multihoja_rechaza_manifiesto_que_omite_hojas(
    client, auth_headers, monkeypatch
):
    from app.routes import pipeline as pipeline_module

    monkeypatch.setattr(
        pipeline_module,
        "require_capability_for_user",
        lambda user_id, capability, settings: "analista",
    )
    response = client.post(
        "/clean/download",
        files={"file": ("meses.xlsx", _multi_sheet_workbook())},
        data={
            "fmt": "xlsx",
            "manifest": json.dumps(
                {
                    "hojas": [
                        {
                            "nombre": "Enero",
                            "procesar": True,
                            "rules": {},
                            "mapping": {},
                            "scope": {},
                            "eliminar_duplicados": False,
                        }
                    ]
                }
            ),
        },
        headers=auth_headers,
    )

    assert response.status_code == 422
    assert "enumerar todas las hojas" in response.json()["detail"]
