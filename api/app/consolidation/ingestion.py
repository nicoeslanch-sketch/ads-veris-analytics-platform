"""Ingesta acotada para fuentes grandes del dominio de consolidación."""

from __future__ import annotations

import csv
import hashlib
import os
import tempfile
from collections.abc import Callable, Iterator, Sequence, Set
from contextlib import contextmanager
from pathlib import Path
from urllib.parse import quote

import httpx
import pandas as pd
from fastapi import HTTPException, status
from openpyxl import load_workbook

from ..config import Settings, get_settings
from ..storage import normalize_user_storage_path


def sha256_file(path: Path, chunk_bytes: int = 1024 * 1024) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        while chunk := source.read(chunk_bytes):
            digest.update(chunk)
    return digest.hexdigest()


@contextmanager
def storage_source_file(
    storage_path: str,
    user_id: str,
    settings: Settings | None = None,
    directory: Path | None = None,
) -> Iterator[tuple[Path, str]]:
    """Descarga a un temporal generado; nunca acepta una ruta local del cliente."""
    cfg = settings or get_settings()
    normalized = normalize_user_storage_path(storage_path, user_id)
    if not cfg.supabase_url or not cfg.supabase_service_role_key:
        raise HTTPException(503, "Storage no está configurado para consolidación.")
    max_bytes = max(1, cfg.consolidation_max_source_mb) * 1024 * 1024
    suffix = Path(normalized).suffix.lower()
    encoded = "/".join(quote(part, safe="") for part in normalized.split("/"))
    url = f"{cfg.supabase_url.rstrip('/')}/storage/v1/object/{cfg.supabase_storage_bucket}/{encoded}"
    headers = {
        "Authorization": f"Bearer {cfg.supabase_service_role_key}",
        "apikey": cfg.supabase_service_role_key,
    }
    temp_path: Path | None = None
    try:
        with tempfile.NamedTemporaryFile(prefix="source-", suffix=suffix, dir=directory, delete=False) as target:
            temp_path = Path(target.name)
            digest = hashlib.sha256()
            received = 0
            try:
                with httpx.stream("GET", url, headers=headers, timeout=180) as response:
                    if response.status_code == 404:
                        raise HTTPException(404, "La fuente de consolidación no existe.")
                    if response.status_code != 200:
                        raise HTTPException(502, "Storage no pudo entregar la fuente de consolidación.")
                    length = response.headers.get("content-length")
                    if length and int(length) > max_bytes:
                        raise HTTPException(status.HTTP_413_REQUEST_ENTITY_TOO_LARGE, "La fuente supera el límite exclusivo de consolidación.")
                    for chunk in response.iter_bytes(1024 * 1024):
                        received += len(chunk)
                        if received > max_bytes:
                            raise HTTPException(status.HTTP_413_REQUEST_ENTITY_TOO_LARGE, "La fuente supera el límite exclusivo de consolidación.")
                        digest.update(chunk)
                        target.write(chunk)
            except httpx.HTTPError as exc:
                raise HTTPException(502, f"No se pudo descargar la fuente: {exc.__class__.__name__}.") from exc
        yield temp_path, digest.hexdigest()
    finally:
        if temp_path is not None:
            temp_path.unlink(missing_ok=True)


def csv_columns(path: Path, delimiter: str = ";") -> list[str]:
    with path.open("r", encoding="utf-8-sig", newline="") as source:
        return [str(value).strip() for value in next(csv.reader(source, delimiter=delimiter))]


def detect_csv_dialect(path: Path) -> tuple[str, str]:
    """Detecta encoding y separador sin depender del nombre del archivo."""
    raw = path.read_bytes()[:65_536]
    encoding = "utf-8-sig"
    try:
        sample = raw.decode(encoding)
    except UnicodeDecodeError:
        encoding = "latin-1"
        sample = raw.decode(encoding)
    try:
        delimiter = csv.Sniffer().sniff(sample, delimiters=";,\t|").delimiter
    except csv.Error:
        delimiter = ";"
    return encoding, delimiter


def tabular_headers(path: Path) -> dict[str, list[str]]:
    """Devuelve hojas y columnas para configurar una fuente genérica."""
    suffix = path.suffix.lower()
    if suffix == ".csv":
        encoding, delimiter = detect_csv_dialect(path)
        with path.open("r", encoding=encoding, newline="") as source:
            header = [str(value).strip() for value in next(csv.reader(source, delimiter=delimiter))]
        return {"Datos": header}
    if suffix == ".xlsx":
        return workbook_headers(path)
    raise ValueError("Formato no soportado. Usa CSV o XLSX.")


def tabular_structure(path: Path, sample_rows: int = 1000) -> dict[str, object]:
    """Perfila estructura con una muestra acotada, sin devolver valores de filas."""
    suffix = path.suffix.lower()
    if suffix == ".csv":
        encoding, delimiter = detect_csv_dialect(path)
        with path.open("rb") as source:
            raw = source.read(1_048_576)
        line_count = max(1, raw.count(b"\n"))
        average_line_bytes = max(1, len(raw) / line_count)
        approximate_rows = max(0, int(path.stat().st_size / average_line_bytes) - 1)
        sample = pd.read_csv(
            path, sep=delimiter, dtype="string", nrows=max(1, sample_rows),
            keep_default_na=False, na_values=[], encoding=encoding,
        )
        return {
            "kind": "CSV", "sheets": [{
                "name": "Datos", "columns": [str(column) for column in sample.columns],
                "approximate_rows": approximate_rows,
                "sample_rows": len(sample),
                "unique_ratio": {
                    str(column): float(sample[column].astype("string").str.strip().replace("", pd.NA).nunique(dropna=True) / max(1, len(sample)))
                    for column in sample.columns
                },
            }],
        }
    if suffix != ".xlsx":
        raise ValueError("Formato no soportado. Usa CSV o XLSX.")
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheets: list[dict[str, object]] = []
        for sheet in workbook.worksheets:
            rows = sheet.iter_rows(values_only=True)
            header: list[str] = []
            for row in rows:
                header = [str(value).strip() if value is not None else "" for value in row]
                if any(header):
                    break
            samples: list[tuple[object, ...]] = []
            for index, row in enumerate(rows):
                if index >= sample_rows:
                    break
                if any(value is not None and str(value).strip() for value in row):
                    samples.append(tuple(row[:len(header)]))
            frame = pd.DataFrame.from_records(samples, columns=header) if header else pd.DataFrame()
            # max_row is an estimate only; formatted workbooks can report excess rows.
            approximate_rows = max(len(frame), int(sheet.max_row or 1) - 1)
            sheets.append({
                "name": sheet.title, "columns": header,
                "approximate_rows": approximate_rows, "sample_rows": len(frame),
                "unique_ratio": {
                    str(column): float(frame[column].astype("string").str.strip().replace("", pd.NA).nunique(dropna=True) / max(1, len(frame)))
                    for column in frame.columns if column
                },
            })
        return {"kind": "Excel", "sheets": sheets}
    finally:
        workbook.close()


def read_tabular_source(
    path: Path,
    *,
    sheet_name: str | None = None,
    usecols: Sequence[str] | None = None,
    chunk_rows: int | None = None,
    checkpoint: Callable[[], None] | None = None,
    metrics: dict[str, int] | None = None,
) -> pd.DataFrame:
    """Lee una tabla genérica conservando los valores como texto."""
    suffix = path.suffix.lower()
    if suffix == ".csv":
        encoding, delimiter = detect_csv_dialect(path)
        rows = max(1, chunk_rows or get_settings().consolidation_csv_chunk_rows)
        chunks: list[pd.DataFrame] = []
        for chunk in pd.read_csv(
            path,
            sep=delimiter,
            usecols=list(usecols) if usecols else None,
            dtype="string",
            chunksize=rows,
            keep_default_na=False,
            na_values=[],
            encoding=encoding,
        ):
            if checkpoint:
                checkpoint()
            chunks.append(chunk)
        if metrics is not None:
            metrics.update({"rows_read": sum(len(chunk) for chunk in chunks), "chunks": len(chunks)})
        if not chunks:
            return pd.DataFrame(columns=list(usecols or []), dtype="string")
        return pd.concat(chunks, ignore_index=True, copy=False).astype("string")

    if suffix != ".xlsx":
        raise ValueError("Formato no soportado. Usa CSV o XLSX.")
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        selected_sheet = sheet_name or workbook.sheetnames[0]
        if selected_sheet not in workbook.sheetnames:
            raise ValueError(f"No existe la hoja '{selected_sheet}'.")
        rows_iter = workbook[selected_sheet].iter_rows(values_only=True)
        header: list[str] | None = None
        for row in rows_iter:
            candidate = [str(value).strip() if value is not None else "" for value in row]
            if any(candidate):
                header = candidate
                break
        if header is None:
            return pd.DataFrame(dtype="string")
        selected = list(usecols) if usecols else list(header)
        missing = [column for column in selected if column not in header]
        if missing:
            raise ValueError(f"Faltan columnas en '{selected_sheet}': {', '.join(missing)}")
        indexes = [header.index(column) for column in selected]
        records: list[tuple[object, ...]] = []
        rows_read = chunks = 0
        batch = max(1, chunk_rows or get_settings().consolidation_chunk_size)
        for row in rows_iter:
            rows_read += 1
            if (rows_read - 1) % batch == 0:
                chunks += 1
                if checkpoint:
                    checkpoint()
            values = tuple(row[index] if index < len(row) else None for index in indexes)
            if any(value is not None and str(value).strip() for value in values):
                records.append(values)
        if metrics is not None:
            metrics.update({"rows_read": rows_read, "chunks": chunks})
        return pd.DataFrame.from_records(records, columns=selected).astype("string")
    finally:
        workbook.close()


def iter_csv_chunks(
    path: Path,
    usecols: Sequence[str] | None = None,
    *,
    delimiter: str = ";",
    chunk_rows: int | None = None,
    checkpoint: Callable[[], None] | None = None,
) -> Iterator[pd.DataFrame]:
    rows = chunk_rows or get_settings().consolidation_csv_chunk_rows
    for chunk in pd.read_csv(
        path,
        sep=delimiter,
        usecols=list(usecols) if usecols else None,
        dtype="string",
        chunksize=max(1, rows),
        keep_default_na=False,
        na_values=[],
        encoding="utf-8-sig",
    ):
        if checkpoint:
            checkpoint()
        yield chunk


def read_csv_selected(
    path: Path,
    usecols: Sequence[str] | None = None,
    *,
    chunk_rows: int | None = None,
    checkpoint: Callable[[], None] | None = None,
    metrics: dict[str, int] | None = None,
) -> pd.DataFrame:
    chunks = list(iter_csv_chunks(path, usecols, chunk_rows=chunk_rows, checkpoint=checkpoint))
    if metrics is not None:
        metrics.update({"rows_read": sum(len(chunk) for chunk in chunks), "chunks": len(chunks)})
    if not chunks:
        return pd.DataFrame(columns=list(usecols or []), dtype="string")
    return pd.concat(chunks, ignore_index=True, copy=False)


def read_csv_unique_for_ids(
    path: Path,
    usecols: Sequence[str],
    authority_ids: Set[str],
    *,
    chunk_rows: int | None = None,
    checkpoint: Callable[[], None] | None = None,
) -> tuple[pd.DataFrame, dict[str, int]]:
    """Reduce una dimensión a claves únicas mientras lee; nunca concatena el CSV."""
    if "ID_aux" not in usecols:
        raise ValueError("La dimensión requiere ID_aux.")
    columns = list(usecols)
    records: dict[str, tuple[object, ...]] = {}
    ambiguous: set[str] = set()
    rows_read = rows_matched = chunks = 0
    for chunk in iter_csv_chunks(
        path,
        columns,
        chunk_rows=chunk_rows,
        checkpoint=checkpoint,
    ):
        chunks += 1
        rows_read += len(chunk)
        chunk = chunk.loc[:, columns]
        keys = chunk["ID_aux"].astype("string").str.strip()
        scoped = chunk.loc[keys.isin(authority_ids)]
        rows_matched += len(scoped)
        for values in scoped.itertuples(index=False, name=None):
            key = str(values[columns.index("ID_aux")]).strip()
            if key in ambiguous:
                continue
            if key in records:
                records.pop(key, None)
                ambiguous.add(key)
            else:
                records[key] = values
        del scoped, keys, chunk
    frame = pd.DataFrame.from_records(list(records.values()), columns=columns).astype("string")
    return frame, {
        "rows_read": rows_read,
        "rows_matched": rows_matched,
        "rows_retained": len(frame),
        "ambiguous_keys": len(ambiguous),
        "chunks": chunks,
    }


def workbook_headers(path: Path) -> dict[str, list[str]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        result: dict[str, list[str]] = {}
        for sheet in workbook.worksheets:
            header: list[str] = []
            for row in sheet.iter_rows(values_only=True):
                values = [str(value).strip() if value is not None else "" for value in row]
                if any(values):
                    header = values
                    break
            result[sheet.title] = header
        return result
    finally:
        workbook.close()


def read_xlsx_selected(
    path: Path,
    *,
    sheet_name: str,
    usecols: Sequence[str],
    filter_equals: tuple[str, str] | None = None,
    chunk_rows: int = 100_000,
    checkpoint: Callable[[], None] | None = None,
    metrics: dict[str, int] | None = None,
) -> pd.DataFrame:
    """Itera filas reales en modo read-only y materializa solo columnas útiles."""
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"No existe la hoja '{sheet_name}'.")
        rows = workbook[sheet_name].iter_rows(values_only=True)
        header = [str(value).strip() if value is not None else "" for value in next(rows)]
        missing = [column for column in usecols if column not in header]
        if missing:
            raise ValueError(f"Faltan columnas en '{sheet_name}': {', '.join(missing)}")
        indexes = [header.index(column) for column in usecols]
        filter_index = header.index(filter_equals[0]) if filter_equals else None
        records: list[tuple[object, ...]] = []
        rows_read = selected_rows = chunks = 0
        for row in rows:
            rows_read += 1
            if (rows_read - 1) % max(1, chunk_rows) == 0:
                chunks += 1
                if checkpoint:
                    checkpoint()
            if filter_equals and str(row[filter_index]).strip() != filter_equals[1]:
                continue
            values = tuple(row[index] if index < len(row) else None for index in indexes)
            if any(value is not None and str(value).strip() for value in values):
                records.append(values)
                selected_rows += 1
        if metrics is not None:
            metrics.update({"rows_read": rows_read, "rows_matched": selected_rows, "chunks": chunks})
        return pd.DataFrame.from_records(records, columns=list(usecols)).astype("string")
    finally:
        workbook.close()


def safe_local_acceptance_path(path: str | os.PathLike[str]) -> Path:
    """Solo para CLI/tests locales; esta función nunca se expone en la API."""
    resolved = Path(path).resolve(strict=True)
    if resolved.suffix.lower() not in {".csv", ".xlsx"}:
        raise ValueError("Formato de fuente no soportado.")
    return resolved


def upload_consolidation_artifact(
    local_path: Path,
    storage_path: str,
    user_id: str,
    settings: Settings | None = None,
) -> None:
    """Sube un artefacto inmutable; un objeto existente no se reemplaza."""
    cfg = settings or get_settings()
    normalized = normalize_user_storage_path(storage_path, user_id)
    encoded = "/".join(quote(part, safe="") for part in normalized.split("/"))
    url = f"{cfg.supabase_url.rstrip('/')}/storage/v1/object/{cfg.supabase_storage_bucket}/{encoded}"
    headers = {
        "Authorization": f"Bearer {cfg.supabase_service_role_key}",
        "apikey": cfg.supabase_service_role_key,
        "Content-Type": "application/octet-stream",
        "x-upsert": "false",
        "Content-Length": str(local_path.stat().st_size),
    }
    with local_path.open("rb") as content:
        try:
            response = httpx.post(url, headers=headers, content=content, timeout=300)
        except httpx.HTTPError as exc:
            raise HTTPException(502, f"No se pudo guardar el artefacto: {exc.__class__.__name__}.") from exc
    if response.status_code == 409:
        # Reanudación segura: nunca sobrescribe. Solo reutiliza el objeto si su
        # contenido inmutable coincide exactamente con el artefacto local.
        expected = sha256_file(local_path)
        remote = hashlib.sha256()
        verification_headers = {
            "Authorization": f"Bearer {cfg.supabase_service_role_key}",
            "apikey": cfg.supabase_service_role_key,
        }
        try:
            with httpx.stream("GET", url, headers=verification_headers, timeout=300) as existing:
                if existing.status_code != 200:
                    raise HTTPException(409, "El artefacto inmutable ya existe y no pudo verificarse.")
                for chunk in existing.iter_bytes(1024 * 1024):
                    remote.update(chunk)
        except httpx.HTTPError as exc:
            raise HTTPException(502, "No se pudo verificar el artefacto existente.") from exc
        if remote.hexdigest() == expected:
            return
        raise HTTPException(409, "El artefacto inmutable ya existe con contenido diferente.")
    if response.status_code not in {200, 201}:
        raise HTTPException(502, "Storage no pudo guardar el artefacto de consolidación.")
