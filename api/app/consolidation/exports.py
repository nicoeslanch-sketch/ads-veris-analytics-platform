"""Artefactos derivados, separados e inmutables."""

from __future__ import annotations

import json
import hashlib
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook, load_workbook

from .ingestion import read_tabular_source
from .pipeline import PipelineOutput
from .resources import ResourceMonitor


def _safe_cell(value: Any) -> Any:
    if isinstance(value, (dict, list, tuple, set)):
        return json.dumps(value, ensure_ascii=False, sort_keys=True, default=str)
    if value is None or pd.isna(value):
        return None
    if isinstance(value, str) and value.startswith(("=", "+", "-", "@")):
        return "'" + value
    return value


def _hash_cell(digest: Any, value: Any) -> None:
    safe = _safe_cell(value)
    encoded = ("" if safe is None else str(safe)).encode("utf-8")
    digest.update(len(encoded).to_bytes(8, "big"))
    digest.update(encoded)


def write_dataframe(
    path: Path,
    frame: pd.DataFrame,
    *,
    sheet_name: str,
    chunk_size: int = 100_000,
    checkpoint: Any | None = None,
) -> dict[str, Any]:
    if path.exists():
        raise FileExistsError(f"El artefacto ya existe: {path.name}")
    workbook = Workbook(write_only=True)
    sheet = workbook.create_sheet(sheet_name)
    sheet.append(list(frame.columns))
    digest = hashlib.sha256()
    for value in frame.columns:
        _hash_cell(digest, value)
    chunks = 0
    for row_index, row in enumerate(frame.itertuples(index=False, name=None), start=1):
        safe_row = [_safe_cell(value) for value in row]
        sheet.append(safe_row)
        for value in safe_row:
            _hash_cell(digest, value)
        if (row_index - 1) % max(1, chunk_size) == 0:
            chunks += 1
            if checkpoint:
                checkpoint()
    workbook.save(path)
    return {"rows": len(frame), "columns": len(frame.columns), "chunks": chunks, "bytes": path.stat().st_size, "logical_sha256": digest.hexdigest()}


def write_pipeline_artifacts(
    output: PipelineOutput,
    directory: Path,
    *,
    monitor: ResourceMonitor | None = None,
    chunk_size: int = 100_000,
) -> dict[str, Path]:
    directory.mkdir(parents=True, exist_ok=True)
    generic = output.manifest.mapping_version.startswith("general")
    annual_path = directory / ("BASE_CONSOLIDADA.xlsx" if generic else "DEMRE_2026_COMPATIBLE.xlsx")
    audit_path = directory / ("AUDITORIA_CONSOLIDACION.xlsx" if generic else "AUDITORIA_CONSOLIDACION_DEMRE_2026.xlsx")
    manifest_path = directory / "manifest.json"
    if monitor:
        with monitor.stage("export_annual") as stage:
            annual_metrics = write_dataframe(
                annual_path, output.annual, sheet_name="BASE DE DATOS", chunk_size=chunk_size,
                checkpoint=lambda: monitor.checkpoint("export_annual"),
            )
            stage.add(rows_read=annual_metrics["rows"], rows_generated=annual_metrics["rows"], chunks=annual_metrics["chunks"], artifact_bytes=annual_metrics["bytes"])
    else:
        annual_metrics = write_dataframe(annual_path, output.annual, sheet_name="BASE DE DATOS", chunk_size=chunk_size)

    def write_audit() -> tuple[int, int]:
        workbook = Workbook(write_only=True)
        row_count = 0
        for name, records in output.audit_tables.items():
            sheet = workbook.create_sheet(name[:31])
            if not records:
                sheet.append(["sin_hallazgos"])
                row_count += 1
                continue
            columns = list(dict.fromkeys(key for record in records for key in record))
            sheet.append(columns)
            row_count += 1
            for record in records:
                sheet.append([_safe_cell(record.get(column)) for column in columns])
                row_count += 1
        issues = workbook.create_sheet("issues")
        issues.append(["code", "severity", "message", "count", "column"])
        row_count += 1
        for issue in output.manifest.issues:
            issues.append([issue.code, issue.severity.value, issue.message, issue.count, issue.column])
            row_count += 1
        workbook.save(audit_path)
        return row_count, audit_path.stat().st_size

    if monitor:
        with monitor.stage("export_audit") as stage:
            audit_rows, audit_bytes = write_audit()
            stage.add(rows_generated=audit_rows, chunks=1, artifact_bytes=audit_bytes)
    else:
        audit_rows, audit_bytes = write_audit()
    output.manifest.resource_metrics = monitor.snapshot() if monitor else output.manifest.resource_metrics
    output.manifest.resource_metrics["artifacts"] = {
        "annual": annual_metrics,
        "audit": {"rows": audit_rows, "bytes": audit_bytes},
    }
    output.manifest.memory_bytes_estimate = output.manifest.resource_metrics.get("peak_rss_bytes", output.manifest.memory_bytes_estimate)
    manifest_path.write_text(output.manifest.model_dump_json(indent=2), encoding="utf-8")
    return {"annual": annual_path, "audit": audit_path, "manifest": manifest_path}


def annual_shape(path: Path) -> tuple[int, int]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        rows = workbook["BASE DE DATOS"].iter_rows(values_only=True)
        header = next(rows)
        count = sum(1 for row in rows if row and row[0] not in (None, ""))
        return count, len(header)
    finally:
        workbook.close()


def logical_sheet_hash(path: Path, *, sheet_name: str = "BASE DE DATOS") -> str:
    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        digest = hashlib.sha256()
        rows = workbook[sheet_name].iter_rows(values_only=True)
        header = next(rows, ())
        width = len(header)
        for value in header:
            _hash_cell(digest, value)
        for row in rows:
            padded = (*row, *(None for _ in range(max(0, width - len(row)))))
            for value in padded[:width]:
                _hash_cell(digest, value)
        return digest.hexdigest()
    finally:
        workbook.close()


def write_historical_consolidated(
    historical_path: Path,
    annual: pd.DataFrame,
    output_path: Path,
    *,
    sheet_name: str = "BASE DE DATOS",
    cohort: int = 2026,
) -> tuple[Path | None, str | None]:
    """Copia por streaming solo si el contrato histórico es exactamente compatible."""
    source = load_workbook(historical_path, read_only=True, data_only=False)
    try:
        if sheet_name not in source.sheetnames:
            return None, "historical_sheet_missing"
        rows = source[sheet_name].iter_rows(values_only=True)
        header = [str(value).strip() if value is not None else "" for value in next(rows)]
        if header != list(annual.columns):
            return None, "historical_schema_incompatible"
        cohort_index = header.index("cohorte") if "cohorte" in header else -1
        workbook = Workbook(write_only=True)
        target = workbook.create_sheet(sheet_name)
        target.append(header)
        historical_count = 0
        for row in rows:
            if not row or row[0] in (None, ""):
                continue
            if cohort_index >= 0 and str(row[cohort_index]).strip() == str(cohort):
                return None, "historical_already_contains_cohort"
            target.append([_safe_cell(value) for value in row])
            historical_count += 1
        for row in annual.itertuples(index=False, name=None):
            target.append([_safe_cell(value) for value in row])
        if output_path.exists():
            raise FileExistsError(output_path.name)
        workbook.save(output_path)
        return output_path, f"historical_rows_preserved={historical_count}"
    finally:
        source.close()


def write_historical_generic(
    historical_path: Path,
    annual: pd.DataFrame,
    output_path: Path,
    *,
    sheet_name: str | None = None,
) -> tuple[Path | None, str | None]:
    """Apila un histórico CSV/XLSX solo cuando el esquema coincide exactamente."""
    historical = read_tabular_source(historical_path, sheet_name=sheet_name)
    if list(historical.columns) != list(annual.columns):
        return None, "historical_schema_incompatible"
    combined = pd.concat([historical, annual], ignore_index=True, copy=False)
    write_dataframe(output_path, combined, sheet_name="BASE DE DATOS")
    return output_path, f"historical_rows_preserved={len(historical)}"
