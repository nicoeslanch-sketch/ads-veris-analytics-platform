"""Parser genérico y auditable de libros de códigos heterogéneos."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
import re
from typing import Any

from openpyxl import load_workbook
import pandas as pd

from .source_detection import normalize_header


@dataclass(frozen=True)
class CodebookResult:
    mapping: dict[str, str]
    conflicts: dict[str, tuple[str, ...]]
    sheet: str
    code_column: str
    label_column: str


def normalize_code(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.endswith(".0") and text[:-2].isdigit():
        return text[:-2]
    return text


def _find_column(header: list[str], requested: str) -> int:
    normalized = normalize_header(requested)
    for index, value in enumerate(header):
        if normalize_header(value) == normalized:
            return index
    raise ValueError(f"No se encontró la columna declarada '{requested}'.")


def parse_codebook(
    path: Path,
    *,
    sheet_name: str,
    code_column: str,
    label_column: str,
    header_scan_rows: int = 30,
) -> CodebookResult:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"No existe la hoja declarada '{sheet_name}'.")
        sheet = workbook[sheet_name]
        iterator = sheet.iter_rows(values_only=True)
        header: list[str] | None = None
        code_index = label_index = -1
        for _ in range(header_scan_rows):
            row = next(iterator, None)
            if row is None:
                break
            candidate = [str(value).strip() if value is not None else "" for value in row]
            try:
                code_index = _find_column(candidate, code_column)
                label_index = _find_column(candidate, label_column)
                header = candidate
                break
            except ValueError:
                continue
        if header is None:
            raise ValueError("No se encontró una cabecera compatible en el libro.")
        labels_by_code: dict[str, set[str]] = {}
        for row in iterator:
            code = normalize_code(row[code_index] if code_index < len(row) else None)
            label = str(row[label_index]).strip() if label_index < len(row) and row[label_index] is not None else ""
            if code and label:
                labels_by_code.setdefault(code, set()).add(label)
        conflicts = {
            code: tuple(sorted(labels))
            for code, labels in labels_by_code.items()
            if len(labels) > 1
        }
        mapping = {
            code: next(iter(labels))
            for code, labels in labels_by_code.items()
            if len(labels) == 1
        }
        return CodebookResult(mapping, conflicts, sheet_name, code_column, label_column)
    finally:
        workbook.close()


def parse_inline_codebook(
    path: Path,
    *,
    sheet_name: str,
    variable: str,
    variable_column: str = "Variable",
    detail_column: str = "Detalle",
) -> CodebookResult:
    """Lee catálogos escritos debajo de una variable como ``1. Etiqueta``.

    Es el formato real del Libro Matrícula y de varias hojas descriptivas
    DEMRE. La sección termina al encontrar la siguiente variable; las etiquetas
    siempre provienen del libro y nunca del código de la aplicación.
    """
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"No existe la hoja declarada '{sheet_name}'.")
        rows = workbook[sheet_name].iter_rows(values_only=True)
        header = next(rows, None)
        if header is None:
            raise ValueError("El libro de códigos está vacío.")
        names = [str(value).strip() if value is not None else "" for value in header]
        variable_index = _find_column(names, variable_column)
        detail_index = _find_column(names, detail_column)
        requested = normalize_header(variable)
        active = False
        labels_by_code: dict[str, set[str]] = {}
        pattern = re.compile(r"^\s*([^\s.)-]+)\s*[.)-]\s*(.+?)\s*$")
        for row in rows:
            raw_variable = row[variable_index] if variable_index < len(row) else None
            current = normalize_header(raw_variable)
            if current:
                if active and current != requested:
                    break
                active = current == requested
            if not active:
                continue
            detail = str(row[detail_index]).strip() if detail_index < len(row) and row[detail_index] is not None else ""
            match = pattern.match(detail)
            if match:
                labels_by_code.setdefault(normalize_code(match.group(1)), set()).add(match.group(2).strip())
        if not active and not labels_by_code:
            raise ValueError(f"No se encontró la variable declarada '{variable}'.")
        conflicts = {code: tuple(sorted(labels)) for code, labels in labels_by_code.items() if len(labels) > 1}
        mapping = {code: next(iter(labels)) for code, labels in labels_by_code.items() if len(labels) == 1}
        if not mapping and not conflicts:
            raise ValueError(f"La variable '{variable}' no contiene códigos legibles.")
        return CodebookResult(mapping, conflicts, sheet_name, variable_column, detail_column)
    finally:
        workbook.close()


def recode_values(values: list[Any], codebook: CodebookResult) -> tuple[list[str | None], dict[str, int]]:
    recoded: list[str | None] = []
    counts = {"mapped": 0, "unmapped": 0, "conflict": 0, "empty": 0}
    for value in values:
        code = normalize_code(value)
        if not code:
            recoded.append(None)
            counts["empty"] += 1
        elif code in codebook.conflicts:
            recoded.append(None)
            counts["conflict"] += 1
        elif code in codebook.mapping:
            recoded.append(codebook.mapping[code])
            counts["mapped"] += 1
        else:
            recoded.append(None)
            counts["unmapped"] += 1
    return recoded, counts


def recode_series(values: pd.Series, codebook: CodebookResult) -> tuple[pd.Series, dict[str, int]]:
    """Recodifica vectorizado y conserva los ceros iniciales del código."""
    codes = values.astype("string").fillna("").str.strip().str.replace(r"\.0$", "", regex=True)
    empty = codes.eq("")
    conflict = codes.isin(codebook.conflicts)
    mapped = codes.isin(codebook.mapping)
    result = codes.map(codebook.mapping).astype("string")
    result = result.mask(empty | conflict | ~mapped, pd.NA)
    return result, {
        "mapped": int(mapped.sum()),
        "unmapped": int((~empty & ~conflict & ~mapped).sum()),
        "conflict": int(conflict.sum()),
        "empty": int(empty.sum()),
    }
