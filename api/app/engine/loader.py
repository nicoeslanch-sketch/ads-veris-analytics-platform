"""Carga de archivos Excel/CSV hacia DataFrames de solo texto (Fase 7, §5.12).

Todo se lee como string: el motor de estandarización decide después qué es
fecha, número o texto. Celda vacía = string vacío (nunca NaN), para que las
comparaciones y conteos sean deterministas.

Mejoras profesionales Fase 7:
- Excel con varias hojas: se elige la hoja con más datos (antes se leía la
  primera en silencio) y se informa cuáles quedaron fuera.
- Filas de título sobre el encabezado ("REPORTE VENTAS 2026"): se detecta la
  fila real de encabezados y se omiten las de arriba, con aviso.
- CSV: el separador se decide mirando varias líneas (no solo la primera).

Fase 8:
- Filas de totales al FINAL ("Total", "Subtotal", "Total general", "Suma"):
  no son datos — duplicarían los ingresos en las métricas. Se omiten con aviso.

`load_dataframe_with_report` devuelve (df, reporte_de_carga);
`load_dataframe` se mantiene como wrapper compatible.
"""

import csv
import hashlib
import io
import re
import threading
import unicodedata
import zipfile
from collections import OrderedDict

import pandas as pd

SUPPORTED_EXTENSIONS = (".csv", ".xlsx")
MAX_ROWS = 200_000
# Fase 12b §30: límites de superficie total (el caché por celdas y openpyxl
# dimensionan la memoria por celdas, no por filas).
MAX_COLUMNS = 300
MAX_TOTAL_CELLS = 4_000_000
_HEADER_SCAN_ROWS = 10

# Metadatos fuera de las columnas del usuario. El motor los usa para auditar
# una fila con su número real en el archivo, incluso después de quitar títulos,
# filas vacías o totales y de hacer reset_index().
SOURCE_ROWS_ATTR = "source_rows"
SOURCE_SHEET_ATTR = "source_sheet"

# Fase 10 §8.2: un .xlsx es un ZIP — 15 MB comprimidos pueden expandirse a
# cientos de MB (zip bomb) y tumbar el proceso al leerlo con pandas.
_MAX_UNCOMPRESSED_BYTES = 250 * 1024 * 1024
_MAX_COMPRESSION_RATIO = 120


def _guard_xlsx_zip(content: bytes) -> None:
    """Rechaza .xlsx corruptos o con expansión anómala ANTES de cargarlos."""
    try:
        with zipfile.ZipFile(io.BytesIO(content)) as zf:
            total_uncompressed = sum(info.file_size for info in zf.infolist())
    except zipfile.BadZipFile:
        raise UnsupportedFileError(
            "El archivo .xlsx está dañado o no es un Excel válido. "
            "Ábrelo en Excel y guárdalo nuevamente como .xlsx."
        )
    if total_uncompressed > _MAX_UNCOMPRESSED_BYTES:
        raise UnsupportedFileError(
            "El Excel se expande a un tamaño demasiado grande para procesarlo. "
            "Divide la base en archivos más pequeños o expórtala como CSV."
        )
    compressed = max(len(content), 1)
    if total_uncompressed / compressed > _MAX_COMPRESSION_RATIO:
        raise UnsupportedFileError(
            "El archivo tiene una compresión anómala y no se puede procesar de "
            "forma segura. Exporta la base como CSV e inténtalo de nuevo."
        )

# Fila-resumen al final de la planilla: su primera celda con texto es una
# etiqueta de total. Solo se revisan las ÚLTIMAS filas (nunca datos del medio).
_TOTAL_ROW_RE = re.compile(
    # Fase 13: coincidencia EXACTA de la etiqueta — "Total Energies" o "Suma
    # Servicios" son datos (una empresa), no un resumen.
    r"^((sub)?total(es)?( general(es)?)?|suma(s|torias?)?|gran total)$"
)
_MAX_TRAILING_TOTAL_ROWS = 3

_VOLATILE_FORMULA_RE = re.compile(
    r"(?i)(?:ALEATORIO(?:\.ENTRE)?|RAND(?:BETWEEN)?|RANDBETWEEN|AHORA|NOW|HOY|TODAY|INDIRECTO|INDIRECT)\s*\("
)
_FORMULA_XML_TAG_RE = re.compile(rb"<(?:[A-Za-z_][\w.-]*:)?f(?:\s|/?>)")

_AUXILIARY_SHEET_NAME_RE = re.compile(
    r"(?i)(?:^|[_\s-])(guia|gu[ií]a|control|leeme|readme|instrucciones?|"
    r"notas?|diccionario|par[aá]metros?|parametros?|configuraci[oó]n|"
    r"portada|caratula|car[aá]tula)(?:$|[_\s-])"
)
_SHEET_STRUCTURE_CACHE_LOCK = threading.Lock()
_SHEET_STRUCTURE_CACHE: "OrderedDict[str, dict[str, dict]]" = OrderedDict()
_SHEET_STRUCTURE_CACHE_SIZE = 8
_SHEET_CLASSIFICATION_CACHE: "OrderedDict[str, list[dict]]" = OrderedDict()


def _sheet_structure_metadata(content: bytes, sheet_names: list[str]) -> dict[str, dict]:
    """Lee solo metadatos estructurales que pandas no conserva.

    La clasificación es una recomendación, por lo que un fallo de esta pasada
    auxiliar nunca impide cargar el libro. Se acota a libros moderados para no
    convertir la pantalla de selección en otro cuello de botella.
    """
    if len(content) > 8 * 1024 * 1024 or len(sheet_names) > 50:
        return {}
    cache_key = hashlib.sha256(content).hexdigest()
    with _SHEET_STRUCTURE_CACHE_LOCK:
        cached = _SHEET_STRUCTURE_CACHE.get(cache_key)
        if cached is not None:
            _SHEET_STRUCTURE_CACHE.move_to_end(cache_key)
            return {name: dict(value) for name, value in cached.items()}
    try:
        import openpyxl

        workbook = openpyxl.load_workbook(
            io.BytesIO(content), data_only=False, read_only=False
        )
        metadata: dict[str, dict] = {}
        for name in sheet_names:
            worksheet = workbook[name]
            formulas = 0
            scan_rows = min(int(worksheet.max_row or 0), 80)
            scan_cols = min(int(worksheet.max_column or 0), 80)
            for row in worksheet.iter_rows(
                min_row=1, max_row=max(scan_rows, 1), min_col=1, max_col=max(scan_cols, 1)
            ):
                formulas += sum(cell.data_type == "f" for cell in row)
            metadata[name] = {
                "formulas_muestra": formulas,
                "celdas_combinadas": len(worksheet.merged_cells.ranges),
                "filas_estimadas": int(worksheet.max_row or 0),
                "columnas_estimadas": int(worksheet.max_column or 0),
            }
        workbook.close()
        with _SHEET_STRUCTURE_CACHE_LOCK:
            _SHEET_STRUCTURE_CACHE[cache_key] = metadata
            _SHEET_STRUCTURE_CACHE.move_to_end(cache_key)
            while len(_SHEET_STRUCTURE_CACHE) > _SHEET_STRUCTURE_CACHE_SIZE:
                _SHEET_STRUCTURE_CACHE.popitem(last=False)
        return {name: dict(value) for name, value in metadata.items()}
    except Exception:
        return {}


def _classify_sheet_sample(
    name: str, sample: pd.DataFrame, structure: dict | None = None
) -> dict:
    """Clasifica una hoja sin excluirla ni modificarla.

    El resultado incluye evidencia para que la interfaz explique y permita
    cambiar cada recomendación. No se usa para bloquear el procesamiento.
    """
    structure = structure or {}
    cleaned = _clean_string_frame(sample)
    populated_rows = int((cleaned != "").any(axis=1).sum()) if len(cleaned) else 0
    populated_columns = int((cleaned != "").any(axis=0).sum()) if len(cleaned.columns) else 0
    non_empty_cells = int((cleaned != "").sum().sum()) if len(cleaned) else 0
    area = max(populated_rows * max(populated_columns, 1), 1)
    density = non_empty_cells / area
    header_row = _detect_header_row(cleaned) if len(cleaned) else 0
    sample_headers = (
        [str(value).strip() for value in cleaned.iloc[header_row].tolist()]
        if populated_rows
        else []
    )
    data_rows = max(populated_rows - header_row - 1, 0)
    auxiliary_name = bool(_AUXILIARY_SHEET_NAME_RE.search(name.strip()))
    formulas = int(structure.get("formulas_muestra", 0) or 0)
    merged = int(structure.get("celdas_combinadas", 0) or 0)
    reasons: list[str] = []

    if auxiliary_name:
        classification = "auxiliar"
        reasons.append("El nombre indica guía, control, lectura o instrucciones.")
    elif populated_columns >= 2 and data_rows >= 5 and density >= 0.35:
        classification = "datos"
        reasons.append(
            f"Tiene estructura tabular ({data_rows} filas de datos y {populated_columns} columnas en la muestra)."
        )
    else:
        classification = "ambigua"
        reasons.append("La muestra no permite confirmar una tabla de datos con suficiente confianza.")

    if header_row:
        reasons.append(f"Hay {header_row} fila(s) de título antes del encabezado.")
    if merged:
        reasons.append(f"Contiene {merged} rango(s) de celdas combinadas.")
    if formulas:
        reasons.append(f"Contiene al menos {formulas} fórmula(s) en la muestra.")
    if data_rows < 5:
        reasons.append("Tiene pocas filas de datos.")
    if populated_columns < 2:
        reasons.append("No tiene al menos dos columnas pobladas.")

    return {
        "nombre": name,
        "clasificacion": classification,
        "recomendacion": "procesar" if classification == "datos" else "conservar_sin_procesar",
        "motivos": reasons,
        "estructura": {
            "filas_muestra": populated_rows,
            "filas_datos_muestra": data_rows,
            "columnas_muestra": populated_columns,
            "celdas_no_vacias_muestra": non_empty_cells,
            "densidad_muestra": round(density, 3),
            "fila_encabezado": header_row + 1 if populated_rows else None,
            # Permite que consumidores internos preseleccionen candidatas por
            # estructura sin materializar cada hoja completa. Son solo los
            # encabezados de la muestra ya leída para esta clasificación.
            "encabezados_muestra": sample_headers,
            **structure,
        },
    }


def _xlsx_contains_formulas(content: bytes) -> bool:
    """Detecta nodos de formula sin materializar otra vez todo el workbook.

    La busqueda es binaria, acotada por bloques y sobre XML de hojas solamente.
    Si encuentra una formula, la auditoria detallada con openpyxl sigue intacta.
    """
    overlap = 96
    with zipfile.ZipFile(io.BytesIO(content)) as workbook:
        sheet_files = (
            name
            for name in workbook.namelist()
            if name.startswith("xl/worksheets/") and name.endswith(".xml")
        )
        for name in sheet_files:
            tail = b""
            with workbook.open(name) as worksheet:
                while chunk := worksheet.read(256 * 1024):
                    sample = tail + chunk
                    if _FORMULA_XML_TAG_RE.search(sample):
                        return True
                    tail = sample[-overlap:]
    return False


def _strip_accents_lower(text: str) -> str:
    normalized = unicodedata.normalize("NFD", text)
    return "".join(c for c in normalized if unicodedata.category(c) != "Mn").lower()


def _looks_like_summary_row(row_values: list[str]) -> bool:
    """La fila califica como resumen SOLO si, además de la etiqueta, el resto
    de sus celdas pobladas son numéricas o está casi vacía (Fase 12b §10: una
    fila de datos real de "Total Energies" o "Suma Servicios" tiene el resto
    de las columnas con texto y NO debe eliminarse)."""
    populated = [v for v in row_values if v]
    others = populated[1:]  # sin la etiqueta
    if len(others) <= 1:
        return True
    numeric_like = sum(
        1 for v in others if re.fullmatch(r"[-$().,%\d\s]+", v) is not None
    )
    return numeric_like == len(others)


def _drop_trailing_total_rows(df: pd.DataFrame, report: dict) -> pd.DataFrame:
    """Omite hasta 3 filas de totales al final del archivo (Fase 8)."""
    dropped = 0
    while len(df) > 1 and dropped < _MAX_TRAILING_TOTAL_ROWS:
        row_values = [str(v).strip() for v in df.iloc[-1]]
        first_text = next((v for v in row_values if v), "")
        if not _TOTAL_ROW_RE.match(_strip_accents_lower(first_text)):
            break
        if not _looks_like_summary_row(row_values):
            break
        df = df.iloc[:-1]
        dropped += 1
    if dropped:
        report["filas_totales_omitidas"] = dropped
        report["avisos"].append(
            f"Se omitieron {dropped} fila(s) de totales al final del archivo: "
            "son un resumen, no datos, y duplicarían tus indicadores."
        )
    return df


class UnsupportedFileError(ValueError):
    pass


def _count_outside_quotes(line: str, sep: str) -> int:
    """Cuenta el separador IGNORANDO lo entrecomillado (Fase 13): en
    'ACME,"Servicio, instalación",100' la coma interna no es separador."""
    count = 0
    in_quotes = False
    for ch in line:
        if ch == '"':
            in_quotes = not in_quotes
        elif ch == sep and not in_quotes:
            count += 1
    return count


def _detect_separator(sample: str) -> str:
    """Separador más consistente en las primeras líneas con contenido."""
    lines = [line for line in sample.splitlines() if line.strip()][:8]
    if not lines:
        return ","
    scores: dict[str, int] = {}
    for sep in (";", ",", "\t"):
        counts = [_count_outside_quotes(line, sep) for line in lines]
        if counts[0] > 0 and len(set(counts)) == 1:
            # Mismo número de separadores en todas las líneas → muy confiable.
            scores[sep] = counts[0] * 100
        else:
            scores[sep] = counts[0]
    best = max(scores, key=lambda k: scores[k])
    return best if scores[best] > 0 else ","


def _clean_string_frame(df: pd.DataFrame) -> pd.DataFrame:
    # Fase 15: los nulos REALES se detectan ANTES de pasar a texto y solo esas
    # celdas quedan vacías. El replace global anterior ({"nan","NaT","None"} →
    # "") borraba valores que el usuario escribió literalmente en el archivo —
    # una categoría llamada "None" o un texto "nan" son DATOS, no vacíos.
    missing = df.isna()
    df = df.astype(str)
    return df.mask(missing, "")


def _csv_source_rows(text: str, separator: str, expected_rows: int) -> list[int]:
    """Números de registro físicos del CSV, respetando líneas entrecomilladas.

    `csv.reader.line_num` apunta a la última línea física consumida por cada
    registro. Si el parser estándar y pandas discrepan ante un CSV irregular,
    se usa una secuencia conservadora en vez de romper la carga.
    """
    try:
        reader = csv.reader(io.StringIO(text), delimiter=separator)
        rows: list[int] = []
        header_seen = False
        for record in reader:
            if not any(str(value).strip() for value in record):
                continue
            if not header_seen:
                header_seen = True
                continue
            rows.append(reader.line_num)
        if len(rows) == expected_rows:
            return rows
    except (csv.Error, UnicodeError):
        pass
    return list(range(2, expected_rows + 2))


def _detect_header_row(raw: pd.DataFrame) -> int:
    """Fila real de encabezados dentro de las primeras _HEADER_SCAN_ROWS.

    Un encabezado tiene varias celdas con contenido; una fila de título suele
    tener 1–2. Se elige la primera fila con ≥60% de celdas no vacías y al
    menos 2 con contenido. Si ninguna califica, se usa la fila 0 (compatible
    con el comportamiento anterior)."""
    total_cols = len(raw.columns)
    if total_cols <= 1:
        return 0
    limit = min(_HEADER_SCAN_ROWS, len(raw))
    for idx in range(limit):
        values = [str(v).strip() for v in raw.iloc[idx].tolist()]
        non_empty = sum(1 for v in values if v)
        if non_empty >= 2 and non_empty / total_cols >= 0.6:
            return idx
    return 0


def _load_excel(
    content: bytes,
    report: dict,
    sheet: str | None = None,
    *,
    book: pd.ExcelFile | None = None,
) -> pd.DataFrame:
    """Carga una hoja; ``book`` permite reutilizar el parser en libros multihoja."""
    if book is None:
        _guard_xlsx_zip(content)
        book = pd.ExcelFile(io.BytesIO(content))
    sheet_names = list(book.sheet_names)
    report["hojas_disponibles"] = sheet_names

    # La misma muestra que antes se usaba solo para elegir la hoja principal
    # alimenta ahora una recomendación editable para TODAS las hojas. Ninguna
    # clasificación excluye datos por sí sola.
    profile_key = hashlib.sha256(content).hexdigest()
    with _SHEET_STRUCTURE_CACHE_LOCK:
        cached_profiles = _SHEET_CLASSIFICATION_CACHE.get(profile_key)
        if cached_profiles is not None:
            _SHEET_CLASSIFICATION_CACHE.move_to_end(profile_key)
            profiles = [dict(profile) for profile in cached_profiles]
        else:
            profiles = []
    if not profiles:
        samples: dict[str, pd.DataFrame] = {}
        for name in sheet_names:
            samples[name] = _clean_string_frame(
                book.parse(name, header=None, nrows=60, dtype=str, keep_default_na=False)
            )
        structure = _sheet_structure_metadata(content, sheet_names)
        profiles = [
            _classify_sheet_sample(name, samples[name], structure.get(name))
            for name in sheet_names
        ]
        with _SHEET_STRUCTURE_CACHE_LOCK:
            _SHEET_CLASSIFICATION_CACHE[profile_key] = profiles
            _SHEET_CLASSIFICATION_CACHE.move_to_end(profile_key)
            while len(_SHEET_CLASSIFICATION_CACHE) > _SHEET_STRUCTURE_CACHE_SIZE:
                _SHEET_CLASSIFICATION_CACHE.popitem(last=False)
    report["clasificacion_hojas"] = profiles
    profiles_by_name = {profile["nombre"]: profile for profile in profiles}

    # Fase 10 §8.3: el usuario puede elegir la hoja; sin elección, se usa la
    # hoja con más celdas con datos (muestra de 60 filas por hoja).
    if sheet is not None and sheet in sheet_names:
        best_sheet = sheet
        report["hoja_usada"] = best_sheet
        if len(sheet_names) > 1:
            report["avisos"].append(
                f"Se usó la hoja '{best_sheet}' (elegida por ti)."
            )
    elif len(sheet_names) == 1:
        # No hay nada que comparar: evita leer una muestra de la misma hoja
        # antes de cargarla completa.
        best_sheet = sheet_names[0]
        report["hoja_usada"] = best_sheet
    else:
        best_sheet = sheet_names[0]
        best_score = -1
        for name in sheet_names:
            score = int(
                profiles_by_name.get(name, {}).get("estructura", {}).get(
                    "celdas_no_vacias_muestra", 0
                )
            )
            if score > best_score:
                best_sheet, best_score = name, score
        report["hoja_usada"] = best_sheet
        if len(sheet_names) > 1:
            others = [s for s in sheet_names if s != best_sheet]
            report["avisos"].append(
                f"El archivo tiene {len(sheet_names)} hojas; se usó '{best_sheet}' "
                f"(la con más datos). Hojas no procesadas: {', '.join(others)}. "
                "Puedes elegir otra hoja desde Estandarización."
            )

    # Fase 15: keep_default_na=False — sin esto, pandas convertía el TEXTO
    # literal "None"/"nan"/"NA" escrito por el usuario en NaN al leer el Excel
    # (los vacíos REALES siguen llegando como NaN y se vacían por la máscara).
    raw = _clean_string_frame(
        book.parse(best_sheet, header=None, dtype=str, keep_default_na=False)
    )
    header_row = _detect_header_row(raw)
    if header_row > 0:
        report["filas_titulo_omitidas"] = header_row
        report["avisos"].append(
            f"Se omitieron {header_row} fila(s) de título sobre los encabezados."
        )
    headers = [str(v).strip() for v in raw.iloc[header_row].tolist()]
    # Encabezados repetidos hacen que df[col] devuelva un DataFrame. Se usa la
    # misma convención de pandas para CSV y se conservan todas las columnas.
    next_suffix: dict[str, int] = {}
    used_headers: set[str] = set()
    unique_headers: list[str] = []
    renamed = 0
    for header in headers:
        base = header or "Columna"
        candidate = base
        suffix = next_suffix.get(base, 1)
        while candidate in used_headers:
            candidate = f"{base}.{suffix}"
            suffix += 1
        next_suffix[base] = suffix
        used_headers.add(candidate)
        unique_headers.append(candidate)
        renamed += int(candidate != base)
    if renamed:
        report["avisos"].append(
            f"El archivo tiene {renamed} encabezado(s) repetidos: se renombraron "
            "(ej: 'Total.1') para conservar todas las columnas."
        )

    data_index = raw.index[header_row + 1 :]
    source_rows = [int(index) + 1 for index in data_index]
    df = raw.iloc[header_row + 1 :].reset_index(drop=True)
    df.columns = unique_headers
    df.attrs[SOURCE_ROWS_ATTR] = source_rows
    df.attrs[SOURCE_SHEET_ATTR] = best_sheet
    return df


def _scan_xlsx_formulas(
    content: bytes,
    sheet: str,
    headers: list[str],
    source_rows: list[int],
    report: dict,
    *,
    workbook=None,
    contains_formulas: bool | None = None,
) -> None:
    """Audita fórmulas en una pasada por el área real de datos seleccionada."""
    formula_report: dict = {
        "disponible": True,
        "total": 0,
        "volatiles": 0,
        "por_columna": {},
        "identificadores_volatiles": [],
    }
    report["formulas"] = formula_report
    if not source_rows or not headers:
        return

    # La enorme mayoria de las bases no contiene formulas. Revisar los tags
    # del XML evita un segundo recorrido celda por celda; cuando hay formulas,
    # conservamos abajo el analisis detallado y sus alertas.
    if contains_formulas is None:
        contains_formulas = _xlsx_contains_formulas(content)
    if not contains_formulas:
        return

    owns_workbook = workbook is None
    try:
        import openpyxl
        from .mapping import detect_columns_extended

        if workbook is None:
            workbook = openpyxl.load_workbook(
                io.BytesIO(content), data_only=False, read_only=True
            )
        worksheet = workbook[sheet]
        allowed_rows = set(source_rows)
        fixed_by_column = {header: 0 for header in headers}
        formula_by_column: dict[str, dict] = {}

        for row in worksheet.iter_rows(
            min_row=min(source_rows),
            max_row=max(source_rows),
            min_col=1,
            max_col=len(headers),
        ):
            row_number = row[0].row
            if row_number not in allowed_rows:
                continue
            for index, cell in enumerate(row):
                header = headers[index]
                if cell.data_type == "f":
                    formula = str(cell.value or "")
                    volatile = bool(_VOLATILE_FORMULA_RE.search(formula))
                    detail = formula_by_column.setdefault(
                        header, {"total": 0, "volatiles": 0, "ejemplos": []}
                    )
                    detail["total"] += 1
                    detail["volatiles"] += int(volatile)
                    if len(detail["ejemplos"]) < 5:
                        detail["ejemplos"].append(
                            {
                                "fila_origen": row_number,
                                "formula": formula[:300],
                                "volatil": volatile,
                            }
                        )
                elif cell.value not in (None, ""):
                    fixed_by_column[header] += 1

        extended = detect_columns_extended(headers)
        for header, detail in formula_by_column.items():
            detail["valores_fijos"] = fixed_by_column[header]
            formula_report["total"] += detail["total"]
            formula_report["volatiles"] += detail["volatiles"]
            match = extended.get(header)
            is_identifier = bool(match and match.grupo == "identificador")
            detail["columna_identificadora"] = is_identifier
            if detail["total"] and detail["valores_fijos"]:
                report["avisos"].append(
                    f"La columna '{header}' tiene {detail['total']} celda(s) calculadas "
                    "por fórmula mezcladas con valores fijos; sus resultados pueden "
                    "cambiar al reabrir el archivo de origen."
                )
            elif detail["total"]:
                report["avisos"].append(
                    f"La columna '{header}' contiene {detail['total']} celda(s) "
                    "calculadas por fórmula; sus resultados pueden cambiar al reabrir "
                    "el archivo de origen."
                )
            if is_identifier and detail["volatiles"]:
                formula_report["identificadores_volatiles"].append(header)
                report["avisos"].append(
                    f"Advertencia fuerte: '{header}' usa {detail['volatiles']} fórmula(s) "
                    "volátil(es) en un identificador. Ese ID no es estable y nunca se "
                    "usará como respaldo para decidir duplicados."
                )
        formula_report["por_columna"] = formula_by_column
        if owns_workbook:
            workbook.close()
    except Exception as exc:  # análisis auxiliar: jamás debe romper la carga
        formula_report.update(
            {
                "disponible": False,
                "error": type(exc).__name__,
                "por_columna": {},
            }
        )
        report["avisos"].append(
            "No se pudo auditar las fórmulas del Excel; el archivo se procesó "
            "normalmente y esta revisión quedó pendiente."
        )


def load_dataframe_with_report(
    filename: str,
    content: bytes,
    sheet: str | None = None,
    *,
    _excel_book: pd.ExcelFile | None = None,
    _formula_workbook=None,
    _contains_formulas: bool | None = None,
) -> tuple[pd.DataFrame, dict]:
    """Carga el archivo y devuelve (df, reporte_de_carga con avisos)."""
    report: dict = {
        "avisos": [],
        "hoja_usada": None,
        "hojas_disponibles": [],
        "filas_titulo_omitidas": 0,
        "filas_totales_omitidas": 0,
    }
    name = (filename or "").lower()
    if name.endswith(".xls"):
        # Fase 10 §8.1: el Excel antiguo (.xls) requiere otra librería y fallaba
        # en ejecución aunque la UI lo aceptara. Mensaje claro en vez de promesa rota.
        raise UnsupportedFileError(
            "El formato Excel antiguo (.xls) no está soportado. Abre el archivo en "
            "Excel y guárdalo como .xlsx (o expórtalo como CSV) para procesarlo."
        )
    if not name.endswith(SUPPORTED_EXTENSIONS):
        raise UnsupportedFileError(
            "Formato no soportado. Sube un archivo Excel (.xlsx) o CSV (.csv)."
        )

    if name.endswith(".csv"):
        try:
            text = content.decode("utf-8-sig")
        except UnicodeDecodeError:
            text = content.decode("latin-1", errors="replace")
        if not text.strip():
            raise UnsupportedFileError("El archivo CSV está vacío.")
        separator = _detect_separator(text)
        report["separador"] = separator
        df = pd.read_csv(
            io.StringIO(text),
            sep=separator,
            dtype=str,
            keep_default_na=False,
            skip_blank_lines=True,
        )
        df.attrs[SOURCE_ROWS_ATTR] = _csv_source_rows(text, separator, len(df))
        df.attrs[SOURCE_SHEET_ATTR] = None
    else:
        df = _load_excel(content, report, sheet=sheet, book=_excel_book)

    # Fase 12b §30: el límite era solo de FILAS — un archivo de 200.000 filas
    # × 500 columnas era "aceptado" y tumbaba pandas/openpyxl. Límite claro
    # de columnas y de celdas totales, con mensaje accionable.
    if len(df.columns) > MAX_COLUMNS:
        raise UnsupportedFileError(
            f"El archivo tiene {len(df.columns)} columnas y el máximo es "
            f"{MAX_COLUMNS}. Elimina columnas que no uses o divide la base."
        )
    if len(df) * max(len(df.columns), 1) > MAX_TOTAL_CELLS:
        raise UnsupportedFileError(
            "El archivo supera el máximo de "
            f"{MAX_TOTAL_CELLS:,} celdas (filas × columnas) para esta versión. "
            "Divide la base en archivos más pequeños.".replace(",", ".")
        )

    # Filas completamente vacías al final (frecuentes en Excel) no son datos.
    # Fase 11: vectorizado por columna (antes era fila por fila con apply).
    source_rows = list(df.attrs.get(SOURCE_ROWS_ATTR, range(2, len(df) + 2)))
    source_sheet = df.attrs.get(SOURCE_SHEET_ATTR)
    # pandas propaga attrs con deepcopy en muchas operaciones. Apartar esta
    # lista mientras filtramos evita copiar miles de numeros por cada columna;
    # se restaura completa antes de devolver el DataFrame.
    df.attrs = {}
    non_empty_mask = pd.Series(False, index=df.index)
    for col in df.columns:
        non_empty_mask |= df[col].astype(str).str.strip() != ""
    keep_positions = [position for position, keep in enumerate(non_empty_mask.tolist()) if keep]
    source_rows = [source_rows[position] for position in keep_positions]
    df = df[non_empty_mask].reset_index(drop=True)

    # Filas de totales al final ("Total", "Suma"): resumen, no datos (Fase 8).
    df = _drop_trailing_total_rows(df, report).reset_index(drop=True)
    source_rows = source_rows[: len(df)]

    if df.empty or len(df.columns) == 0:
        raise UnsupportedFileError("El archivo no tiene datos que procesar.")
    if len(df) > MAX_ROWS:
        raise UnsupportedFileError(
            f"El archivo supera el máximo de {MAX_ROWS:,} filas para esta versión.".replace(",", ".")
        )
    df = df.reset_index(drop=True)
    df.attrs[SOURCE_ROWS_ATTR] = source_rows
    df.attrs[SOURCE_SHEET_ATTR] = source_sheet
    if name.endswith(".xlsx"):
        _scan_xlsx_formulas(
            content,
            report["hoja_usada"],
            [str(column) for column in df.columns],
            source_rows,
            report,
            workbook=_formula_workbook,
            contains_formulas=_contains_formulas,
        )
    return df, report


def load_dataframes_with_reports(
    filename: str,
    content: bytes,
    sheets: list[str],
) -> tuple[dict[str, tuple[pd.DataFrame, dict]], list[str]]:
    """Carga varias hojas compartiendo una sola apertura del libro.

    Cada hoja pasa por las mismas validaciones, detecciÃ³n de encabezados y
    auditorÃ­a de fÃ³rmulas que la carga individual. Solo se evita repetir el
    parseo inmutable del archivo completo para cada hoja.
    """
    if not (filename or "").lower().endswith(".xlsx"):
        raise UnsupportedFileError("La carga multihoja requiere un archivo .xlsx.")
    _guard_xlsx_zip(content)
    book = pd.ExcelFile(io.BytesIO(content))
    available = list(book.sheet_names)
    requested = list(dict.fromkeys(sheets))
    missing = [name for name in requested if name not in available]
    if missing:
        book.close()
        raise UnsupportedFileError(
            "No existen estas hojas en el archivo: " + ", ".join(missing)
        )

    contains_formulas = _xlsx_contains_formulas(content)
    formula_workbook = None
    try:
        if requested and contains_formulas:
            import openpyxl

            formula_workbook = openpyxl.load_workbook(
                io.BytesIO(content), data_only=False, read_only=True
            )
        loaded = {
            name: load_dataframe_with_report(
                filename,
                content,
                sheet=name,
                _excel_book=book,
                _formula_workbook=formula_workbook,
                _contains_formulas=contains_formulas,
            )
            for name in requested
        }
        return loaded, available
    finally:
        if formula_workbook is not None:
            formula_workbook.close()
        book.close()


def load_dataframe(filename: str, content: bytes) -> pd.DataFrame:
    """Wrapper compatible: solo el DataFrame (el reporte se descarta)."""
    df, _ = load_dataframe_with_report(filename, content)
    return df
