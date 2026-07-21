"""Endpoints del pipeline de datos (SPEC §6 + Fase 7). Todos exigen JWT de Supabase.

Cada endpoint acepta el archivo de dos formas:
- `file` (multipart): para archivos pequeños o desarrollo local.
- `storage_path` (form): ruta dentro del bucket de Supabase Storage; la API
  descarga el archivo con la service_role key (flujo preferido en producción).

El trabajo pesado (descarga de Storage y pandas) es síncrono y corre en el
threadpool (`run_in_threadpool`): así el event loop queda libre y varios
usuarios pueden procesar archivos a la vez sin bloquearse entre ellos.

Fase 7:
- **Caché del pipeline** (§5.7): estandarizar+limpiar el mismo archivo con las
  mismas reglas se calcula UNA vez; cambiar el periodo del dashboard ya no
  re-corre todo el motor (gran ahorro de CPU en Render).
- **POST /clean/assisted**: limpieza dirigida por variables del usuario
  (Plan Analista/Gold, 2/mes + addons). La interpretación de instrucciones es
  determinista hoy; su costura IA vive en engine/directed.py.
- `mapping` (roles corregidos por el usuario, §5.10) y `scope` (columnas
  objetivo) disponibles en /clean, /clean/download y /metrics.
- Costura de refinado IA (§5.13) cableada tras la limpieza, apagada por flag.
"""

import copy
import hashlib
import io
import json
import os
import re
import threading
import time
import zipfile
from collections import OrderedDict
from uuid import UUID

from fastapi import (
    APIRouter,
    BackgroundTasks,
    Depends,
    File,
    Form,
    HTTPException,
    UploadFile,
    status,
)
from fastapi.concurrency import run_in_threadpool
from fastapi.responses import StreamingResponse

from .. import quota
from ..auth import AuthenticatedUser, get_current_user
from ..capabilities import Capability, require_capability_for_user
from ..config import Settings, get_settings
from ..engine.ai_refine import refine_with_ai
from ..engine.audit import AUDIT_COLUMNS, build_audit_dataframe
from ..engine.clean import DEFAULT_RULES, analyze_and_clean
from ..engine.directed import (
    MAX_INSTRUCTIONS_CHARS,
    interpret_cleaning_instructions,
)
from ..engine.export import safe_export_dataframe
from ..engine.loader import (
    SOURCE_ROWS_ATTR,
    UnsupportedFileError,
    load_dataframes_with_reports,
    load_dataframe_with_report,
)
from ..engine.ai_classifier import classify_columns_with_ai
from ..engine.mapping import detect_column_roles, detect_columns_extended
from ..engine.metrics import (
    CurrencyDetection,
    compute_metrics,
    detect_currency,
    is_transaction_profile,
)
from ..engine.multi_sheet import (
    build_analysis_frame,
    detect_relationships,
    is_unit_cost_column,
    relation_stats,
    validate_analysis_scope,
)
from ..engine.standardize import normalize_headers, standardize_dataframe
from ..restore_cache import (
    build_restore_snapshot,
    fetch_dataset_mapping,
    fetch_latest_cleaning_config,
    fetch_latest_restore_record,
    fetch_restore_record,
    fetch_restore_state_bundle,
    fetch_restore_state_metadata,
    reserve_restore_snapshot_revision,
    store_restore_snapshot,
    valid_restore_snapshot,
)
from ..storage import download_from_storage, normalize_user_storage_path
from ..version import ENGINE_VERSION

router = APIRouter()

MAX_UPLOAD_BYTES = 15 * 1024 * 1024  # multipart es solo para archivos pequeños
PREVIEW_ROWS = 5

# Carga y estandarización son etapas inmutables compartidas por /standardize,
# /clean, /metrics y descargas. Un presupuesto único evita duplicar memoria sin
# límite: una base grande desplaza las entradas antiguas por LRU.
_FRAME_CACHE_LOCK = threading.Lock()
_FRAME_CACHE: "OrderedDict[tuple, tuple[object, dict]]" = OrderedDict()
_FRAME_CACHE_CELL_BUDGET = 1_600_000
_FRAME_CACHE_MAX_ENTRY_CELLS = 1_200_000

# Reopening the app should not repeatedly fetch and deserialize the same
# validated snapshots. Writes invalidate this short, per-user production cache.
_RESTORE_RESPONSE_CACHE_LOCK = threading.Lock()
_RESTORE_RESPONSE_CACHE: "OrderedDict[str, tuple[float, dict]]" = OrderedDict()
_RESTORE_RESPONSE_CACHE_TTL_SECONDS = 10 * 60
_RESTORE_RESPONSE_CACHE_MAX_USERS = 8


def _restore_response_cache_get(cache_key: str) -> dict | None:
    if get_settings().app_env != "production":
        return None
    now = time.monotonic()
    with _RESTORE_RESPONSE_CACHE_LOCK:
        cached = _RESTORE_RESPONSE_CACHE.get(cache_key)
        if cached is None:
            return None
        stored_at, response = cached
        if now - stored_at > _RESTORE_RESPONSE_CACHE_TTL_SECONDS:
            _RESTORE_RESPONSE_CACHE.pop(cache_key, None)
            return None
        _RESTORE_RESPONSE_CACHE.move_to_end(cache_key)
        return copy.deepcopy(response)


def _restore_response_cache_store(cache_key: str, response: dict) -> None:
    if get_settings().app_env != "production" or response.get("dataset") is None:
        return
    with _RESTORE_RESPONSE_CACHE_LOCK:
        _RESTORE_RESPONSE_CACHE[cache_key] = (time.monotonic(), copy.deepcopy(response))
        _RESTORE_RESPONSE_CACHE.move_to_end(cache_key)
        while len(_RESTORE_RESPONSE_CACHE) > _RESTORE_RESPONSE_CACHE_MAX_USERS:
            _RESTORE_RESPONSE_CACHE.popitem(last=False)


def _restore_response_cache_invalidate(user_id: str) -> None:
    with _RESTORE_RESPONSE_CACHE_LOCK:
        prefix = f"{user_id}:"
        for cache_key in list(_RESTORE_RESPONSE_CACHE):
            if cache_key.startswith(prefix):
                _RESTORE_RESPONSE_CACHE.pop(cache_key, None)


def _restore_response_cache_key(
    user_id: str,
    record: dict,
    state: dict | None,
) -> str | None:
    if not isinstance(state, dict):
        return None
    revision = state.get("revision")
    updated_at = state.get("updated_at")
    if not isinstance(revision, int) or revision <= 0 or not isinstance(updated_at, str):
        return None
    return (
        f"{user_id}:{record['id']}:{record.get('status', '')}:"
        f"{revision}:{updated_at}"
    )


def _clone_frame(df):
    cloned = df.copy(deep=True)
    cloned.attrs = copy.deepcopy(df.attrs)
    return cloned


def _frame_cache_get(key: tuple):
    with _FRAME_CACHE_LOCK:
        cached = _FRAME_CACHE.get(key)
        if cached is None:
            return None
        _FRAME_CACHE.move_to_end(key)
        frame, report = cached
    return _clone_frame(frame), copy.deepcopy(report)


def _frame_cache_has(key: tuple) -> bool:
    with _FRAME_CACHE_LOCK:
        return key in _FRAME_CACHE


def _frame_cache_store(key: tuple, frame, report: dict) -> None:
    cells = len(frame) * max(len(frame.columns), 1)
    if cells > _FRAME_CACHE_MAX_ENTRY_CELLS:
        return
    stored = (_clone_frame(frame), copy.deepcopy(report))
    with _FRAME_CACHE_LOCK:
        _FRAME_CACHE[key] = stored
        _FRAME_CACHE.move_to_end(key)
        total = sum(
            len(item[0]) * max(len(item[0].columns), 1)
            for item in _FRAME_CACHE.values()
        )
        while len(_FRAME_CACHE) > 1 and total > _FRAME_CACHE_CELL_BUDGET:
            _, removed = _FRAME_CACHE.popitem(last=False)
            total -= len(removed[0]) * max(len(removed[0].columns), 1)


def _frame_key(kind: str, filename: str, content: bytes, sheet: str | None, extra: str = "") -> tuple:
    return (
        kind,
        hashlib.sha1(content).digest(),
        os.path.splitext(filename)[1].lower(),
        sheet or "",
        extra,
    )


def _normalize_user_storage_path(storage_path: str, user: AuthenticatedUser) -> str:
    """El bucket organiza los archivos por carpeta {user_id}/...; la API descarga
    con la service_role key (salta RLS), así que la propiedad se valida aquí."""
    return normalize_user_storage_path(storage_path, user.id)


# lib/datasets.ts antepone Date.now() al nombre para evitar colisiones en
# Storage ("1784231134931_base3_distribuidora_grande.xlsx") — es un detalle
# de almacenamiento interno, no el nombre que el usuario reconoce. El nombre
# que se muestra (y el que llega a /metrics, /reportes, descargas…) debe ser
# el original, sin ese prefijo.
_STORAGE_TIMESTAMP_PREFIX_RE = re.compile(r"^\d{10,}_")


def _display_filename(basename: str) -> str:
    stripped = _STORAGE_TIMESTAMP_PREFIX_RE.sub("", basename, count=1)
    return stripped or basename


async def _read_input(
    file: UploadFile | None,
    storage_path: str | None,
    user: AuthenticatedUser,
) -> tuple[str, bytes]:
    if file is not None:
        content = await file.read()
        if len(content) > MAX_UPLOAD_BYTES:
            raise HTTPException(
                status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
                detail="El archivo supera los 15 MB. Súbelo a Supabase Storage y envía storage_path.",
            )
        return file.filename or "archivo.csv", content
    if storage_path:
        safe_storage_path = _normalize_user_storage_path(storage_path, user)
        content = await run_in_threadpool(download_from_storage, safe_storage_path)
        return _display_filename(os.path.basename(safe_storage_path)), content
    raise HTTPException(
        status_code=422,
        detail="Envía un archivo (campo 'file') o una ruta de Storage (campo 'storage_path').",
    )


def _load_or_400(filename: str, content: bytes, sheet: str | None = None):
    key = _frame_key("raw", filename, content, sheet)
    cached = _frame_cache_get(key)
    if cached is not None:
        return cached
    try:
        frame, report = load_dataframe_with_report(filename, content, sheet=sheet)
    except UnsupportedFileError as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc))
    _frame_cache_store(key, frame, report)
    return frame, report


def _standardize_frame_cached(
    filename: str,
    content: bytes,
    sheet: str | None,
    mapping: dict | None,
    original=None,
):
    mapping_key = json.dumps(mapping or {}, sort_keys=True)
    key = _frame_key("standardized", filename, content, sheet, mapping_key)
    cached = _frame_cache_get(key)
    if cached is not None:
        return cached
    if original is None:
        original, _ = _load_or_400(filename, content, sheet=sheet)
    standardized, report = standardize_dataframe(original, mapping=mapping)
    _frame_cache_store(key, standardized, report)
    return standardized, report


def _clean_sheet_param(sheet: str | None) -> str | None:
    """Nombre de hoja elegido por el usuario (Fase 10 §8.3), saneado."""
    if not sheet:
        return None
    value = sheet.strip()
    if len(value) > 100:
        raise HTTPException(status_code=422, detail="Nombre de hoja demasiado largo.")
    return value or None


def _parse_json_field(raw: str | None, field: str) -> dict:
    if not raw:
        return {}
    if len(raw) > 20_000:
        raise HTTPException(
            status_code=422,
            detail=f"El campo '{field}' es demasiado grande.",
        )
    try:
        value = json.loads(raw)
    except json.JSONDecodeError:
        raise HTTPException(
            status_code=422,
            detail=f"El campo '{field}' debe ser JSON válido.",
        )
    if not isinstance(value, dict):
        raise HTTPException(
            status_code=422,
            detail=f"El campo '{field}' debe ser un objeto JSON.",
        )
    return value


def _validate_restore_state(raw: str | None) -> dict:
    if not raw:
        return {}
    state = _parse_json_field(raw, "restore_state")
    active_sheet = _clean_sheet_param(state.get("active_sheet"))
    available = state.get("available_sheets", [])
    excluded = state.get("excluded_sheets", [])
    if not isinstance(available, list) or not isinstance(excluded, list):
        raise HTTPException(
            status_code=422,
            detail="restore_state.available_sheets y excluded_sheets deben ser listas.",
        )

    def clean_names(values: list) -> list[str]:
        names: list[str] = []
        for value in values[:200]:
            name = _clean_sheet_param(value) if isinstance(value, str) else None
            if name and name not in names:
                names.append(name)
        return names

    available_names = clean_names(available)
    excluded_names = [name for name in clean_names(excluded) if name in available_names]
    sheet_errors = state.get("sheet_errors", {})
    if not isinstance(sheet_errors, dict) or not all(
        isinstance(key, str) and isinstance(value, str)
        for key, value in sheet_errors.items()
    ):
        raise HTTPException(status_code=422, detail="restore_state.sheet_errors debe ser un objeto.")
    selected = state.get("selected_sheets")
    selected_names = (
        [name for name in clean_names(selected) if name in available_names]
        if isinstance(selected, list)
        else [name for name in available_names if name not in excluded_names]
    )
    analysis_raw = state.get("analysis_scope")
    if analysis_raw is not None and not isinstance(analysis_raw, dict):
        raise HTTPException(
            status_code=422,
            detail="restore_state.analysis_scope debe ser un objeto o null.",
        )
    try:
        analysis_scope = (
            validate_analysis_scope(analysis_raw, available_names)
            if available_names and isinstance(analysis_raw, dict)
            else None
        )
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc))
    selection_mode = state.get("selection_mode", "all")
    if selection_mode not in {"all", "custom"}:
        raise HTTPException(status_code=422, detail="restore_state.selection_mode no es válido.")
    if analysis_scope is not None:
        analysis_scope["_selection_mode"] = selection_mode
    return {
        "active_sheet": active_sheet,
        "available_sheets": available_names,
        "excluded_sheets": excluded_names,
        "selected_sheets": selected_names,
        "sheet_errors": {
            key: value[:500] for key, value in sheet_errors.items() if key in available_names
        },
        "analysis_scope": analysis_scope,
        "selection_mode": selection_mode,
        "combine_sheets": bool(state.get("combine_sheets", False)),
    }


# ── Validación estricta de entradas (Fase 10 §14.1) ─────────────────────────
# La UI ya construye estos objetos bien, pero un cliente directo puede
# saltársela: claves desconocidas o tipos incorrectos → 422 claro.

_VALID_RULE_KEYS = {
    "fechas", "textos", "duplicados", "tipos", "nulos",
    "columnas_vacias", "fuera_de_rango",
}
_VALID_MAPPING_ROLES = {
    "fecha", "cliente", "producto", "categoria", "monto", "costo",
    "cantidad", "canal", "sucursal", "vendedor",
}


def _validate_rules(rules: dict) -> dict:
    for key, value in rules.items():
        if key not in _VALID_RULE_KEYS:
            raise HTTPException(
                status_code=422,
                detail=f"Regla desconocida: '{key}'. Válidas: {', '.join(sorted(_VALID_RULE_KEYS))}.",
            )
        if not isinstance(value, bool):
            raise HTTPException(
                status_code=422,
                detail=f"La regla '{key}' debe ser true o false.",
            )
    return rules


def _validate_mapping(mapping: dict | None) -> dict | None:
    if not mapping:
        return mapping
    if len(mapping) > 30:
        raise HTTPException(status_code=422, detail="El mapeo tiene demasiadas entradas.")
    for role, col in mapping.items():
        if str(role).strip().lower() not in _VALID_MAPPING_ROLES:
            raise HTTPException(
                status_code=422,
                detail=f"Rol de mapeo desconocido: '{role}'.",
            )
        if col is not None and not isinstance(col, str):
            raise HTTPException(
                status_code=422,
                detail=f"El mapeo del rol '{role}' debe ser el nombre de una columna (texto).",
            )
    return mapping


def _validate_scope(scope: dict | None) -> dict | None:
    if not scope:
        return scope
    for key in scope:
        if key not in {"incluir", "excluir"}:
            raise HTTPException(
                status_code=422,
                detail=f"Clave de alcance desconocida: '{key}' (usa 'incluir'/'excluir').",
            )
    for key in ("incluir", "excluir"):
        values = scope.get(key)
        if values is None:
            continue
        if not isinstance(values, list) or not all(isinstance(v, str) for v in values):
            raise HTTPException(
                status_code=422,
                detail=f"'{key}' debe ser una lista de nombres de columna.",
            )
        if len(values) > 200:
            raise HTTPException(status_code=422, detail=f"'{key}' tiene demasiadas columnas.")
    return scope


_MANIFEST_ENTRY_KEYS = {
    "nombre",
    "procesar",
    "rules",
    "mapping",
    "scope",
    "eliminar_duplicados",
    "status",
    "error",
    "revision",
}


def _parse_sheet_manifest(raw: str | None) -> dict | None:
    """Valida el contrato explícito de descarga multihoja.

    El manifiesto, no el caché del proceso, es la fuente de verdad de las hojas
    elegidas por el usuario y de la configuración aplicada a cada una.
    """
    if not raw:
        return None
    if len(raw) > 100_000:
        raise HTTPException(status_code=422, detail="El manifiesto de hojas es demasiado grande.")
    try:
        value = json.loads(raw)
    except json.JSONDecodeError:
        raise HTTPException(status_code=422, detail="El campo 'manifest' debe ser JSON válido.")
    if not isinstance(value, dict) or set(value) != {"hojas"}:
        raise HTTPException(
            status_code=422,
            detail="El manifiesto debe ser un objeto con una única lista 'hojas'.",
        )
    entries = value.get("hojas")
    if not isinstance(entries, list) or not entries or len(entries) > 50:
        raise HTTPException(
            status_code=422,
            detail="El manifiesto debe incluir entre 1 y 50 hojas.",
        )

    normalized: list[dict] = []
    seen: set[str] = set()
    for index, entry in enumerate(entries, start=1):
        if not isinstance(entry, dict):
            raise HTTPException(status_code=422, detail=f"La hoja {index} del manifiesto no es válida.")
        unknown = set(entry) - _MANIFEST_ENTRY_KEYS
        if unknown:
            raise HTTPException(
                status_code=422,
                detail=f"La hoja {index} contiene campos desconocidos: {', '.join(sorted(unknown))}.",
            )
        name = _clean_sheet_param(entry.get("nombre"))
        if not name:
            raise HTTPException(status_code=422, detail=f"La hoja {index} no tiene nombre.")
        if name in seen:
            raise HTTPException(status_code=422, detail=f"La hoja '{name}' está repetida en el manifiesto.")
        seen.add(name)
        if not isinstance(entry.get("procesar"), bool):
            raise HTTPException(
                status_code=422,
                detail=f"'procesar' debe ser true o false para la hoja '{name}'.",
            )

        rules = entry.get("rules", {})
        mapping = entry.get("mapping", {})
        scope = entry.get("scope", {})
        if not isinstance(rules, dict) or not isinstance(mapping, dict) or not isinstance(scope, dict):
            raise HTTPException(
                status_code=422,
                detail=f"rules, mapping y scope deben ser objetos en la hoja '{name}'.",
            )
        remove_duplicates = entry.get("eliminar_duplicados", False)
        if not isinstance(remove_duplicates, bool):
            raise HTTPException(
                status_code=422,
                detail=f"eliminar_duplicados debe ser booleano en la hoja '{name}'.",
            )
        revision = entry.get("revision", 0)
        if not isinstance(revision, int) or revision < 0:
            raise HTTPException(
                status_code=422,
                detail=f"revision debe ser un entero no negativo en la hoja '{name}'.",
            )
        normalized.append(
            {
                "nombre": name,
                "procesar": entry["procesar"],
                "rules": _validate_rules(rules),
                "mapping": _validate_mapping(mapping) or {},
                "scope": _validate_scope(scope) or {},
                "eliminar_duplicados": remove_duplicates,
                "status": str(entry.get("status", "pendiente"))[:30],
                "error": str(entry.get("error", ""))[:500],
                "revision": revision,
            }
        )

    if not any(entry["procesar"] for entry in normalized):
        raise HTTPException(status_code=422, detail="Selecciona al menos una hoja para procesar.")
    return {"hojas": normalized}


def _parse_analysis_scope(raw: str | None, available_sheets: list[str]) -> dict:
    value = _parse_json_field(raw, "analysis_scope") if raw else None
    try:
        scope = validate_analysis_scope(value, available_sheets)
        # `_selection_mode` pertenece solo al almacenamiento de restauracion.
        # Se acepta por compatibilidad con snapshots/clientes anteriores, pero
        # nunca debe circular por metricas, IA, exportaciones ni respuestas.
        scope.pop("_selection_mode", None)
        return scope
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc))


def _processed_manifest_frames(
    filename: str,
    content: bytes,
    manifest: dict,
    cache_dataset_id: str | None = None,
    selected_sheets: set[str] | None = None,
) -> tuple[dict[str, object], dict[str, dict[str, str]], dict[str, dict]]:
    """Procesa el manifiesto en orden y conserva la configuracion por hoja.

    Para metricas se reciben solo las hojas del alcance. Antes, cambiar de
    Enero a Mayo copiaba y validaba las diez hojas limpias aunque el tablero
    necesitara una sola; la exportacion y el detector de relaciones conservan
    el comportamiento completo al omitir ``selected_sheets``.
    """
    frames: dict[str, object] = {}
    mappings: dict[str, dict[str, str]] = {}
    results: dict[str, dict] = {}
    for entry in manifest["hojas"]:
        if not entry["procesar"] or (
            selected_sheets is not None and entry["nombre"] not in selected_sheets
        ):
            continue
        result = _analyze_cached(
            filename,
            content,
            entry["rules"],
            apply=True,
            mapping=entry["mapping"] or None,
            scope=entry["scope"] or None,
            sheet=entry["nombre"],
            eliminar_duplicados=entry["eliminar_duplicados"],
            cache_dataset_id=cache_dataset_id,
            cache_revision=entry.get("revision") or None,
        )
        frames[entry["nombre"]] = result["_df_limpio"].copy()
        mappings[entry["nombre"]] = result.get("mapeo", entry["mapping"])
        results[entry["nombre"]] = result
    return frames, mappings, results


def _validate_scope_currencies(
    analysis_scope: dict,
    mappings: dict[str, dict[str, str]],
    results: dict[str, dict],
) -> None:
    """Impide combinar indicadores monetarios de monedas incompatibles."""
    currencies: set[str] = set()
    for sheet in analysis_scope["sheets"]:
        mapping = mappings.get(sheet, {})
        if not any(mapping.get(role) for role in ("monto", "costo")):
            continue
        currency = results.get(sheet, {}).get("_moneda")
        if currency is None:
            continue
        if currency.mixta:
            raise HTTPException(
                status_code=422,
                detail="Hay monedas incompatibles dentro de una de las hojas seleccionadas.",
            )
        currencies.add(currency.dominante)
    if len(currencies) > 1:
        raise HTTPException(
            status_code=422,
            detail="Las hojas seleccionadas usan monedas incompatibles y no se pueden analizar juntas.",
        )


# ── Caché del pipeline (§5.7 + Fase 11) ──────────────────────────────────────
# Un mismo archivo con las mismas reglas/mapeo/alcance se procesa UNA vez.
#
# Fase 11: el tope ANTIGUO era por entrada (600k celdas) y excluía justamente
# a los archivos grandes — un 50.000×20 se reprocesaba completo en CADA
# módulo (Resumen, Explorar, Reportes, retomar…), que era la lentitud
# reportada. Ahora el límite es un PRESUPUESTO TOTAL de celdas: caben varias
# bases chicas o una grande, y se desalojan las más antiguas (LRU) hasta que
# la nueva quepa. Presupuesto para Render 512 MB: 2,4M celdas × ~80 B ≈ 190 MB
# en el peor caso.

_CACHE_LOCK = threading.Lock()
_CLEAN_CACHE: "OrderedDict[tuple, dict]" = OrderedDict()
_CACHE_TOTAL_CELL_BUDGET = 2_400_000
# Una entrada individual jamás puede superar el presupuesto completo
# (el loader ya limita a 200.000 filas).
_CACHE_MAX_ENTRY_CELLS = 2_400_000
_AUDIT_CACHE_LOCK = threading.Lock()
_AUDIT_CACHE: "OrderedDict[tuple, object]" = OrderedDict()
_AUDIT_CACHE_MAX_ROWS = 100_000
_EXPORT_CACHE_LOCK = threading.Lock()
_EXPORT_CACHE: "OrderedDict[tuple, tuple[bytes, str, str]]" = OrderedDict()
_EXPORT_INFLIGHT: dict[tuple, threading.Event] = {}
_EXPORT_CACHE_MAX_ENTRIES = 3


def _cache_entry_cells(result: dict, apply: bool) -> int:
    rows = result["resumen"]["filas_despues" if apply else "filas_antes"]
    cols = result["resumen"]["columnas_despues" if apply else "columnas_antes"]
    return rows * max(cols, 1)


def _cache_store(key: tuple, result: dict, apply: bool) -> None:
    cells = _cache_entry_cells(result, apply)
    if cells > _CACHE_MAX_ENTRY_CELLS:
        return
    with _CACHE_LOCK:
        _CLEAN_CACHE[key] = result
        _CLEAN_CACHE.move_to_end(key)
        # Desalojar LRU hasta caber en el presupuesto total.
        def total() -> int:
            return sum(
                _cache_entry_cells(r, bool(r["resumen"]["aplicado"]))
                for r in _CLEAN_CACHE.values()
            )
        while len(_CLEAN_CACHE) > 1 and total() > _CACHE_TOTAL_CELL_BUDGET:
            _CLEAN_CACHE.popitem(last=False)


def _cache_key(
    content: bytes,
    rules: dict | None,
    apply: bool,
    mapping: dict | None,
    scope: dict | None,
    sheet: str | None = None,
    eliminar_duplicados: bool = False,
    cache_dataset_id: str | None = None,
    cache_revision: int | None = None,
) -> tuple:
    effective_rules = {**DEFAULT_RULES, **(rules or {})}
    return (
        ENGINE_VERSION,
        cache_dataset_id or "",
        int(cache_revision or 0),
        hashlib.sha1(content).digest(),
        json.dumps(effective_rules, sort_keys=True),
        apply,
        json.dumps(mapping or {}, sort_keys=True),
        json.dumps(scope or {}, sort_keys=True),
        sheet or "",
        bool(eliminar_duplicados),
    )


def _analyze_cached(
    filename: str,
    content: bytes,
    rules: dict | None,
    apply: bool,
    mapping: dict | None = None,
    scope: dict | None = None,
    sheet: str | None = None,
    eliminar_duplicados: bool = False,
    cache_dataset_id: str | None = None,
    cache_revision: int | None = None,
    preloaded: tuple[object, dict] | None = None,
) -> dict:
    """analyze_and_clean con caché. El dict cacheado JAMÁS se muta: los
    endpoints construyen su respuesta con una copia superficial."""
    key = _cache_key(
        content,
        rules,
        apply,
        mapping,
        scope,
        sheet,
        eliminar_duplicados,
        cache_dataset_id,
        cache_revision,
    )
    with _CACHE_LOCK:
        cached = _CLEAN_CACHE.get(key)
        if cached is not None:
            _CLEAN_CACHE.move_to_end(key)
            return cached

    if preloaded is None:
        df, load_report = _load_or_400(filename, content, sheet=sheet)
    else:
        df, load_report = preloaded
        _frame_cache_store(
            _frame_key("raw", filename, content, sheet), df, load_report
        )

    # Moneda (Fase 10 §4.4): se detecta sobre los valores CRUDOS — la
    # estandarización quita los símbolos y después ya no hay evidencia.
    standardized = _standardize_frame_cached(
        filename,
        content,
        sheet,
        mapping,
        original=df,
    )
    standardized_frame, standardization_report = standardized
    effective_mapping = {
        **_validated_detected_mapping(standardized_frame, standardization_report),
        **(mapping or {}),
    }
    raw_columns = list(df.columns)
    raw_by_standardized = {
        standardized_name: raw_columns[index]
        for index, standardized_name in enumerate(standardized_frame.columns)
        if index < len(raw_columns)
    }
    monto_col = raw_by_standardized.get(effective_mapping.get("monto"))
    costo_col = raw_by_standardized.get(effective_mapping.get("costo"))
    currency = detect_currency(
        df[monto_col] if monto_col in df.columns else None,
        df[costo_col] if costo_col in df.columns else None,
    )

    result = analyze_and_clean(
        df,
        rules,
        apply,
        mapping=effective_mapping,
        scope=scope,
        eliminar_duplicados=eliminar_duplicados,
        standardized=standardized,
    )
    result["_moneda"] = currency
    result["avisos"] = list(load_report.get("avisos", [])) + list(result.get("avisos", []))
    result["carga"] = {
        "hoja_usada": load_report.get("hoja_usada"),
        "hojas_disponibles": load_report.get("hojas_disponibles", []),
        "clasificacion_hojas": load_report.get("clasificacion_hojas", []),
        "filas_titulo_omitidas": load_report.get("filas_titulo_omitidas", 0),
        "formulas": load_report.get("formulas"),
    }

    # Costura de refinado IA (§5.13): preparada, apagada por flag.
    settings = get_settings()
    if apply and settings.ai_refine_enabled and result.get("_df_limpio") is not None:
        refined, notas = refine_with_ai(result["_df_limpio"], result.get("reporte_calidad", {}))
        result["_df_limpio"] = refined
        if notas:
            result["avisos"] = result["avisos"] + [f"IA: {n}" for n in notas]

    _cache_store(key, result, apply)
    return result


def _public_clean_response(result: dict, filename: str, extra: dict | None = None) -> dict:
    """Copia sin los campos internos "_" (el dict cacheado no se toca)."""
    response = {k: v for k, v in result.items() if not k.startswith("_")}
    currency = result.get("_moneda")
    if isinstance(currency, CurrencyDetection):
        response["moneda"] = currency.dominante
        response["moneda_mixta"] = currency.mixta
        response["moneda_detalle"] = currency.to_dict()
    response["archivo"] = filename
    if extra:
        response.update(extra)
    return response


def _request_excel_recalculation(workbook) -> None:
    """Marca el libro para que Excel recalcule fórmulas conservadas al abrirlo."""
    calculation = getattr(workbook, "calculation", None)
    if calculation is None:
        return
    calculation.calcMode = "auto"
    calculation.fullCalcOnLoad = True
    calculation.forceFullCalc = True


# ── Trabajo pesado con pandas: SIEMPRE fuera del event loop ─────────────────


def _validated_detected_mapping(frame, report: dict) -> dict[str, str]:
    """Descarta roles criticos incompatibles con los valores observados."""
    detected = detect_column_roles(list(frame.columns))
    expected_types = {
        "fecha": "fecha",
        "monto": "numero",
        "costo": "numero",
        "cantidad": "numero",
    }
    for role, expected_type in expected_types.items():
        column = detected.get(role)
        if not column:
            continue
        confidence = float(report.get("column_confidence", {}).get(column, 0.0))
        if report.get("column_types", {}).get(column) != expected_type or confidence < 0.60:
            detected.pop(role, None)
    return detected


def _standardize_sync(filename: str, content: bytes, sheet: str | None = None) -> dict:
    df_original, load_report = _load_or_400(filename, content, sheet=sheet)
    df_std, report = _standardize_frame_cached(
        filename,
        content,
        sheet,
        mapping=None,
        original=df_original,
    )

    # Fase 9: mapeo universal — rol extendido (64 roles) por columna según el
    # diccionario de palabras clave, con método y confianza del match.
    extended = detect_columns_extended(list(df_std.columns))
    settings = get_settings()
    if settings.ai_classifier_enabled:
        # Costura IA (apagada por defecto): clasificar lo que el diccionario
        # no reconoció. Hoy el stub devuelve {}.
        sin_match = [c for c in df_std.columns if c not in extended]
        if sin_match:
            extended.update(classify_columns_with_ai(sin_match, df_std))

    # Vista previa antes/después con los mismos encabezados normalizados.
    before = df_original.copy()
    before.columns = df_std.columns
    detected_mapping = _validated_detected_mapping(df_std, report)
    return {
        "archivo": filename,
        "filas": len(df_std),
        "columnas": len(df_std.columns),
        "column_types": report["column_types"],
        "column_confidence": report["column_confidence"],
        "mapeo": detected_mapping,
        "mapeo_extendido": {col: match.to_dict() for col, match in extended.items()},
        "cambios": report["cambios"],
        "mojibake_auditoria": report.get("mojibake_auditoria", []),
        "avisos": list(load_report.get("avisos", [])) + list(report.get("avisos", [])),
        "carga": {
            "hoja_usada": load_report.get("hoja_usada"),
            "hojas_disponibles": load_report.get("hojas_disponibles", []),
            "clasificacion_hojas": load_report.get("clasificacion_hojas", []),
            "filas_titulo_omitidas": load_report.get("filas_titulo_omitidas", 0),
            "formulas": load_report.get("formulas"),
        },
        "preview": {
            "columnas": list(df_std.columns),
            "antes": [[str(v) for v in row] for row in before.head(PREVIEW_ROWS).itertuples(index=False, name=None)],
            "despues": [[str(v) for v in row] for row in df_std.head(PREVIEW_ROWS).itertuples(index=False, name=None)],
        },
    }


def _clean_sync(
    filename: str,
    content: bytes,
    rules: dict,
    apply: bool,
    mapping: dict | None = None,
    scope: dict | None = None,
    extra: dict | None = None,
    sheet: str | None = None,
    eliminar_duplicados: bool = False,
    cache_dataset_id: str | None = None,
    cache_revision: int | None = None,
) -> dict:
    result = _analyze_cached(
        filename,
        content,
        rules,
        apply,
        mapping=mapping,
        scope=scope,
        sheet=sheet,
        eliminar_duplicados=eliminar_duplicados,
        cache_dataset_id=cache_dataset_id,
        cache_revision=cache_revision,
    )
    return _public_clean_response(result, filename, extra)


def _extract_columns_sync(
    filename: str, content: bytes, sheet: str | None = None
) -> tuple[list[str], dict[str, str]]:
    """Columnas normalizadas + roles detectados, sin correr el motor completo.
    Se usa para interpretar las instrucciones ANTES de gastar CPU (y cupo)."""
    df, _ = _load_or_400(filename, content, sheet=sheet)
    normalize_headers(df)
    columns = list(df.columns)
    return columns, detect_column_roles(columns)


def _export_annotations(result: dict, df) -> tuple[dict, dict, list[tuple]]:
    """Marcas visuales y observaciones auditables para una hoja limpia."""
    from ..engine.standardize import (
        map_unique,
        parse_date,
        parse_number,
        physical_missing_mask,
        semantic_missing_mask,
    )

    role_labels: dict[str, str] = {
        "fecha": "fecha",
        "monto": "monto",
        "costo": "costo",
        "cantidad": "cantidad",
        "cliente": "cliente",
        "producto": "producto",
        "categoria": "categoría",
        "canal": "canal",
        "sucursal": "sucursal",
        "vendedor": "vendedor",
    }
    column_types: dict = result["column_types"]
    col_role = {col_name: role for role, col_name in result["mapeo"].items()}
    total = max(len(df), 1)
    source_rows = list(result.get("_source_rows_limpio") or range(2, len(df) + 2))
    source_sheet = (result.get("carga") or {}).get("hoja_usada")
    yellow: dict[tuple[int, str], str] = {}
    red: dict[tuple[int, str], str] = {}

    for col in df.columns:
        ctype = column_types.get(col, "texto")
        role = col_role.get(col)
        series = df[col]
        missing = (
            physical_missing_mask(series)
            | semantic_missing_mask(series, role, column_type=ctype)
        ).reset_index(drop=True)
        fill_rate = float((~missing).sum()) / total
        # Parsear por valor único y trabajar con máscaras evita repetir una
        # llamada Python por cada celda del libro grande.
        if ctype in {"fecha", "numero"}:
            parser = parse_date if ctype == "fecha" else parse_number
            parsed = map_unique(series.astype(str).reset_index(drop=True), parser)
            invalid = (~missing) & parsed.isna()
            invalid_message = (
                "Fecha no interpretable: se conservó el valor original — revisar."
                if ctype == "fecha"
                else "Número no interpretable: se conservó el valor original — revisar."
            )
            for row_idx in invalid[invalid].index:
                yellow[(int(row_idx), col)] = invalid_message

        if ctype == "fecha":
            for row_idx in missing[missing].index:
                yellow[(row_idx, col)] = "Fecha faltante: revisar."
        elif fill_rate >= 0.7:
            label = role_labels.get(role, col) if role else col
            for row_idx in missing[missing].index:
                red[(row_idx, col)] = f"Dato faltante en una columna casi completa ({label})."

        # ── Fase 18: porcentajes fuera de rango o con escala mixta ──
        # Un descuento de 1.1 (110%) o de -0.2 casi siempre es un error de
        # captura; se marca y se observa, nunca se corrige solo.
        header = str(col).casefold()
        if ctype == "numero" and any(
            token in header for token in ("pct", "porcentaje", "percent", "descuento")
        ):
            parsed_pct = map_unique(series.astype(str).reset_index(drop=True), parse_number)
            frac_scale = int(((parsed_pct >= 0) & (parsed_pct <= 1)).sum())
            over_one = (parsed_pct > 1) & (parsed_pct <= 100)
            out_mask = (parsed_pct < 0) | (parsed_pct > 100)
            if frac_scale >= int(over_one.sum()):
                # La columna trabaja en fracciones (0.15 = 15%): todo lo que
                # supere 1 equivale a más de 100%.
                out_mask = out_mask | over_one
            for row_idx in out_mask[out_mask].index[:500]:
                key = (int(row_idx), col)
                if key not in yellow and key not in red:
                    yellow[key] = (
                        "Porcentaje fuera del rango 0–100%: revisar (se conservó "
                        "el valor original)."
                    )

    # ── Fase 18: coherencia monto ≈ cantidad × precio × (1 − descuento) ──
    # Solo cuando el archivo trae todas las piezas legibles; una diferencia
    # relevante se observa como posible inconsistencia, sin corregir nada.
    roles_map = result.get("mapeo", {})
    monto_col = roles_map.get("monto")
    cantidad_col = roles_map.get("cantidad")
    precio_col = next(
        (
            str(column) for column in df.columns
            if "precio" in str(column).casefold() and str(column) != monto_col
        ),
        None,
    )
    descuento_col = next(
        (str(column) for column in df.columns if "descuento" in str(column).casefold()),
        None,
    )
    if monto_col in df.columns and cantidad_col in df.columns and precio_col:
        base = df.reset_index(drop=True)
        monto_vals = map_unique(base[monto_col].astype(str), parse_number)
        cantidad_vals = map_unique(base[cantidad_col].astype(str), parse_number)
        precio_vals = map_unique(base[precio_col].astype(str), parse_number)
        if descuento_col in (base.columns if descuento_col else []):
            descuento_vals = map_unique(base[descuento_col].astype(str), parse_number)
            # Solo descuentos en escala de fracción [0, 1]; el resto de filas
            # no se evalúa (su escala es dudosa y no queremos falsos positivos).
            descuento_ok = (descuento_vals >= 0) & (descuento_vals <= 1)
        else:
            descuento_vals = pd.Series(0.0, index=base.index)
            descuento_ok = pd.Series(True, index=base.index)
        candidatos = (
            monto_vals.notna()
            & cantidad_vals.notna()
            & precio_vals.notna()
            & descuento_vals.notna()
            & descuento_ok
            & (monto_vals >= 0)
        )
        esperado = cantidad_vals * precio_vals * (1 - descuento_vals)
        tolerancia = esperado.abs().clip(lower=50.0) * 0.02
        incoherentes = candidatos & ((monto_vals - esperado).abs() > tolerancia)
        for row_idx in incoherentes[incoherentes].index[:300]:
            key = (int(row_idx), monto_col)
            if key not in yellow and key not in red:
                yellow[key] = (
                    "Posible monto inconsistente: difiere de cantidad × precio "
                    f"× (1 − descuento) ≈ {esperado.iat[row_idx]:,.0f}".replace(",", ".")
                    + " — revisar."
                )

    observations = [
        (source_rows[row], source_sheet, col, "revisar", message)
        for (row, col), message in yellow.items()
    ] + [
        (source_rows[row], source_sheet, col, "faltante", message)
        for (row, col), message in red.items()
    ] + [
        (
            detail["fila_origen"],
            detail.get("hoja_origen"),
            "*",
            "duplicado_eliminado",
            detail["motivo"],
        )
        for detail in result.get("_filas_duplicadas_eliminadas", [])
    ]

    # ── Fase 18: cobertura de negocio adicional en Observaciones ──
    # 1. Ambigüedades numéricas y de fecha detectadas por columna. Numeric
    # summaries use the complete column (not the inference sample) and expose
    # row samples so Observaciones reconciles with Auditoria.
    numeric_ambiguities = result.get("_ambiguedades_numericas", {})
    for column, detail in numeric_ambiguities.items():
        count = int(detail.get("cantidad", 0) or 0)
        convention = str(detail.get("convencion", "decimal")).upper()
        sample_rows = ", ".join(str(row) for row in detail.get("filas_muestra", []))
        observations.append((
            "-",
            source_sheet,
            column,
            "ambiguedad_numerica_resumen",
            f"{count} valores como '1,234' se interpretaron como {convention}."
            + (f" Filas de muestra: {sample_rows}." if sample_rows else ""),
        ))
    for aviso in result.get("avisos", []):
        text = str(aviso)
        if "ambigu" not in text.casefold():
            continue
        if "1,234" in text and numeric_ambiguities:
            continue
        match = re.search(r"columna '([^']+)'", text)
        observations.append(
            ("-", source_sheet, match.group(1) if match else "*", "revisar", text[:400])
        )

    # 2. Duplicados exactos DETECTADOS pero conservados (la eliminación
    #    requiere confirmación del usuario y el archivo debe decirlo).
    removed_rows = {
        int(detail["fila_origen"])
        for detail in result.get("_filas_duplicadas_eliminadas", [])
    }
    for detail in result.get("_filas_duplicadas_detectadas", []):
        source_row = int(detail["fila_origen"])
        if source_row in removed_rows:
            continue
        observations.append((
            source_row,
            source_sheet,
            "*",
            "duplicado_conservado",
            "Fila duplicada exacta conservada: la eliminación requiere tu confirmación en Limpieza.",
        ))

    # 3. Identificadores repetidos con contenido distinto (no son duplicados
    #    eliminables; alguien debe decidir cuál registro es el correcto).
    duplicate_detail = result.get("duplicados_detalle") or {}
    conflict_examples = duplicate_detail.get("ejemplos_conflictos_id", []) or []
    for conflict in conflict_examples:
        filas = ", ".join(str(fila) for fila in (conflict.get("filas_origen") or [])[:10])
        observations.append((
            "-",
            source_sheet,
            conflict.get("columna", "*"),
            "conflicto_id",
            (
                f"El identificador '{conflict.get('identificador')}' se repite con "
                f"contenido distinto (filas {filas}): revisa cuál es el registro correcto."
            ),
        ))
    total_conflicts = int(duplicate_detail.get("conflictos_id", 0) or 0)
    if total_conflicts > len(conflict_examples):
        observations.append((
            "-",
            source_sheet,
            "*",
            "conflicto_id",
            (
                f"En total {total_conflicts} identificador(es) se repiten con contenido "
                "distinto; se listan los primeros ejemplos."
            ),
        ))
    return yellow, red, observations


def _safe_excel_sheet_name(name: str, used: set[str]) -> str:
    """Nombre válido y único sin perder silenciosamente una hoja."""
    base = re.sub(r"[\\/*?:\[\]]", "_", str(name)).strip().strip("'") or "Hoja"
    base = base[:31]
    candidate = base
    suffix = 2
    while candidate.casefold() in used:
        tail = f"_{suffix}"
        candidate = f"{base[: 31 - len(tail)]}{tail}"
        suffix += 1
    used.add(candidate.casefold())
    return candidate


def _analysis_export_type_columns(
    frame,
    results: dict[str, dict],
    mapping: dict[str, str] | None = None,
) -> tuple[set[str], set[str]]:
    """Propaga los tipos ya inferidos a una tabla apilada/relacionada.

    Una relación puede conservar columnas numéricas que no son un rol KPI
    (por ejemplo ``Precio_Unitario``) o renombrarlas con el nombre de la hoja
    derecha. Limitar el tipado a monto/costo/cantidad las convertía de nuevo
    en texto al exportar ``Datos_relacionados``.
    """
    source_kinds: dict[str, set[str]] = {}
    numeric_renames: set[str] = set()
    date_renames: set[str] = set()
    frame_columns = {str(column) for column in frame.columns}
    for sheet_name, result in results.items():
        for column, kind in result.get("column_types", {}).items():
            source = str(column)
            source_kinds.setdefault(source, set()).add(str(kind))
            # join_related_frames solo agrega este sufijo cuando una columna
            # de la tabla derecha colisiona con otra de la izquierda. La
            # coincidencia debe ser exacta: un origen numérico "ID" jamás
            # debe convertir "ID_Producto" en número por simple prefijo.
            renamed = f"{source}_{sheet_name}"
            # El sufijo solo puede provenir de una colisión del merge si la
            # columna original también sigue presente en la tabla izquierda.
            # Sin esa evidencia, ``ID_Referencia`` puede ser simplemente una
            # columna textual original y no el renombre de un ``ID`` numérico.
            if source in frame_columns and renamed in frame_columns:
                if kind == "numero":
                    numeric_renames.add(renamed)
                elif kind == "fecha":
                    date_renames.add(renamed)

    numeric = {
        source
        for source, kinds in source_kinds.items()
        if source in frame_columns and kinds == {"numero"}
    } | numeric_renames
    dates = {
        source
        for source, kinds in source_kinds.items()
        if source in frame_columns and kinds == {"fecha"}
    } | date_renames
    resolved_mapping = mapping or {}
    numeric.update(
        column
        for role, column in resolved_mapping.items()
        if role in {"monto", "costo", "cantidad"} and column in frame.columns
    )
    numeric.update(
        str(column)
        for column in frame.columns
        if str(column).casefold().startswith(("utilidad_", "margen_", "costo_"))
    )
    date_role = resolved_mapping.get("fecha")
    if date_role in frame.columns:
        dates.add(date_role)
    return numeric, dates


def _write_clean_sheet(
    wb,
    title: str,
    df,
    yellow: dict,
    red: dict,
    index: int | None = None,
    numeric_columns: set[str] | None = None,
    date_columns: set[str] | None = None,
) -> None:
    from openpyxl.styles import Font, PatternFill

    ws = wb.create_sheet(title, index)
    exported = safe_export_dataframe(
        df,
        numeric_columns=numeric_columns,
        date_columns=date_columns,
        canonical_numeric=True,
    )
    ws.append(list(exported.columns))
    # Fase 18: la hoja limpia debe seguir siendo LEGIBLE al abrirla — un
    # encabezado plano y columnas colapsadas hacían parecer que la limpieza
    # empeoró el archivo. Encabezado con color de marca, primera fila fija y
    # anchos calculados desde el contenido real.
    header_fill = PatternFill(start_color="1A3A52", end_color="1A3A52", fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
    for row in exported.itertuples(index=False, name=None):
        ws.append(list(row))
    ws.freeze_panes = "A2"
    from openpyxl.utils import get_column_letter

    sample = exported.head(200)
    for position, column in enumerate(exported.columns, start=1):
        contenido = int(sample[column].astype(str).str.len().max()) if len(sample) else 0
        width = max(len(str(column)), contenido, 8) + 2
        ws.column_dimensions[get_column_letter(position)].width = min(width, 42)

    yellow_fill = PatternFill(start_color="FFEB3B", end_color="FFEB3B", fill_type="solid")
    red_fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
    columns = list(df.columns)
    positions = {column: index + 1 for index, column in enumerate(columns)}
    for row, column in yellow:
        ws.cell(row=row + 2, column=positions[column]).fill = yellow_fill
    for row, column in red:
        ws.cell(row=row + 2, column=positions[column]).fill = red_fill
    percentage_markers = ("pct", "porcentaje", "percent")
    for column, position in positions.items():
        normalized = re.sub(r"[^a-z0-9]+", "_", str(column).casefold()).strip("_")
        if any(marker in normalized.split("_") for marker in percentage_markers) or normalized.startswith("margen_"):
            for cell in ws.iter_rows(
                min_row=2, max_row=ws.max_row, min_col=position, max_col=position
            ):
                cell[0].number_format = "0.00%"
    for column in date_columns or set():
        position = positions.get(column)
        if position is None:
            continue
        for cell in ws.iter_rows(
            min_row=2, max_row=ws.max_row, min_col=position, max_col=position
        ):
            if cell[0].data_type == "d":
                cell[0].number_format = "dd/mm/yyyy"


def _write_observations_sheet(
    wb, observations: list[tuple], title: str = "Observaciones"
) -> None:
    from openpyxl.styles import Font

    ws = wb.create_sheet(title)
    ws.append(["Fila origen", "Hoja", "Columna", "Tipo", "Detalle"])
    for cell in ws[1]:
        cell.font = Font(bold=True)
    if observations:
        for source_row, source_sheet, column, kind, message in observations:
            ws.append([source_row, source_sheet or "CSV", column, kind, message])
    else:
        ws.append([
            "—",
            "—",
            "—",
            "—",
            (
                "Sin observaciones dentro de las reglas ejecutadas. Esto no certifica "
                "completitud ni ausencia de problemas fuera de ese alcance."
            ),
        ])
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 24
    ws.column_dimensions["E"].width = 64


def _build_export_audit(
    filename: str,
    content: bytes,
    sheet: str | None,
    result: dict,
    rules: dict,
    mapping: dict | None,
    scope: dict | None,
    cache_revision: int | None = None,
    original_frame=None,
):
    audit_key = (
        ENGINE_VERSION,
        hashlib.sha256(content).digest(),
        sheet or "",
        json.dumps({**DEFAULT_RULES, **(rules or {})}, sort_keys=True),
        json.dumps(mapping or {}, sort_keys=True),
        json.dumps(scope or {}, sort_keys=True),
        int(cache_revision or 0),
        tuple(int(row) for row in result.get("_source_rows_limpio", [])),
    )
    with _AUDIT_CACHE_LOCK:
        cached_audit = _AUDIT_CACHE.get(audit_key)
        if cached_audit is not None:
            _AUDIT_CACHE.move_to_end(audit_key)
            return cached_audit.copy(deep=True)
    original = (
        _clone_frame(original_frame)
        if original_frame is not None
        else _load_or_400(filename, content, sheet=sheet)[0]
    )
    original_headers = [str(column) for column in original.columns]
    original_source_rows = list(
        original.attrs.get(SOURCE_ROWS_ATTR, range(2, len(original) + 2))
    )
    normalize_headers(original)
    confidence = {
        column: details.get("confianza_tipo")
        for column, details in result.get("reporte_calidad", {}).items()
    }
    audit = build_audit_dataframe(
        filename=filename,
        original=original,
        cleaned=result["_df_limpio"],
        original_source_rows=[int(row) for row in original_source_rows],
        cleaned_source_rows=[
            int(row) for row in result.get("_source_rows_limpio", [])
        ],
        source_sheet=(result.get("carga") or {}).get("hoja_usada"),
        column_types=result.get("column_types", {}),
        column_confidence=confidence,
        mapping=result.get("mapeo", mapping or {}),
        rules=rules,
        scope=scope,
        removed_rows=result.get("_filas_duplicadas_eliminadas", []),
        detected_duplicate_rows=result.get("_filas_duplicadas_detectadas", []),
        source_sha256=hashlib.sha256(content).hexdigest(),
        original_headers=original_headers,
        revision=cache_revision,
    )
    with _AUDIT_CACHE_LOCK:
        _AUDIT_CACHE[audit_key] = audit.copy(deep=True)
        _AUDIT_CACHE.move_to_end(audit_key)
        total_rows = sum(len(frame) for frame in _AUDIT_CACHE.values())
        while len(_AUDIT_CACHE) > 1 and total_rows > _AUDIT_CACHE_MAX_ROWS:
            _, removed = _AUDIT_CACHE.popitem(last=False)
            total_rows -= len(removed)
    return audit


def _write_audit_sheet(wb, audit, title: str = "Auditoria") -> None:
    from openpyxl.styles import Font

    ws = wb.create_sheet(title)
    exported = safe_export_dataframe(audit)
    ws.append(list(exported.columns))
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in exported.itertuples(index=False, name=None):
        ws.append(list(row))
    if audit.empty:
        ws.append([
            "—",
            "—",
            "—",
            "—",
            "—",
            "—",
            "sin_hallazgos_en_alcance",
            "sin_cambios_detectados",
            1.0,
            "no_requerida",
            ENGINE_VERSION,
            json.dumps(
                {
                    "nota": (
                        "No certifica completitud; solo indica que las reglas "
                        "ejecutadas no generaron registros de auditoria."
                    )
                },
                ensure_ascii=False,
            ),
        ])
    ws.freeze_panes = "A2"
    ws.column_dimensions["E"].width = 28
    ws.column_dimensions["F"].width = 28
    ws.column_dimensions["L"].width = 60


def _write_manifest_sheet(wb, records: list[dict], title: str = "Manifest") -> None:
    from openpyxl.styles import Font

    columns = [
        "hoja", "estado", "procesada", "filas_antes", "filas_despues",
        "reglas", "mapeo", "error", "version_motor", "source_sha256",
        "alcance_analisis", "procedencia_analisis",
    ]
    ws = wb.create_sheet(title)
    ws.append(columns)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for record in records:
        ws.append([
            record.get("hoja"),
            record.get("estado"),
            record.get("procesada"),
            record.get("filas_antes"),
            record.get("filas_despues"),
            json.dumps(record.get("reglas", {}), ensure_ascii=False, sort_keys=True),
            json.dumps(record.get("mapeo", {}), ensure_ascii=False, sort_keys=True),
            record.get("error", ""),
            ENGINE_VERSION,
            record.get("source_sha256"),
            json.dumps(record.get("analysis_scope"), ensure_ascii=False, sort_keys=True),
            json.dumps(record.get("analysis_provenance"), ensure_ascii=False, sort_keys=True),
        ])
    ws.freeze_panes = "A2"
    ws.column_dimensions["F"].width = 36
    ws.column_dimensions["G"].width = 36
    ws.column_dimensions["H"].width = 42


def _clean_download_sync(
    filename: str,
    content: bytes,
    rules: dict,
    fmt: str,
    mapping: dict | None = None,
    scope: dict | None = None,
    sheet: str | None = None,
    eliminar_duplicados: bool = False,
) -> tuple[bytes, str, str]:
    result = _analyze_cached(
        filename,
        content,
        rules,
        apply=True,
        mapping=mapping,
        scope=scope,
        sheet=sheet,
        eliminar_duplicados=eliminar_duplicados,
    )
    df = result["_df_limpio"].copy()
    stem = re.sub(r"[^\w\-]", "_", os.path.splitext(filename)[0])
    numeric_columns = {
        column for column, kind in result.get("column_types", {}).items() if kind == "numero"
    }
    date_columns = {
        column for column, kind in result.get("column_types", {}).items() if kind == "fecha"
    }
    df_export = safe_export_dataframe(
        df, numeric_columns=numeric_columns, canonical_numeric=True
    )
    audit = _build_export_audit(
        filename, content, sheet, result, rules, mapping, scope
    )

    if fmt == "csv":
        # CSV no soporta hojas: se entrega un ZIP con datos y auditoría
        # separada, ambos legibles y sin contaminar las columnas empresariales.
        output = io.BytesIO()
        with zipfile.ZipFile(output, "w", compression=zipfile.ZIP_DEFLATED) as archive:
            archive.writestr(
                f"{stem}_limpio.csv",
                df_export.to_csv(index=False, sep=";").encode("utf-8-sig"),
            )
            archive.writestr(
                f"{stem}_auditoria.csv",
                safe_export_dataframe(audit).to_csv(index=False, sep=";").encode("utf-8-sig"),
            )
            archive.writestr(
                "manifest.json",
                json.dumps(
                    {
                        "archivo_origen": filename,
                        "version_motor": ENGINE_VERSION,
                        "source_sha256": hashlib.sha256(content).hexdigest(),
                        "filas_datos": len(df_export),
                        "registros_auditoria": len(audit),
                        "nota": (
                            "La auditoria cubre las reglas ejecutadas y no certifica "
                            "completitud fuera de ese alcance."
                        ),
                    },
                    ensure_ascii=False,
                    indent=2,
                ).encode("utf-8"),
            )
        return (
            output.getvalue(),
            f"{stem}_limpio_con_auditoria.zip",
            "application/zip",
        )

    import openpyxl

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    yellow, red, observations = _export_annotations(result, df)
    _write_clean_sheet(
        wb,
        "Datos_limpios",
        df,
        yellow,
        red,
        numeric_columns=numeric_columns,
        date_columns=date_columns,
    )
    _write_observations_sheet(wb, observations)
    _write_audit_sheet(wb, audit)
    output = io.BytesIO()
    _request_excel_recalculation(wb)
    wb.save(output)
    output.seek(0)
    return (
        output.getvalue(),
        f"{stem}_limpio.xlsx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def _clean_download_book_uncached_sync(
    filename: str,
    content: bytes,
    manifest: dict,
    export_format: str,
    analysis_scope: dict | None,
    cache_dataset_id: str | None = None,
) -> tuple[bytes, str, str]:
    """Exportacion Fase 17: libro completo o ZIP multihoja auditable."""
    if not filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=422, detail="La descarga multihoja requiere un archivo .xlsx.")

    import openpyxl
    import pandas as pd

    entries = manifest["hojas"]
    processed_names = [entry["nombre"] for entry in entries if entry["procesar"]]
    missing_names = [
        name
        for name in processed_names
        if not _frame_cache_has(_frame_key("raw", filename, content, name))
    ]
    try:
        preloaded, available = load_dataframes_with_reports(
            filename, content, missing_names
        )
    except UnsupportedFileError as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc))
    declared = [entry["nombre"] for entry in entries]
    if set(declared) != set(available) or len(declared) != len(available):
        raise HTTPException(
            status_code=422,
            detail="El manifiesto debe enumerar todas las hojas del Excel exactamente una vez.",
        )

    source_hash = hashlib.sha256(content).hexdigest()
    observations: list[tuple] = []
    frames: dict[str, object] = {}
    mappings: dict[str, dict[str, str]] = {}
    results: dict[str, dict] = {}
    annotations: dict[str, tuple[dict, dict]] = {}
    audit_frames: list[object] = []
    records: list[dict] = []
    processing_errors: list[str] = []

    for entry in entries:
        name = entry["nombre"]
        base_record = {
            "hoja": name,
            "reglas": entry["rules"],
            "mapeo": entry["mapping"],
            "source_sha256": source_hash,
            "analysis_scope": analysis_scope,
            "revision": entry.get("revision", 0),
        }
        if not entry["procesar"]:
            observations.append(("-", name, "*", "hoja_no_procesada", "Conservada sin modificaciones."))
            records.append({**base_record, "estado": "no_procesada", "procesada": False})
            continue
        try:
            result = _analyze_cached(
                filename,
                content,
                entry["rules"],
                apply=True,
                mapping=entry["mapping"] or None,
                scope=entry["scope"] or None,
                sheet=name,
                eliminar_duplicados=entry["eliminar_duplicados"],
                cache_dataset_id=cache_dataset_id,
                cache_revision=entry.get("revision") or None,
                preloaded=preloaded.get(name),
            )
            frame = result["_df_limpio"].copy()
            frames[name] = frame
            mappings[name] = result.get("mapeo", entry["mapping"])
            results[name] = result
            yellow, red, sheet_observations = _export_annotations(result, frame)
            annotations[name] = (yellow, red)
            observations.extend(sheet_observations)
            audit_frames.append(
                _build_export_audit(
                    filename, content, name, result, entry["rules"],
                    entry["mapping"] or None, entry["scope"] or None,
                    entry.get("revision") or None,
                    preloaded.get(name, (None, {}))[0],
                )
            )
            records.append({
                **base_record,
                "estado": "procesada",
                "procesada": True,
                "mapeo": mappings[name],
                "filas_antes": result["resumen"]["filas_antes"],
                "filas_despues": result["resumen"]["filas_despues"],
                "error": "",
            })
        except Exception as exc:
            error = str(exc)[:500]
            processing_errors.append(f"{name}: {error}")
            observations.append(("-", name, "*", "error_procesamiento", error))
            records.append({
                **base_record,
                "estado": "error",
                "procesada": False,
                "error": error,
            })

    ambiguity_by_column: dict[str, int] = {}
    for result in results.values():
        for column, detail in result.get("_ambiguedades_numericas", {}).items():
            ambiguity_by_column[column] = ambiguity_by_column.get(column, 0) + int(
                detail.get("cantidad", 0) or 0
            )
    ambiguity_total = sum(ambiguity_by_column.values())
    if ambiguity_total:
        breakdown = ", ".join(
            f"{count} en {column}" for column, count in sorted(ambiguity_by_column.items())
        )
        observations.append((
            "-",
            "Todas las hojas procesadas",
            "*",
            "ambiguedad_numerica_total",
            f"{ambiguity_total} valores ambiguos fueron interpretados por la convención de su columna: {breakdown}.",
        ))

    if processing_errors:
        raise HTTPException(
            status_code=422,
            detail="No se exportó un libro parcial. Corrige o reintenta estas hojas: "
            + "; ".join(processing_errors),
        )

    analysis_frame = None
    analysis_mapping: dict[str, str] = {}
    provenance = None
    # En modo single las hojas procesadas ya son la salida. Crear una copia
    # adicional llamada Datos_relacionados era engañoso y duplicaba el peso.
    if analysis_scope and analysis_scope["mode"] != "single":
        _validate_scope_currencies(analysis_scope, mappings, results)
        try:
            analysis_frame, analysis_mapping, provenance = build_analysis_frame(
                frames, mappings, analysis_scope
            )
        except (KeyError, ValueError) as exc:
            raise HTTPException(status_code=422, detail=str(exc))
    for record in records:
        record["analysis_provenance"] = provenance

    analysis_numeric_columns: set[str] = set()
    analysis_date_columns: set[str] = set()
    if analysis_frame is not None:
        analysis_numeric_columns, analysis_date_columns = _analysis_export_type_columns(
            analysis_frame, results, analysis_mapping
        )

    audit_parts = [frame for frame in audit_frames if not frame.empty]
    audit = (
        pd.concat(audit_parts, ignore_index=True)
        if audit_parts
        else pd.DataFrame(columns=AUDIT_COLUMNS)
    )
    stem = re.sub(r"[^\w\-]", "_", os.path.splitext(filename)[0])
    manifest_payload = {
        "archivo_origen": filename,
        "version_motor": ENGINE_VERSION,
        "source_sha256": source_hash,
        "analysis_scope": analysis_scope,
        "analysis_provenance": provenance,
        "hojas": records,
    }

    if export_format == "csv":
        output = io.BytesIO()
        with zipfile.ZipFile(output, "w", compression=zipfile.ZIP_DEFLATED) as archive:
            for name, frame in frames.items():
                safe_name = re.sub(r"[^\w\-]", "_", name) or "Hoja"
                archive.writestr(
                    f"{safe_name}_limpio.csv",
                    safe_export_dataframe(
                        frame,
                        numeric_columns={
                            column
                            for column, kind in results[name].get("column_types", {}).items()
                            if kind == "numero"
                        },
                        canonical_numeric=True,
                    ).to_csv(index=False, sep=";").encode("utf-8-sig"),
                )
            archive.writestr(
                "Auditoria.csv",
                safe_export_dataframe(audit).to_csv(index=False, sep=";").encode("utf-8-sig"),
            )
            if analysis_frame is not None:
                analysis_name = (
                    "Datos_combinados.csv"
                    if analysis_scope["mode"] == "append"
                    else "Datos_relacionados.csv"
                )
                archive.writestr(
                    analysis_name,
                    safe_export_dataframe(
                        analysis_frame,
                        numeric_columns=analysis_numeric_columns,
                        date_columns=analysis_date_columns,
                        canonical_numeric=True,
                    ).to_csv(index=False, sep=";").encode("utf-8-sig"),
                )
            archive.writestr(
                "manifest.json",
                json.dumps(manifest_payload, ensure_ascii=False, indent=2).encode("utf-8"),
            )
        return output.getvalue(), f"{stem}_multihoja_limpio.zip", "application/zip"

    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=False)
    for name, frame in frames.items():
        original_sheet = wb[name]
        index = wb.worksheets.index(original_sheet)
        wb.remove(original_sheet)
        yellow, red = annotations[name]
        _write_clean_sheet(
            wb,
            name,
            frame,
            yellow,
            red,
            index=index,
            numeric_columns={
                column
                for column, kind in results[name].get("column_types", {}).items()
                if kind == "numero"
            },
            date_columns={
                column
                for column, kind in results[name].get("column_types", {}).items()
                if kind == "fecha"
            },
        )

    used_names = {name.casefold() for name in wb.sheetnames}
    if analysis_frame is not None:
        preferred = "Datos_combinados" if analysis_scope["mode"] == "append" else "Datos_relacionados"
        _write_clean_sheet(
            wb,
            _safe_excel_sheet_name(preferred, used_names),
            analysis_frame,
            {},
            {},
            numeric_columns=analysis_numeric_columns,
            date_columns=analysis_date_columns,
        )
    _write_observations_sheet(wb, observations, _safe_excel_sheet_name("Observaciones", used_names))
    _write_audit_sheet(wb, audit, _safe_excel_sheet_name("Auditoria", used_names))
    _write_manifest_sheet(wb, records, _safe_excel_sheet_name("Manifest", used_names))
    output = io.BytesIO()
    _request_excel_recalculation(wb)
    wb.save(output)
    return (
        output.getvalue(),
        f"{stem}_multihoja_limpio.xlsx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def _clean_download_book_sync(
    filename: str,
    content: bytes,
    manifest: dict,
    export_format: str,
    analysis_scope: dict | None,
    cache_dataset_id: str | None = None,
) -> tuple[bytes, str, str]:
    """Una sola exportación por contenido+revisión+alcance; clics repetidos reutilizan bytes."""
    key = (
        ENGINE_VERSION,
        hashlib.sha256(content).digest(),
        export_format,
        json.dumps(manifest, sort_keys=True, separators=(",", ":")),
        json.dumps(analysis_scope or {}, sort_keys=True, separators=(",", ":")),
        cache_dataset_id or "",
    )
    while True:
        with _EXPORT_CACHE_LOCK:
            cached = _EXPORT_CACHE.get(key)
            if cached is not None:
                _EXPORT_CACHE.move_to_end(key)
                return cached
            event = _EXPORT_INFLIGHT.get(key)
            if event is None:
                event = threading.Event()
                _EXPORT_INFLIGHT[key] = event
                producer = True
            else:
                producer = False
        if producer:
            break
        event.wait(timeout=120)

    try:
        exported = _clean_download_book_uncached_sync(
            filename, content, manifest, export_format, analysis_scope, cache_dataset_id
        )
        with _EXPORT_CACHE_LOCK:
            _EXPORT_CACHE[key] = exported
            _EXPORT_CACHE.move_to_end(key)
            while len(_EXPORT_CACHE) > _EXPORT_CACHE_MAX_ENTRIES:
                _EXPORT_CACHE.popitem(last=False)
        return exported
    finally:
        with _EXPORT_CACHE_LOCK:
            finished = _EXPORT_INFLIGHT.pop(key, None)
            if finished is not None:
                finished.set()


def _metrics_multi_sync(
    filename: str,
    content: bytes,
    manifest: dict,
    analysis_scope: dict,
    date_from: str | None,
    date_to: str | None,
    cache_dataset_id: str | None = None,
) -> dict:
    frames, mappings, results = _processed_manifest_frames(
        filename,
        content,
        manifest,
        cache_dataset_id,
        set(analysis_scope["sheets"]),
    )
    _validate_scope_currencies(analysis_scope, mappings, results)
    try:
        frame, mapping, provenance = build_analysis_frame(frames, mappings, analysis_scope)
    except (KeyError, ValueError) as exc:
        raise HTTPException(status_code=422, detail=str(exc))
    hint_sheet = (
        analysis_scope["join"]["left_sheet"]
        if analysis_scope["mode"] == "join"
        else analysis_scope["active_sheet"]
    )
    currency_hint = results[hint_sheet].get("_moneda")
    computed = compute_metrics(
        frame, mapping, date_from=date_from, date_to=date_to, currency_hint=currency_hint
    )
    computed["archivo"] = filename
    qualities = [results[name]["resumen"]["calidad_despues"] for name in analysis_scope["sheets"]]
    computed["calidad_datos"] = round(sum(qualities) / max(len(qualities), 1), 1)
    computed["analysis_scope"] = analysis_scope
    computed["analysis_provenance"] = provenance
    pipeline_warnings: list[str] = []
    detected_duplicates = 0
    removed_duplicates = 0
    for name in analysis_scope["sheets"]:
        sheet_result = results[name]
        detected_duplicates += int(
            sheet_result.get("problemas", {}).get("duplicados", 0) or 0
        )
        removed_duplicates += int(
            sheet_result.get("correcciones", {}).get("filas_duplicadas_eliminadas", 0)
            or 0
        )
        pipeline_warnings.extend(
            f"{name}: {warning}"
            for warning in sheet_result.get("avisos", [])
            if any(
                marker in str(warning).casefold()
                for marker in ("ambigu", "duplic", "repetici", "fecha", "porcent")
            )
        )
    computed["advertencias"] = list(dict.fromkeys([
        *computed.get("advertencias", []),
        *pipeline_warnings,
    ]))
    computed["duplicados"] = {
        "detectados": detected_duplicates,
        "eliminados": removed_duplicates,
        "conservados": max(detected_duplicates - removed_duplicates, 0),
    }
    return computed


def _relationships_sync(
    filename: str,
    content: bytes,
    manifest: dict,
    manual: dict | None = None,
    cache_dataset_id: str | None = None,
    focus: dict | None = None,
) -> dict:
    selected_sheets: set[str] | None = None
    focused_transaction_sheets: set[str] = set()
    manifest_names = {entry["nombre"] for entry in manifest["hojas"] if entry["procesar"]}
    if isinstance(manual, dict):
        requested_pair = {
            str(manual.get("left_sheet", "")),
            str(manual.get("right_sheet", "")),
        }
        if requested_pair <= manifest_names and len(requested_pair) == 2:
            selected_sheets = requested_pair
    elif isinstance(focus, dict):
        requested = focus.get("sheets")
        if isinstance(requested, list) and all(isinstance(name, str) for name in requested):
            focused_transaction_sheets = {
                name for name in requested if name in manifest_names
            }
            if focused_transaction_sheets:
                # Primero procesa solo las ventas pedidas. El reporte de carga
                # ya contiene las muestras de encabezados de todo el libro y
                # permite descubrir maestras de costo sin procesarlas dos veces.
                selected_sheets = set(focused_transaction_sheets)
    frames, mappings, results = _processed_manifest_frames(
        filename, content, manifest, cache_dataset_id, selected_sheets
    )

    if focused_transaction_sheets and results:
        first_result = next(iter(results.values()))
        profiles = first_result.get("carga", {}).get("clasificacion_hojas", [])
        profiles_by_name = {
            profile.get("nombre"): profile
            for profile in profiles
            if isinstance(profile, dict) and isinstance(profile.get("nombre"), str)
        }
        reference_sheets: set[str] = set()
        for entry in manifest["hojas"]:
            name = entry["nombre"]
            if not entry["procesar"] or name in frames:
                continue
            profile = profiles_by_name.get(name, {})
            structure = profile.get("estructura", {}) if isinstance(profile, dict) else {}
            headers = structure.get("encabezados_muestra", []) if isinstance(structure, dict) else []
            headers = [str(header) for header in headers if str(header).strip()]
            detected = detect_column_roles(headers) if headers else {}
            effective_mapping = {
                **detected,
                **(entry.get("mapping") or {}),
            }
            has_unit_cost = is_unit_cost_column(effective_mapping.get("costo"))
            reference_profile = bool(
                has_unit_cost
                and not is_transaction_profile(headers, effective_mapping)
            )
            # Compatibilidad con libros previos cuyo manifiesto no conserva
            # encabezados de muestra, pero sí una hoja llamada Productos.
            if reference_profile or "producto" in name.casefold():
                reference_sheets.add(name)
        if reference_sheets:
            extra_frames, extra_mappings, extra_results = _processed_manifest_frames(
                filename,
                content,
                manifest,
                cache_dataset_id,
                reference_sheets,
            )
            frames.update(extra_frames)
            mappings.update(extra_mappings)
            results.update(extra_results)

    def guard_currency(candidate: dict) -> dict:
        """Bloquea ventas y costos declarados en monedas incompatibles."""
        inspected = dict(candidate)
        left_name = inspected["left_sheet"]
        right_name = inspected["right_sheet"]
        left_mapping = mappings.get(left_name, {})
        right_mapping = mappings.get(right_name, {})
        left_is_transaction = is_transaction_profile(
            frames[left_name].columns, left_mapping
        )
        right_is_transaction = is_transaction_profile(
            frames[right_name].columns, right_mapping
        )
        enriches_cost = bool(
            inspected.get("purpose") == "enriquecer_costos"
            or (
                left_is_transaction
                and is_unit_cost_column(right_mapping.get("costo"))
                and not right_is_transaction
            )
        )
        inspected.setdefault(
            "purpose",
            "enriquecer_costos" if enriches_cost else "otra_relacion",
        )
        inspected.setdefault(
            "recommended",
            bool(inspected.get("safe") and enriches_cost),
        )
        inspected["currency_compatible"] = True
        if not enriches_cost:
            return inspected
        left_currency = results.get(left_name, {}).get("_moneda")
        right_currency = results.get(right_name, {}).get("_moneda")
        if not isinstance(left_currency, CurrencyDetection) or not isinstance(
            right_currency, CurrencyDetection
        ):
            return inspected
        if left_currency.mixta or right_currency.mixta:
            inspected["safe"] = False
            inspected["recommended"] = False
            inspected["currency_compatible"] = False
            inspected["reason"] = (
                "Una de las hojas contiene monedas mezcladas; no se pueden "
                "calcular costos ni utilidad con esa relacion."
            )
        elif left_currency.dominante != right_currency.dominante:
            inspected["safe"] = False
            inspected["recommended"] = False
            inspected["currency_compatible"] = False
            left_label = (
                left_currency.dominante
                if left_currency.detectadas
                else f"sin moneda declarada (se asumio {left_currency.dominante})"
            )
            right_label = (
                right_currency.dominante
                if right_currency.detectadas
                else f"sin moneda declarada (se asumio {right_currency.dominante})"
            )
            inspected["reason"] = (
                f"Las ventas de {left_name} estan en {left_label} y "
                f"los costos de {right_name} estan {right_label}; "
                "no se relacionan sin una conversion explicita."
            )
        return inspected

    candidates = [
        guard_currency(candidate)
        for candidate in detect_relationships(frames, mappings)
    ]
    candidates.sort(
        key=lambda item: (
            not item["safe"],
            not item.get("recommended", False),
            0 if item.get("purpose") == "enriquecer_costos" else 1,
            -item["overlap"],
            -item["coverage_left"],
        )
    )
    safe = [candidate for candidate in candidates if candidate["safe"]]
    manual_result = None
    if manual is not None:
        allowed = {"left_sheet", "right_sheet", "left_keys", "right_keys", "type"}
        if not isinstance(manual, dict) or set(manual) - allowed:
            raise HTTPException(status_code=422, detail="La relación manual no es válida.")
        left_name = str(manual.get("left_sheet", ""))
        right_name = str(manual.get("right_sheet", ""))
        left_keys = manual.get("left_keys")
        right_keys = manual.get("right_keys")
        if (
            left_name not in frames
            or right_name not in frames
            or left_name == right_name
            or not isinstance(left_keys, list)
            or not isinstance(right_keys, list)
            or not left_keys
            or len(left_keys) != len(right_keys)
            or len(left_keys) > 2
            or not all(isinstance(key, str) and key for key in left_keys + right_keys)
        ):
            raise HTTPException(status_code=422, detail="Las hojas o claves manuales no son válidas.")
        stats = relation_stats(frames[left_name], left_keys, frames[right_name], right_keys)
        manual_result = guard_currency({
            "left_sheet": left_name,
            "right_sheet": right_name,
            "left_keys": left_keys,
            "right_keys": right_keys,
            "type": "left",
            **stats.to_dict(),
        })
    return {
        "candidates": candidates,
        "safe_count": len(safe),
        "manual": manual_result,
        "message": None if safe else "No encontramos una conexion segura entre estas hojas. Puedes analizarlas por separado.",
    }


def _metrics_sync(
    filename: str,
    content: bytes,
    mapping: dict | None,
    date_from: str | None,
    date_to: str | None,
    sheet: str | None = None,
    eliminar_duplicados: bool = False,
    rules: dict | None = None,
    scope: dict | None = None,
    cache_dataset_id: str | None = None,
    cache_revision: int | None = None,
) -> dict:
    # Las métricas siempre se calculan sobre datos estandarizados y limpios.
    # Con el caché (§5.7), cambiar el periodo NO re-corre el pipeline completo.
    result = _analyze_cached(
        filename,
        content,
        rules=rules,
        apply=True,
        mapping=mapping,
        scope=scope,
        sheet=sheet,
        eliminar_duplicados=eliminar_duplicados,
        cache_dataset_id=cache_dataset_id,
        cache_revision=cache_revision,
    )
    df_clean = result["_df_limpio"]
    computed = compute_metrics(
        df_clean, result.get("mapeo", mapping), date_from=date_from, date_to=date_to,
        currency_hint=result.get("_moneda"),
    )
    computed["archivo"] = filename
    computed["calidad_datos"] = result["resumen"]["calidad_despues"]
    # Las decisiones conservadoras del pipeline (duplicados no confirmados,
    # fechas/números ambiguos) también deben acompañar a los KPI. Sin esto el
    # Resumen parecía definitivo aunque Limpieza sí hubiese advertido el riesgo.
    pipeline_warnings = [
        str(warning)
        for warning in result.get("avisos", [])
        if any(
            marker in str(warning).casefold()
            for marker in ("ambigu", "duplic", "repetici", "fecha", "porcent")
        )
    ]
    computed["advertencias"] = list(dict.fromkeys([
        *computed.get("advertencias", []),
        *pipeline_warnings,
    ]))
    detected_duplicates = int(result.get("problemas", {}).get("duplicados", 0) or 0)
    removed_duplicates = int(
        result.get("correcciones", {}).get("filas_duplicadas_eliminadas", 0) or 0
    )
    computed["duplicados"] = {
        "detectados": detected_duplicates,
        "eliminados": removed_duplicates,
        "conservados": max(detected_duplicates - removed_duplicates, 0),
    }
    return computed


def _restore_response(
    record: dict,
    snapshot: dict,
    source: str,
    *,
    sheet_sessions: dict[str, dict] | None = None,
    restore_state: dict | None = None,
) -> dict:
    response = {
        "dataset": {
            "id": record["id"],
            "name": record["name"],
            "source": record.get("source", "excel_csv"),
            "storage_path": record["storage_path"],
            "status": record["status"],
        },
        "standardization": snapshot["standardization"],
        "cleaning": snapshot.get("cleaning"),
        "metrics": snapshot.get("metrics"),
        "mapping": snapshot.get("mapping"),
        "eliminar_duplicados": bool(snapshot.get("eliminar_duplicados", False)),
        "source": source,
    }
    if sheet_sessions is not None:
        response["sheet_sessions"] = sheet_sessions
    if restore_state is not None:
        stored_scope = restore_state.get("analysis_scope")
        selection_mode = restore_state.get("selection_mode") or (
            stored_scope.get("_selection_mode")
            if isinstance(stored_scope, dict)
            else None
        ) or "all"
        public_scope = None
        if isinstance(stored_scope, dict):
            scope_candidate = {
                key: value
                for key, value in stored_scope.items()
                if key != "_selection_mode"
            }
            if all(
                key in scope_candidate for key in ("mode", "sheets", "active_sheet")
            ):
                try:
                    public_scope = validate_analysis_scope(
                        scope_candidate,
                        restore_state.get("available_sheets", []),
                    )
                except ValueError:
                    public_scope = None
        response.update(
            {
                "active_sheet": restore_state.get("active_sheet"),
                "available_sheets": restore_state.get("available_sheets", []),
                "excluded_sheets": restore_state.get("excluded_sheets", []),
                "selected_sheets": restore_state.get(
                    "selected_sheets",
                    [
                        name for name in restore_state.get("available_sheets", [])
                        if name not in restore_state.get("excluded_sheets", [])
                    ],
                ),
                "sheet_errors": restore_state.get("sheet_errors", {}),
                "analysis_scope": public_scope,
                "selection_mode": selection_mode,
                "combine_sheets": bool(restore_state.get("combine_sheets", False)),
            }
        )
    return response


def _build_and_store_restore_snapshot(
    dataset_id: str,
    user_id: str,
    filename: str,
    content: bytes,
    cleaning: dict | None,
    mapping: dict | None,
    sheet: str | None,
    eliminar_duplicados: bool,
    revision: int,
    restore_state: dict | None = None,
    persist: bool = True,
) -> dict:
    """Construye una hoja con resultados del servidor y la guarda atómicamente."""
    standardization = _standardize_sync(filename, content, sheet)
    effective_sheet = sheet or standardization.get("carga", {}).get("hoja_usada")
    directed = (cleaning or {}).get("dirigida") or {}
    cleaning_scope = (
        {
            "incluir": directed.get("columnas_incluir", []),
            "excluir": directed.get("columnas_excluir", []),
        }
        if directed
        else None
    )
    metrics = (
        _metrics_sync(
            filename,
            content,
            mapping,
            None,
            None,
            effective_sheet,
            eliminar_duplicados,
            rules=(cleaning or {}).get("reglas_activas"),
            scope=cleaning_scope,
            cache_dataset_id=dataset_id,
            cache_revision=revision,
        )
        if cleaning is not None
        else None
    )
    snapshot = build_restore_snapshot(
        standardization,
        cleaning,
        metrics,
        mapping,
        eliminar_duplicados,
        revision=revision,
        source_sha256=hashlib.sha256(content).hexdigest(),
        rules=(cleaning or {}).get("reglas_activas"),
        sheet=effective_sheet,
    )
    effective_state = dict(restore_state or {})
    effective_state.setdefault("active_sheet", effective_sheet)
    if not effective_state.get("available_sheets"):
        effective_state["available_sheets"] = standardization.get("carga", {}).get(
            "hojas_disponibles", []
        )
    if persist:
        stored = store_restore_snapshot(
            dataset_id, user_id, snapshot, restore_state=effective_state
        )
        if not stored:
            raise HTTPException(
                status_code=503,
                detail="El resultado se calculó, pero no pudo confirmarse su persistencia. Vuelve a intentar.",
            )
    return snapshot


def _store_standardization_restore_snapshot(
    dataset_id: str,
    user_id: str,
    content: bytes,
    standardization: dict,
    sheet: str | None,
    revision: int,
    restore_state: dict | None = None,
) -> dict:
    """Persiste una sesión de hoja aunque aún no haya limpieza/métricas."""

    effective_sheet = sheet or standardization.get("carga", {}).get("hoja_usada")
    snapshot = build_restore_snapshot(
        standardization,
        None,
        None,
        None,
        False,
        revision=revision,
        source_sha256=hashlib.sha256(content).hexdigest(),
        rules=None,
        sheet=effective_sheet,
    )
    effective_state = dict(restore_state or {})
    effective_state.setdefault("active_sheet", effective_sheet)
    if not effective_state.get("available_sheets"):
        effective_state["available_sheets"] = standardization.get("carga", {}).get(
            "hojas_disponibles", []
        )
    stored = store_restore_snapshot(
        dataset_id, user_id, snapshot, restore_state=effective_state
    )
    if not stored:
        raise HTTPException(
            status_code=503,
            detail="La hoja se estandarizó, pero no pudo confirmarse su persistencia. Vuelve a intentar.",
        )
    return snapshot


def _restore_latest_sync(user_id: str) -> dict:
    """Restore from a persistent snapshot, or rebuild once as a fallback."""
    record = fetch_latest_restore_record(user_id)
    if record is None:
        return {"dataset": None, "source": "empty"}
    production_cache = get_settings().app_env == "production"
    authoritative_state = (
        fetch_restore_state_metadata(record["id"], user_id)
        if production_cache
        else None
    )
    cache_key = _restore_response_cache_key(user_id, record, authoritative_state)
    if cache_key is not None:
        cached_response = _restore_response_cache_get(cache_key)
        if cached_response is not None:
            return cached_response

    if authoritative_state is not None:
        bundle = fetch_restore_state_bundle(
            record["id"], user_id, state=authoritative_state
        )
    elif production_cache:
        bundle = None
    else:
        bundle = fetch_restore_state_bundle(record["id"], user_id)
    if bundle is not None:
        state = bundle["state"]
        valid_by_key: dict[str, dict] = {}
        for row in bundle["sheets"]:
            if (
                row.get("source_sha256") != state.get("source_sha256")
                or row.get("engine_version") != state.get("engine_version")
                or state.get("engine_version") != ENGINE_VERSION
            ):
                continue
            raw = row.get("snapshot")
            snapshot_status = (
                "limpio" if isinstance(raw, dict) and isinstance(raw.get("cleaning"), dict)
                else "estandarizado"
            )
            valid = valid_restore_snapshot(
                raw,
                snapshot_status,
                expected_revision=row.get("revision"),
                expected_source_sha256=row.get("source_sha256"),
                expected_rules_hash=row.get("rules_hash"),
                expected_mapping_hash=row.get("mapping_hash"),
                expected_sheet=row.get("sheet"),
            )
            if valid is not None and row.get("engine_version") == valid.get("engine_version"):
                valid_by_key[str(row.get("sheet_key"))] = valid

        active_sheet = state.get("active_sheet")
        active_key = active_sheet or "__single__"
        active = valid_by_key.get(active_key)
        if active is None and valid_by_key:
            # Recuperación conservadora ante metadata global incompleta: usa
            # la hoja de mayor revisión validada, nunca un JSON sin validar.
            active = max(valid_by_key.values(), key=lambda item: item["revision"])
            active_sheet = active.get("sheet")
        if active is not None:
            sessions = {
                snapshot["sheet"]: {
                    "standardization": snapshot["standardization"],
                    "cleaning": snapshot.get("cleaning"),
                    "metrics": snapshot.get("metrics"),
                    "mapping": snapshot.get("mapping"),
                    "eliminar_duplicados": bool(
                        snapshot.get("eliminar_duplicados", False)
                    ),
                }
                for snapshot in valid_by_key.values()
                if snapshot.get("sheet") is not None
            }
            state = {**state, "active_sheet": active_sheet}
            response = _restore_response(
                record,
                active,
                "snapshot",
                sheet_sessions=sessions,
                restore_state=state,
            )
            if cache_key is not None:
                _restore_response_cache_store(cache_key, response)
            return response

    # La revisión se reserva ANTES de descargar y recalcular el archivo.
    revision = reserve_restore_snapshot_revision(record["id"], user_id)
    storage_path = normalize_user_storage_path(record["storage_path"], user_id)
    content = download_from_storage(storage_path)
    filename = _display_filename(os.path.basename(storage_path))
    mapping = fetch_dataset_mapping(record["id"])
    cleaning = None
    eliminar_duplicados = False
    if record["status"] == "limpio":
        rules, eliminar_duplicados = fetch_latest_cleaning_config(record["id"], user_id)
        cleaning = _clean_sync(
            filename,
            content,
            rules,
            True,
            mapping,
            None,
            None,
            None,
            eliminar_duplicados,
        )
    # Si falta la migración de revisiones, la restauración calculada sigue
    # funcionando pero NO se escribe con una revisión local insegura.
    safe_revision = revision or 1
    snapshot = _build_and_store_restore_snapshot(
        record["id"],
        user_id,
        filename,
        content,
        cleaning,
        mapping,
        None,
        eliminar_duplicados,
        safe_revision,
        persist=revision is not None,
    )
    response = _restore_response(record, snapshot, "computed")
    if production_cache and revision is not None:
        refreshed_state = fetch_restore_state_metadata(record["id"], user_id)
        refreshed_key = _restore_response_cache_key(user_id, record, refreshed_state)
        if refreshed_state and refreshed_state.get("revision") == revision and refreshed_key:
            _restore_response_cache_store(refreshed_key, response)
    return response


def _store_restore_state_sync(
    dataset_id: str,
    user_id: str,
    restore_state: dict,
    revision: int,
    settings: Settings,
) -> dict:
    """Persiste solo estado global reutilizando un snapshot ya validado.

    La RPC v2 exige una escritura de snapshot junto al estado. Para no
    recalcular ni degradar una limpieza existente, se valida una fila vigente,
    se conserva íntegra y solo se actualiza su revisión reservada. Si no hay
    una base segura, se rechaza: nunca se fabrica ni se usa la RPC antigua.
    """

    record = fetch_restore_record(dataset_id, user_id, settings)
    if record is None:
        raise HTTPException(status_code=404, detail="El dataset no existe.")
    if record.get("status") == "error":
        raise HTTPException(
            status_code=409,
            detail="El dataset está en estado de error y no admite restauración.",
        )
    bundle = fetch_restore_state_bundle(dataset_id, user_id, settings)
    if not isinstance(bundle, dict):
        raise HTTPException(
            status_code=409,
            detail="Aún no existe un snapshot válido para guardar este estado.",
        )
    authoritative = bundle.get("state")
    rows = bundle.get("sheets")
    if not isinstance(authoritative, dict) or not isinstance(rows, list):
        raise HTTPException(
            status_code=409,
            detail="El estado persistido del dataset está incompleto.",
        )
    if authoritative.get("dataset_id") not in (None, dataset_id) or authoritative.get(
        "user_id"
    ) not in (None, user_id):
        raise HTTPException(status_code=403, detail="El dataset no pertenece al usuario.")
    if authoritative.get("engine_version") != ENGINE_VERSION:
        raise HTTPException(
            status_code=409,
            detail="El snapshot fue generado por otra versión del motor y debe recalcularse.",
        )

    requested_available = restore_state.get("available_sheets", [])
    if not requested_available:
        raise HTTPException(
            status_code=422,
            detail="restore_state.available_sheets no puede estar vacío.",
        )
    authoritative_available = authoritative.get("available_sheets", [])
    if isinstance(authoritative_available, list) and authoritative_available:
        if set(requested_available) != set(authoritative_available):
            raise HTTPException(
                status_code=409,
                detail="Las hojas del estado no coinciden con el dataset vigente.",
            )
    active_sheet = restore_state.get("active_sheet")
    if active_sheet and active_sheet not in requested_available:
        raise HTTPException(
            status_code=422,
            detail="restore_state.active_sheet debe pertenecer a available_sheets.",
        )

    source_sha256 = authoritative.get("source_sha256")
    valid_by_key: dict[str, dict] = {}
    for row in rows:
        if not isinstance(row, dict):
            continue
        if (
            row.get("source_sha256") != source_sha256
            or row.get("engine_version") != ENGINE_VERSION
        ):
            continue
        raw = row.get("snapshot")
        snapshot_status = (
            "limpio"
            if isinstance(raw, dict) and isinstance(raw.get("cleaning"), dict)
            else "estandarizado"
        )
        valid = valid_restore_snapshot(
            raw,
            snapshot_status,
            expected_revision=row.get("revision"),
            expected_source_sha256=row.get("source_sha256"),
            expected_rules_hash=row.get("rules_hash"),
            expected_mapping_hash=row.get("mapping_hash"),
            expected_sheet=row.get("sheet"),
        )
        if valid is not None:
            valid_by_key[str(row.get("sheet_key"))] = valid
    if not valid_by_key:
        raise HTTPException(
            status_code=409,
            detail="No existe un snapshot vigente que pueda reutilizarse con seguridad.",
        )

    preferred_keys = [
        active_sheet,
        authoritative.get("active_sheet"),
        "__single__",
    ]
    base_snapshot = next(
        (
            valid_by_key[key]
            for key in preferred_keys
            if isinstance(key, str) and key in valid_by_key
        ),
        None,
    )
    if base_snapshot is None:
        base_snapshot = max(
            valid_by_key.values(), key=lambda snapshot: snapshot["revision"]
        )
    snapshot = copy.deepcopy(base_snapshot)
    snapshot["revision"] = revision
    stored = store_restore_snapshot(
        dataset_id,
        user_id,
        snapshot,
        settings=settings,
        restore_state=restore_state,
    )
    if not stored:
        raise HTTPException(
            status_code=503,
            detail=(
                "El estado no pudo confirmarse mediante la escritura atómica. "
                "Vuelve a intentar."
            ),
        )
    return {
        "status": "guardado",
        "revision": revision,
        "restore_state": restore_state,
    }


# ── Endpoints ─────────────────────────────────────────────────────────────────


@router.post("/restore/latest")
async def restore_latest(
    user: AuthenticatedUser = Depends(get_current_user),
    settings: Settings = Depends(get_settings),
) -> dict:
    """One round-trip restoration; pandas is only used when no snapshot exists."""
    # Fase 14 (P0): el fallback sin snapshot corre el pipeline completo —
    # misma puerta comercial que el dashboard. El frontend (DatasetBootstrap)
    # trata el 403 como "nada que restaurar", sin romper la navegación.
    await run_in_threadpool(
        require_capability_for_user, user.id, Capability.VIEW_DASHBOARD, settings
    )
    return await run_in_threadpool(_restore_latest_sync, user.id)


@router.post("/restore/state")
async def persist_restore_state(
    dataset_id: UUID = Form(...),
    restore_state: str = Form(...),
    user: AuthenticatedUser = Depends(get_current_user),
    settings: Settings = Depends(get_settings),
) -> dict:
    """Guarda selección/errores multihoja sin recalcular datos.

    Contrato multipart: ``dataset_id`` + ``restore_state`` JSON. La revisión
    proviene exclusivamente de PostgreSQL y la escritura usa la RPC v2
    guardada; no existe fallback directo ni degradación al snapshot legacy.
    """

    state = _validate_restore_state(restore_state)
    await run_in_threadpool(
        require_capability_for_user,
        user.id,
        Capability.VIEW_DASHBOARD,
        settings,
    )
    dataset_key = str(dataset_id)
    revision = await run_in_threadpool(
        reserve_restore_snapshot_revision,
        dataset_key,
        user.id,
        settings,
    )
    if revision is None:
        raise HTTPException(
            status_code=503,
            detail="No se pudo reservar una revisión segura para el estado.",
        )
    response = await run_in_threadpool(
        _store_restore_state_sync,
        dataset_key,
        user.id,
        state,
        revision,
        settings,
    )
    _restore_response_cache_invalidate(user.id)
    return response


@router.post("/standardize")
async def standardize(
    background_tasks: BackgroundTasks,
    file: UploadFile | None = File(None),
    storage_path: str | None = Form(None),
    dataset_id: str | None = Form(None),
    sheet: str | None = Form(None),
    restore_state: str | None = Form(None),
    user: AuthenticatedUser = Depends(get_current_user),
    settings: Settings = Depends(get_settings),
) -> dict:
    _restore_response_cache_invalidate(user.id)
    # La revisión se reserva al entrar al endpoint, antes de cualquier descarga
    # o cálculo que pueda invertir el orden de dos peticiones concurrentes.
    revision = (
        await run_in_threadpool(
            reserve_restore_snapshot_revision, dataset_id, user.id, settings
        )
        if dataset_id
        else None
    )
    # Fase 13: las cuentas nuevas nacen SIN plan — pueden navegar, pero
    # procesar archivos requiere un plan activo o la prueba gratuita vigente
    # (las cuentas existentes conservan su plan básico y no notan el cambio).
    # threadpool: la puerta consulta Supabase por HTTP y no debe bloquear el loop.
    await run_in_threadpool(
        require_capability_for_user, user.id, Capability.STANDARDIZE, settings
    )
    filename, content = await _read_input(file, storage_path, user)
    sheet_name = _clean_sheet_param(sheet)
    state = _validate_restore_state(restore_state)
    result = await run_in_threadpool(_standardize_sync, filename, content, sheet_name)
    if revision is not None:
        result["revision"] = revision
    if dataset_id and revision is not None:
        await run_in_threadpool(
            _store_standardization_restore_snapshot,
            dataset_id,
            user.id,
            content,
            result,
            sheet_name,
            revision,
            state,
        )
    return result


@router.post("/clean")
async def clean(
    background_tasks: BackgroundTasks,
    file: UploadFile | None = File(None),
    storage_path: str | None = Form(None),
    dataset_id: str | None = Form(None),
    rules: str | None = Form(None),
    apply: bool = Form(False),
    eliminar_duplicados: bool = Form(False),
    mapping: str | None = Form(None),
    sheet: str | None = Form(None),
    restore_state: str | None = Form(None),
    user: AuthenticatedUser = Depends(get_current_user),
    settings: Settings = Depends(get_settings),
) -> dict:
    if apply:
        _restore_response_cache_invalidate(user.id)
    revision = (
        await run_in_threadpool(
            reserve_restore_snapshot_revision, dataset_id, user.id, settings
        )
        if apply and dataset_id
        else None
    )
    await run_in_threadpool(
        require_capability_for_user, user.id, Capability.CLEAN, settings
    )
    filename, content = await _read_input(file, storage_path, user)
    rules_dict = _validate_rules(_parse_json_field(rules, "rules"))
    mapping_dict = _validate_mapping(_parse_json_field(mapping, "mapping") or None)
    sheet_name = _clean_sheet_param(sheet)
    state = _validate_restore_state(restore_state)
    result = await run_in_threadpool(
        _clean_sync, filename, content, rules_dict, apply, mapping_dict, None, None,
        sheet_name, eliminar_duplicados, dataset_id, revision,
    )
    if revision is not None:
        result["revision"] = revision
    if apply and dataset_id and revision is not None:
        await run_in_threadpool(
            _build_and_store_restore_snapshot,
            dataset_id,
            user.id,
            filename,
            content,
            result,
            mapping_dict,
            sheet_name,
            eliminar_duplicados,
            revision,
            state,
        )
    return result


@router.post("/clean/assisted")
async def clean_assisted(
    background_tasks: BackgroundTasks,
    file: UploadFile | None = File(None),
    storage_path: str | None = Form(None),
    dataset_id: str | None = Form(None),
    instructions: str = Form(...),
    rules: str | None = Form(None),
    eliminar_duplicados: bool = Form(False),
    mapping: str | None = Form(None),
    sheet: str | None = Form(None),
    restore_state: str | None = Form(None),
    user: AuthenticatedUser = Depends(get_current_user),
    settings: Settings = Depends(get_settings),
) -> dict:
    """Limpieza dirigida por variables del usuario (Fase 7 §3).

    Flujo: capacidad (Plan Analista/Gold) → cupo (2/mes + addons) →
    interpretar instrucciones (costura IA determinista) → correr el motor
    dirigido → registrar el consumo SOLO si corrió OK. Si las instrucciones
    no se reconocen, responde 422 y el intento NO se descuenta."""
    _restore_response_cache_invalidate(user.id)
    revision = (
        await run_in_threadpool(
            reserve_restore_snapshot_revision, dataset_id, user.id, settings
        )
        if dataset_id
        else None
    )
    await run_in_threadpool(
        require_capability_for_user, user.id, Capability.AI_CLEANING, settings
    )

    instructions = (instructions or "").strip()
    if not instructions:
        raise HTTPException(status_code=422, detail="Escribe qué variables o columnas quieres limpiar.")
    if len(instructions) > MAX_INSTRUCTIONS_CHARS:
        raise HTTPException(
            status_code=422,
            detail=f"Las instrucciones superan los {MAX_INSTRUCTIONS_CHARS} caracteres. Sé breve y específico.",
        )

    filename, content = await _read_input(file, storage_path, user)
    rules_dict = _validate_rules(_parse_json_field(rules, "rules"))
    mapping_dict = _validate_mapping(_parse_json_field(mapping, "mapping") or None)
    sheet_name = _clean_sheet_param(sheet)
    state = _validate_restore_state(restore_state)
    # 1) Cupo ANTES de gastar CPU (lanza 429 con CTA a Planes si no quedan intentos).
    quota_info = await run_in_threadpool(quota.check_cleaning_quota, user.id, settings)

    # 2) Interpretar instrucciones sobre las columnas reales del archivo.
    columns, auto_roles = await run_in_threadpool(
        _extract_columns_sync, filename, content, sheet_name
    )
    roles = {**auto_roles, **(mapping_dict or {})}
    plan = interpret_cleaning_instructions(instructions, columns, roles)
    if not plan.reconocido:
        raise HTTPException(
            status_code=422,
            detail=" ".join(plan.avisos) + " El intento NO se descontó de tu cupo.",
        )

    # 3) Correr el motor con las reglas y el alcance dirigidos. Un alcance que
    #    queda VACÍO (las exclusiones cubren todo) jamás se reinterpreta como
    #    "todas las columnas" (Fase 10 §6.3).
    effective_scope = (set(plan.columnas_incluir) or set(columns)) - set(plan.columnas_excluir)
    if not effective_scope:
        raise HTTPException(
            status_code=422,
            detail="Tus instrucciones excluyen todas las columnas: no habría nada que "
            "limpiar. Ajusta las columnas a incluir. El intento NO se descontó de tu cupo.",
        )
    merged_rules = {**rules_dict, **plan.reglas_forzadas}
    scope = {"incluir": plan.columnas_incluir, "excluir": plan.columnas_excluir}
    result = await run_in_threadpool(
        _clean_sync, filename, content, merged_rules, True, mapping_dict, scope,
        None, sheet_name, eliminar_duplicados, dataset_id, revision,
    )
    if revision is not None:
        result["revision"] = revision

    # 4) Registrar el consumo (best-effort) SOLO tras un run exitoso.
    consume_addon = bool(quota_info and quota_info.get("consume_addon"))
    await run_in_threadpool(quota.record_cleaning_usage, user.id, settings, consume_addon)

    if quota_info:
        base = quota_info["base"]
        usadas = quota_info["usadas_mes"] + 1
        addons = quota_info["addons"] - (1 if consume_addon else 0)
        cupo = {
            "disponible": True,
            "usadas_mes": usadas,
            "base": base,
            "addons": max(addons, 0),
            "restantes": max(base - usadas, 0) + max(addons, 0),
        }
    else:
        cupo = {"disponible": False, "usadas_mes": 0, "base": settings.ai_cleaning_monthly_limit, "addons": 0}

    result["dirigida"] = {"instrucciones": instructions, **plan.to_dict(), "cupo": cupo}
    if dataset_id and revision is not None:
        await run_in_threadpool(
            _build_and_store_restore_snapshot,
            dataset_id,
            user.id,
            filename,
            content,
            result,
            mapping_dict,
            sheet_name,
            eliminar_duplicados,
            revision,
            state,
        )
    return result


@router.post("/clean/download")
async def clean_download(
    file: UploadFile | None = File(None),
    storage_path: str | None = Form(None),
    dataset_id: str | None = Form(None),
    rules: str | None = Form(None),
    eliminar_duplicados: bool = Form(False),
    fmt: str = Form("xlsx"),
    format: str | None = Form(None),
    mapping: str | None = Form(None),
    scope: str | None = Form(None),
    sheet: str | None = Form(None),
    manifest: str | None = Form(None),
    combinar_hojas: bool = Form(False),
    analysis_scope: str | None = Form(None),
    user: AuthenticatedUser = Depends(get_current_user),
    settings: Settings = Depends(get_settings),
) -> StreamingResponse:
    """Devuelve el dataset con la limpieza aplicada como archivo descargable (xlsx o csv)."""
    await run_in_threadpool(
        require_capability_for_user, user.id, Capability.DOWNLOAD_CLEAN_DATASET, settings
    )
    export_format = (format or fmt).strip().lower()
    if export_format not in {"csv", "xlsx"}:
        raise HTTPException(
            status_code=422,
            detail="El campo 'fmt' debe ser 'csv' o 'xlsx'.",
        )
    filename, content = await _read_input(file, storage_path, user)
    sheet_manifest = _parse_sheet_manifest(manifest)
    if combinar_hojas and sheet_manifest is None:
        raise HTTPException(
            status_code=422,
            detail="Para combinar hojas debes enviar un manifiesto explícito.",
        )
    if sheet_manifest is not None:
        if export_format not in {"xlsx", "csv"}:
            raise HTTPException(
                status_code=422,
                detail="La descarga multihoja solo está disponible en formato XLSX.",
            )
        file_bytes, out_name, media_type = await run_in_threadpool(
            _clean_download_book_sync,
            filename,
            content,
            sheet_manifest,
            export_format,
            _parse_analysis_scope(
                analysis_scope,
                [entry["nombre"] for entry in sheet_manifest["hojas"]],
            ) if analysis_scope else (
                validate_analysis_scope(
                    {
                        "mode": "append",
                        "sheets": [entry["nombre"] for entry in sheet_manifest["hojas"] if entry["procesar"]],
                        "active_sheet": next(
                            entry["nombre"] for entry in sheet_manifest["hojas"] if entry["procesar"]
                        ),
                    },
                    [entry["nombre"] for entry in sheet_manifest["hojas"]],
                ) if combinar_hojas else None
            ),
            dataset_id,
        )
        return StreamingResponse(
            iter([file_bytes]),
            media_type=media_type,
            headers={"Content-Disposition": f'attachment; filename="{out_name}"'},
        )
    rules_dict = _validate_rules(_parse_json_field(rules, "rules"))
    mapping_dict = _validate_mapping(_parse_json_field(mapping, "mapping") or None)
    scope_dict = _validate_scope(_parse_json_field(scope, "scope") or None)
    file_bytes, out_name, media_type = await run_in_threadpool(
        _clean_download_sync, filename, content, rules_dict, export_format,
        mapping_dict, scope_dict, _clean_sheet_param(sheet), eliminar_duplicados,
    )
    return StreamingResponse(
        iter([file_bytes]),
        media_type=media_type,
        headers={"Content-Disposition": f'attachment; filename="{out_name}"'},
    )


@router.post("/metrics")
async def metrics(
    file: UploadFile | None = File(None),
    storage_path: str | None = Form(None),
    dataset_id: str | None = Form(None),
    mapping: str | None = Form(None),
    rules: str | None = Form(None),
    scope: str | None = Form(None),
    revision: int | None = Form(None),
    eliminar_duplicados: bool = Form(False),
    date_from: str | None = Form(None),
    date_to: str | None = Form(None),
    sheet: str | None = Form(None),
    manifest: str | None = Form(None),
    analysis_scope: str | None = Form(None),
    user: AuthenticatedUser = Depends(get_current_user),
    settings: Settings = Depends(get_settings),
) -> dict:
    # Fase 14 (P0): /metrics reprocesa el archivo completo (caché aparte) —
    # sin esta puerta, una cuenta sin plan tenía el dashboard gratis.
    await run_in_threadpool(
        require_capability_for_user, user.id, Capability.VIEW_DASHBOARD, settings
    )
    filename, content = await _read_input(file, storage_path, user)
    sheet_manifest = _parse_sheet_manifest(manifest)
    if sheet_manifest is not None:
        if not analysis_scope:
            raise HTTPException(status_code=422, detail="Las metricas multihoja requieren analysis_scope.")
        available = [entry["nombre"] for entry in sheet_manifest["hojas"]]
        parsed_analysis_scope = _parse_analysis_scope(analysis_scope, available)
        return await run_in_threadpool(
            _metrics_multi_sync,
            filename,
            content,
            sheet_manifest,
            parsed_analysis_scope,
            date_from,
            date_to,
            dataset_id,
        )
    mapping_dict = _validate_mapping(_parse_json_field(mapping, "mapping") or None)
    rules_dict = _validate_rules(_parse_json_field(rules, "rules"))
    scope_dict = _validate_scope(_parse_json_field(scope, "scope") or None)
    return await run_in_threadpool(
        _metrics_sync, filename, content, mapping_dict, date_from, date_to,
        _clean_sheet_param(sheet), eliminar_duplicados, rules_dict, scope_dict,
        dataset_id, revision,
    )


@router.post("/sheets/relationships")
async def sheet_relationships(
    file: UploadFile | None = File(None),
    storage_path: str | None = Form(None),
    dataset_id: str | None = Form(None),
    manifest: str = Form(...),
    relationship: str | None = Form(None),
    focus: str | None = Form(None),
    user: AuthenticatedUser = Depends(get_current_user),
    settings: Settings = Depends(get_settings),
) -> dict:
    await run_in_threadpool(
        require_capability_for_user, user.id, Capability.VIEW_DASHBOARD, settings
    )
    filename, content = await _read_input(file, storage_path, user)
    sheet_manifest = _parse_sheet_manifest(manifest)
    if sheet_manifest is None:
        raise HTTPException(status_code=422, detail="Envia un manifiesto de hojas.")
    return await run_in_threadpool(
        _relationships_sync,
        filename,
        content,
        sheet_manifest,
        _parse_json_field(relationship, "relationship") if relationship else None,
        dataset_id,
        _parse_json_field(focus, "focus") if focus else None,
    )
